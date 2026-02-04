from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from weasyprint import HTML
from django.template.loader import render_to_string
from datetime import datetime

import logging


from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpRequest, HttpResponse, JsonResponse
from django.contrib.auth import login as auth_login
from django.contrib.auth.views import LoginView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.views.generic import FormView, ListView
from django.urls import reverse, reverse_lazy
from django.views.decorators.http import require_http_methods


from myapp.models import QuizResult
from gamification.models import Badge, Achievement, DascoinTransaction
from .forms import UserRegisterForm, UserUpdateForm, ProfileUpdateForm
from .models import Profile, Role
 
# Получаем логгер для записи в журнал аудита
audit_logger = logging.getLogger('audit')




class RegisterView(LoginRequiredMixin, FormView):
    form_class = UserRegisterForm
    template_name = 'users/register.html'
    success_url = reverse_lazy('home')

    def form_valid(self, form):
        user = form.save()
        # Логирование действия
        audit_logger.info(
            'Зарегистрировался на платформе', 
            extra={
                'user': user.email if user.is_authenticated else 'Anonymous'
            }
        )
                
        form.save()
        return super().form_valid(form)



@login_required
def profile(request: HttpRequest) -> HttpResponse:
    """
    Отображает страницу профиля пользователя.

    Args:
        request (HttpRequest): Объект запроса.

    Returns:
        HttpResponse: Ответ с отрендеренным шаблоном профиля.
        Шаблон включает формы для редактирования профиля.
    """
    user = request.user
    try:
        profile = user.profile
    except Profile.DoesNotExist:
        return render(request, 'users/profile_error.html', {
            'error_message': 'Профиль пользователя не найден. Пожалуйста, обратитесь к администратору.'
        })

    if request.method == 'POST':
        user_form = UserUpdateForm(request.POST, instance=request.user)
        profile_form = ProfileUpdateForm(request.POST, request.FILES, instance=profile)
        if user_form.is_valid() and profile_form.is_valid():
            user_form.save()
            profile_form.save()
            return redirect('users:profile')
    else:
        user_form = UserUpdateForm(instance=request.user)
        profile_form = ProfileUpdateForm(instance=profile)

    # Логирование действия
    audit_logger.info(
        'Перешёл в свой профиль', 
        extra={
            'user': request.user.email if request.user.is_authenticated else 'Anonymous'
        }
    )

    return render(request, 'users/profile.html', {
        'user_form': user_form,
        'profile_form': profile_form,
        'dascoin_points': profile.dascoin_points,
        'recent_badges': profile.get_recent_badges(),
        'recent_achievements': profile.get_recent_achievements(),
        'total_badges': profile.get_available_badges_count(),
        'total_achievements': profile.get_achievements().count(),
    })


@login_required
def all_badges(request: HttpRequest) -> HttpResponse:
    """Отображает все бейджи пользователя"""
    user = request.user
    profile = user.profile
    
    user_badges = profile.get_badges()
    total_badges_received = user_badges.count()
    all_badges_count = Badge.objects.filter(is_active=True).count() # Всего активных бейджей (не используется)
    total_badges_available = profile.get_available_badges_count()
    progress_percent = int((total_badges_received / all_badges_count * 100)) if all_badges_count > 0 else 0
    
    context = {
        'user_badges': user_badges,
        'total_badges': total_badges_available,
        'progress_percent': progress_percent,
    }
    
    return render(request, 'users/includes/_all_badges.html', context)


@login_required
def all_achievements(request: HttpRequest) -> HttpResponse:
    """Отображает все достижения пользователя"""
    user = request.user
    profile = user.profile
    
    user_achievements = profile.get_achievements()
    
    context = {
        'user_achievements': user_achievements,
    }
    
    return render(request, 'users/includes/_all_achievements.html', context)





@login_required
def quiz_report(request, quiz_id):
    quiz_result = get_object_or_404(QuizResult, id=quiz_id, user=request.user)
    answers = quiz_result.answers.select_related('question', 'selected_answer').order_by('question__id').all()

    # Создаем словарь, где ключ - вопрос, значение - список выбранных ответов
    multiple_choice_answers = {}

    for answer in answers:
        if answer.question.question_type == 'multiple':
            # Если вопрос еще не в словаре, добавляем с пустым списком
            if answer.question not in multiple_choice_answers:
                multiple_choice_answers[answer.question] = []
            # Добавляем выбранный ответ (если он есть)
            if answer.selected_answer:
                multiple_choice_answers[answer.question].append(answer.selected_answer)

    context = {
        'quiz_result': quiz_result,
        'answers': answers,
        'multiple_choice_answers': multiple_choice_answers,
    }
    return render(request, 'users/includes/_quiz_report.html', context)


class CustomLoginView(LoginView):
    """
    Кастомный класс расширяющий функционал встроенной в Django функции авторизации.
    Если пользователь зарегистрирован, но администратор сайта не поставил ему в админ панеле галочку - подтверждение,
    то пользователю, при попытке входа, будет выводится сообщение "Ваш аккаунт ожидает подтверждения администратором."
    """
    template_name = "users/login.html"
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        # Заменяем label для поля username на email
        form.fields['username'].label = 'Логин'
        form.fields['username'].help_text = 'Введите ваш логин'
        return form

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated:
            return redirect('users:profile')
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        user = form.get_user()
        try:
            profile = user.profile
        except Profile.DoesNotExist:
            return render(self.request, 'users/profile_error.html', {
                'error_message': 'Профиль пользователя не найден. Пожалуйста, обратитесь к администратору.'
            })
        if not profile.is_approved:
            # Логирование действия
            audit_logger.info(
                'Хочет авторизоваться, но профиль не подтверждён', 
                extra={
                    'user': user.email if user.is_authenticated else 'Anonymous'
                }
            )
            messages.error(self.request, "Ваш аккаунт ожидает подтверждения администратором.")
            return redirect('users:login')

        # Проверяем, состоит ли пользователь в группе "Внешний пользователь"
        if user.groups.filter(name='Внешний пользователь').exists():
            audit_logger.info(
                'Вошёл в систему (Внешний пользователь)', 
                extra={
                    'user': user.email if user.is_authenticated else 'Anonymous'
                }
            )
            auth_login(self.request, user)
            
            
            # Редирект после авторизации для внешних пользователей
            return redirect('homepage')
        
        audit_logger.info(
            'Вошёл в систему', 
            extra={
                'user': user.email if user.is_authenticated else 'Anonymous'
            }
        )
        auth_login(self.request, user)
        
        
        # Проверяем, нужно ли показать модальное окно при первом входе
        if not profile.first_login_shown:
            profile.first_login_shown = True
            profile.save()
            # Добавляем флаг в сессию для показа модального окна
            self.request.session['show_intro_modal'] = True
        
        return redirect(self.get_success_url())




class TransactionsListView(LoginRequiredMixin, ListView):
    """CBV для отображения истории транзакций DASCOIN пользователя"""
    model = DascoinTransaction
    template_name = 'users/transactions.html'
    context_object_name = 'transactions'
    paginate_by = 20
    ordering = ['-created_at']
    
    def get_queryset(self):
        """Возвращает транзакции только для текущего пользователя с возможностью фильтрации"""
        queryset = DascoinTransaction.objects.filter(user=self.request.user).order_by('-created_at')
        
        # Фильтрация по типу транзакции
        transaction_type = self.request.GET.get('type')
        if transaction_type and transaction_type in ['award', 'deduct', 'set', 'correction']:
            queryset = queryset.filter(transaction_type=transaction_type)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Добавляет дополнительный контекст"""
        context = super().get_context_data(**kwargs)
        
        context['total_transactions'] = self.get_queryset().count()
        context['current_filter'] = self.request.GET.get('type', '')
        
        # Статистика по типам транзакций
        all_transactions = DascoinTransaction.objects.filter(user=self.request.user)
        context['stats'] = {
            'award': all_transactions.filter(transaction_type='award').count(),
            'deduct': all_transactions.filter(transaction_type='deduct').count(),
            'set': all_transactions.filter(transaction_type='set').count(),
            'correction': all_transactions.filter(transaction_type='correction').count(),
        }
        
        # Безопасный back_url
        referer = self.request.META.get('HTTP_REFERER', '')
        current_host = self.request.get_host()
        if referer and current_host in referer:
            context['back_url'] = referer
        else:
            context['back_url'] = reverse('users:profile')
        # Логирование действия
        audit_logger.info(
            'Смотрит историю транзакций DASCOIN', 
            extra={
                'user': self.request.user.email if self.request.user.is_authenticated else 'Anonymous'
            }
        )
        
        return context


@login_required
def export_transactions_excel(request):
    transactions = DascoinTransaction.objects.filter(user=request.user).order_by('-created_at')
    wb = Workbook()
    ws = wb.active
    ws.title = "Транзакции DASCOIN"
    headers = ['Дата', 'Тип', 'Изменение', 'До', 'После', 'Причина', 'Администратор']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    for tx in transactions:
        ws.append([
            tx.created_at.strftime("%d.%m.%Y %H:%M"),
            tx.get_transaction_type_display(),
            tx.points_change,
            tx.points_before,
            tx.points_after,
            tx.reason or "Не указана",
            tx.admin_user.get_full_name() if tx.admin_user else "Система"
        ])
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"transactions_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response


@login_required
def export_transactions_pdf(request):
    """Экспорт транзакций в PDF с помощью WeasyPrint"""
    transactions = DascoinTransaction.objects.filter(user=request.user).order_by('-created_at')
    html_string = render_to_string('users/transactions_pdf.html', {
        'transactions': transactions,
        'user': request.user,
        'generated_at': datetime.now(),
        'total_transactions': transactions.count(),
    })
    html = HTML(string=html_string)
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = f"transactions_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    audit_logger.info(
        'Экспортировал транзакции в PDF', 
        extra={
            'user': request.user.email if request.user.is_authenticated else 'Anonymous'
        }
    )
    return response


@login_required
@require_http_methods(["POST"])
def clear_intro_modal_flag(request):
    """Очищает флаг показа модального окна из сессии"""
    if 'show_intro_modal' in request.session:
        del request.session['show_intro_modal']
    return JsonResponse({'status': 'success'})




@login_required
def quiz_attempts_report(request: HttpRequest) -> HttpResponse:
    """
    Отображает отчёт по попыткам тестов и заданий пользователя.
    Если тест не содержит открытых ответов - показывает лучшую попытку.
    Если тест содержит открытые ответы - показывает проверенные попытки.
    Также отображаются все отправленные на проверку задания.
    """
    from quizzes.models import Quiz, Question, HomeworkSubmission
    
    user = request.user
    
    # Получаем все результаты тестов пользователя
    all_results = QuizResult.objects.filter(user=user).order_by('-completed_at')
    
    # Группируем результаты по названию теста
    quiz_results_dict = {}
    for result in all_results:
        quiz_title = result.quiz_title
        if quiz_title not in quiz_results_dict:
            quiz_results_dict[quiz_title] = []
        quiz_results_dict[quiz_title].append(result)
    
    # Формируем список отчётов
    report_data = []
    
    for quiz_title, results in quiz_results_dict.items():
        # Находим Quiz по названию
        quiz = Quiz.objects.filter(name=quiz_title).first()
        
        if not quiz:
            continue
        
        # Проверяем, есть ли открытые вопросы в тесте
        has_open_questions = Question.objects.filter(
            quiz=quiz,
            question_type=Question.TEXT
        ).exists()
        
        # Определяем, какой результат показывать
        if has_open_questions:
            # Если есть открытые вопросы - показываем проверенные попытки
            reviewed_results = [r for r in results if r.status == 'reviewed']
            if reviewed_results:
                # Берём самую новую проверенную попытку (первая в списке, т.к. results отсортированы по дате убывания)
                result = reviewed_results[0]
            else:
                # Если нет проверенных, пропускаем этот тест
                continue
        else:
            # Если нет открытых вопросов - показываем лучшую попытку
            result = max(results, key=lambda r: r.percent)
        
        # Определяем статус
        if has_open_questions:
            if result.status == 'reviewed':
                status = 'Пройден' if result.passed else 'Не пройден'
            elif result.status == 'pending':
                status = 'На проверке'
            else:
                status = 'Завершен'
        else:
            status = 'Пройден' if result.passed else 'Не пройден'
        
        # Получаем курс
        course_name = result.course.title if result.course else 'Не указан'
        course_slug = result.course.slug if result.course else None
        
        # Определяем тип теста
        quiz_type = 'Тест'  # По умолчанию - тест в материалах курса
        if result.course:
            # Проверяем, является ли тест финальным тестом курса
            if result.course.final_quiz and result.course.final_quiz.id == quiz.id:
                quiz_type = 'Финальный тест курса'
            else:
                # Проверяем, является ли тест финальным тестом урока
                from courses.models import Lesson
                lesson_with_quiz = Lesson.objects.filter(final_quiz=quiz).first()
                if lesson_with_quiz:
                    quiz_type = 'Финальный тест урока'
        
        # Подсчитываем правильные ответы
        # score может быть дробным из-за частичных баллов за открытые вопросы
        # Округляем до ближайшего целого для отображения
        correct_count = round(result.score)
        total_count = result.total_questions
        
        report_data.append({
            'item_type': 'quiz',
            'quiz_name': quiz_title,
            'quiz_id': quiz.id,
            'course_name': course_name,
            'course_slug': course_slug,
            'quiz_type': quiz_type,
            'correct_count': correct_count,
            'total_count': total_count,
            'percent': round(result.percent, 1),
            'status': status,
            'completed_at': result.completed_at,
        })
    
    # Добавляем отправленные на проверку задания
    homework_submissions = HomeworkSubmission.objects.filter(user=user).select_related('homework', 'course')
    
    for submission in homework_submissions:
        # Определяем статус задания
        if submission.status == 'pending':
            status = 'Ожидает проверки'
        elif submission.status == 'correct':
            status = 'Правильно'
        elif submission.status == 'incorrect':
            status = 'Не правильно'
        else:
            status = submission.get_status_display()
        
        # Получаем курс
        course_name = submission.course.title if submission.course else 'Не указан'
        course_slug = submission.course.slug if submission.course else None
        
        report_data.append({
            'item_type': 'homework',
            'quiz_name': submission.homework.name,
            'quiz_id': None,
            'homework_id': submission.homework.id,
            'submission_id': submission.id,
            'course_name': course_name,
            'course_slug': course_slug,
            'quiz_type': 'Задание',
            'correct_count': None,
            'total_count': None,
            'percent': None,
            'status': status,
            'completed_at': submission.submitted_at,
        })
    
    # Сортируем по дате завершения (новые первыми)
    report_data.sort(key=lambda x: x['completed_at'], reverse=True)
    
    context = {
        'report_data': report_data,
    }
    
    return render(request, 'users/quiz_attempts_report.html', context)


