from django.utils import timezone
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse, reverse_lazy
from django.db.models import Max
from django.views.generic import DetailView, ListView, TemplateView, View, CreateView, DeleteView
from django.contrib.auth.models import User
from django.db import transaction, models
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.decorators.http import require_POST, require_http_methods
from django.http import JsonResponse, HttpResponse, HttpRequest, HttpResponseForbidden
from django.core.cache import cache
from django.template.loader import render_to_string
from weasyprint import HTML
from datetime import datetime, timedelta
from .forms import CourseForm,  LessonForm
from .models import Course, Lesson, UserLessonTrajectory, Trajectory, UserCourseTrajectory, TrajectoryCourse, Certificate
from myapp.models import UserProgress, UserCourse, QuizResult
from myapp.views import is_admin, is_author_or_admin
import logging
from builder.models import CategoryName, Incident
from django.contrib.auth import get_user_model
from django.utils.decorators import method_decorator
from gamification.utils import award_dascoin_points, award_course_badge, award_trajectory_badge, award_first_lesson_badge
from .utils import issue_certificate, get_user_certificates
from quizzes.models import Quiz, HomeworkSubmission
from builder.views.incidents_views import _get_user_cache_version, INCIDENTS_PAGE_CACHE_TIMEOUT




logger = logging.getLogger(__name__)
audit_logger = logging.getLogger('audit')




def auto_unlock_quiz_if_lessons_completed(user, course):
    """
    Автоматически разблокирует финальный тест курса, если пользователь завершил все уроки
    после того как тест был заблокирован.
    """

    if not course.final_quiz:
        return False
    
    from quizzes.models import QuizLock
    quiz_lock = QuizLock.objects.filter(user=user, quiz=course.final_quiz).first()
    
    if quiz_lock and quiz_lock.is_locked:
        quiz_lock.is_locked = False
        quiz_lock.locked_at = None
        quiz_lock.save()
        return True
    
    return False




@method_decorator(login_required, name='dispatch')
class UserCourseTrajectoryDetailView(DetailView):
    """
    Деталка по траектории пользователя: показывает прогресс по курсам в траектории.
    """

    model = UserCourseTrajectory
    template_name = 'courses/user_course_trajectory_detail.html'
    context_object_name = 'user_trajectory'


    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user_trajectory: UserCourseTrajectory = self.object
        trajectory_courses = TrajectoryCourse.objects.filter(trajectory=user_trajectory.trajectory).order_by('order')
        user_courses = {uc.course_id: uc for uc in user_trajectory.user.started_courses.all()}
        progress = []
        all_completed = True
        next_available_course = None
        
        for tc in trajectory_courses:
            uc = user_courses.get(tc.course_id)
            if not (uc and uc.status == 'completed'):
                all_completed = False
            progress.append({
                'course': tc.course,
                'order': tc.order,
                'user_course': uc,
                'available': self._is_course_available(user_trajectory, tc, user_courses)
            })
            
            # Находим первый доступный курс
            if next_available_course is None and self._is_course_available(user_trajectory, tc, user_courses):
                if not uc or uc.status != 'completed':
                    next_available_course = tc.course
        
        # --- автоапдейт статуса и опыта ---
        if all_completed and not user_trajectory.completed:
            user_trajectory.completed = True
            user_trajectory.save(update_fields=['completed'])
            
            # Начисляем DASCOIN за завершение траектории
            award_dascoin_points(user_trajectory.user, 100, f"Завершение траектории {user_trajectory.trajectory.name}")
            
            # Выдаем бейдж за траекторию
            award_trajectory_badge(user_trajectory.user, user_trajectory.trajectory.name)
            
            # Выдаем сертификат за траекторию (если настроено)
            issue_certificate(user_trajectory.user, trajectory=user_trajectory.trajectory)
            
            # Логируем завершение траектории
            audit_logger.info(
                f'Завершил траекторию {user_trajectory.trajectory.name}', 
                extra={'user': user_trajectory.user.username}
            )
        
        context['trajectory_progress'] = progress
        context['next_available_course'] = next_available_course
        return context


    def _is_course_available(self, user_trajectory, tc, user_courses):
        """
        Курс доступен, если он первый в траектории или предыдущий завершён.
        """

        if tc.order == 1:
            return True
        prev_tc = TrajectoryCourse.objects.filter(trajectory=tc.trajectory, order=tc.order-1).first()
        if not prev_tc:
            return True
        prev_uc = user_courses.get(prev_tc.course_id)
        return prev_uc and prev_uc.status == 'completed'




class CourseDetailView(DetailView):
    """
    Деталка курса.

    """

    model = Course
    slug_url_kwarg = 'slug'
    template_name = 'courses/course_detail.html'
    context_object_name = 'course'


    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        user = request.user
        
        if user.is_authenticated:
            # Проверяем доступ через менеджер
            available_courses = Course.objects.available_for_user(user)
            if self.object not in available_courses:
                return render(request, 'courses/course_access_denied.html', {
                    'course': self.object, 
                    'reason': 'У вас нет доступа к этому курсу.'
                })
            
            # --- Блокировка доступа к курсу вне очереди траектории ---
            user_trajectories = UserCourseTrajectory.objects.filter(user=user, trajectory__courses=self.object)
            for ut in user_trajectories:
                tc = TrajectoryCourse.objects.filter(trajectory=ut.trajectory, course=self.object).first()
                if tc and tc.order > 1:
                    prev_tc = TrajectoryCourse.objects.filter(trajectory=ut.trajectory, order=tc.order-1).first()
                    if prev_tc:
                        from myapp.models import UserCourse
                        prev_uc = UserCourse.objects.filter(user=user, course=prev_tc.course).first()
                        if not prev_uc or prev_uc.status != 'completed':
                            return render(request, 'courses/course_access_denied.html', {'course': self.object, 'reason': 'Курс недоступен. Сначала завершите предыдущий курс в траектории.'})
        return self.render_to_response(self.get_context_data())


    def post(self, request, *args, **kwargs):
        """Обработка начала курса"""
        course = self.get_object()
        user = request.user
        
        if not user.is_authenticated:
            return redirect('users:login')
        
        # Получаем или создаем запись курса
        user_course, created = UserCourse.objects.get_or_create(
            user=user,
            course=course,
            defaults={'status': 'available'}
        )
        
        # Обновляем статус при нажатии кнопки
        if 'start_course' in request.POST and user_course.status == 'available':
            user_course.status = 'started'
            user_course.save()
        
        return redirect('courses:course_detail', slug=course.slug)



    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        request = self.request
        course = self.object
        user = request.user
        
        # Инициализация переменных
        user_course = None
        progress = 0
        completed_lessons = 0
        completed_quizzes = 0
        completed_homeworks = 0
        total_lessons = course.lessons.count()
        total_quizzes = course.quizzes.count()
        total_homeworks = course.homeworks.count()
        # Задания считаются как тесты для общего подсчёта материалов
        total_materials = total_lessons + total_quizzes + total_homeworks
        next_lesson = None
        all_completed = False
        completed_lessons_ids = None
        completed_quizzes_ids = None
        completed_homeworks_ids = None
        trajectory = None
        show_final_quiz = False
        show_completion_animation = False
        quiz_statuses = {}
        
        # Аудит
        audit_logger.info(
            f'Перешёл к курсу {course.title}', 
            extra={'user': user.username if user.is_authenticated else 'Anonymous'}
        )

        if user.is_authenticated:
            user_course = UserCourse.objects.filter(user=user, course=course).first()
            
            # Проверяем deadline и блокируем курс, если срок истек
            # Для курс-инцидентов блокировка не применяется (обрабатывается ниже)
            if user_course and user_course.deadline and user_course.status not in ['completed', 'blocked']:
                if timezone.now() > user_course.deadline and not course.is_incident:
                    user_course.status = 'blocked'
                    user_course.save(update_fields=['status'])

            # Если курс начат или завершен
            if user_course and user_course.status in ['started', 'completed']:
                trajectory = UserLessonTrajectory.objects.filter(user=user, course=course).first()
                
                if trajectory:
                    lessons = trajectory.lessons.all()
                    total_lessons = lessons.count()
                    lesson_ids = lessons.values_list('id', flat=True)
                    
                    completed_lessons = UserProgress.objects.filter(
                        user=user,
                        course=course,
                        completed=True,
                        lesson_id__in=lesson_ids
                    ).count()

                    completed_lessons_ids = list(
                        UserProgress.objects.filter(
                            user=user,
                            course=course,
                            completed=True,
                            lesson_id__in=lesson_ids
                        ).values_list('lesson_id', flat=True)
                    )
                    
                    # Получаем пройденные тесты в рамках этого курса
                    completed_quizzes_ids = list(
                        QuizResult.objects.filter(
                            user=user,
                            course=course,
                            quiz_title__in=[quiz.name for quiz in course.quizzes],
                            passed=True
                        ).values_list('quiz_title', flat=True)
                    )
                    completed_quizzes = len(completed_quizzes_ids)
                    
                    # Получаем выполненные задания в рамках этого курса
                    completed_homeworks_ids = list(
                        HomeworkSubmission.objects.filter(
                            user=user,
                            course=course,
                            homework__in=course.homeworks,
                            status='correct'
                        ).values_list('homework_id', flat=True).distinct()
                    )
                    completed_homeworks = len(completed_homeworks_ids)
                    
                    # Получаем статус каждого теста (для отображения иконки ожидания проверки)
                    quiz_statuses = {}
                    for quiz in course.quizzes:
                        latest_result = QuizResult.objects.filter(
                            user=user,
                            course=course,
                            quiz_title=quiz.name
                        ).order_by('-completed_at').first()
                        if latest_result:
                            quiz_statuses[quiz.name] = latest_result.status
                        else:
                            quiz_statuses[quiz.name] = None

                    max_completed_order = UserProgress.objects.filter(
                        user=user,
                        course=course,
                        completed=True,
                        lesson_id__in=lesson_ids
                    ).select_related('lesson').aggregate(max_order=Max('lesson__order'))['max_order'] or 0

                    # Специальная логика для курса "Чек-ап стоматологической клиники"
                    if course.title == "Чек-ап стоматологической клиники":
                        # Для этого курса next_lesson всегда None, так как второй урок заменяется формой метрик
                        next_lesson = None
                    else:
                        # Ищем следующий материал (урок или тест) в порядке курса
                        next_lesson = self._get_next_material(user, course, trajectory, max_completed_order)
                    
                else:
                    lessons = course.lessons.all()
                    completed_lessons = UserProgress.objects.filter(
                        user=user,
                        course=course,
                        completed=True
                    ).count()

                    completed_lessons_ids = list(
                        UserProgress.objects.filter(
                            user=user,
                            course=course,
                            completed=True
                        ).values_list('lesson_id', flat=True)
                    )
                    
                    # Получаем пройденные тесты в рамках этого курса (только уникальные по quiz_title)
                    completed_quizzes_ids = list(
                        QuizResult.objects.filter(
                            user=user,
                            course=course,
                            quiz_title__in=[quiz.name for quiz in course.quizzes],
                            passed=True
                        ).values_list('quiz_title', flat=True).distinct()
                    )
                    completed_quizzes = len(completed_quizzes_ids)
                    
                    # Получаем выполненные задания в рамках этого курса
                    completed_homeworks_ids = list(
                        HomeworkSubmission.objects.filter(
                            user=user,
                            course=course,
                            homework__in=course.homeworks,
                            status='correct'
                        ).values_list('homework_id', flat=True).distinct()
                    )
                    completed_homeworks = len(completed_homeworks_ids)
                    
                    # Получаем статус каждого теста (для отображения иконки ожидания проверки)
                    quiz_statuses = {}
                    for quiz in course.quizzes:
                        latest_result = QuizResult.objects.filter(
                            user=user,
                            course=course,
                            quiz_title=quiz.name
                        ).order_by('-completed_at').first()
                        if latest_result:
                            quiz_statuses[quiz.name] = latest_result.status
                        else:
                            quiz_statuses[quiz.name] = None

                    max_completed_order = UserProgress.objects.filter(
                        user=user,
                        course=course,
                        completed=True
                    ).select_related('lesson').aggregate(max_order=Max('lesson__order'))['max_order'] or 0

                    # Специальная логика для курса "Чек-ап стоматологической клиники"
                    if course.title == "Чек-ап стоматологической клиники":
                        # Для этого курса next_lesson всегда None, так как второй урок заменяется формой метрик
                        next_lesson = None
                    else:
                        # Ищем следующий материал (урок или тест) в порядке курса
                        next_lesson = self._get_next_material(user, course, None, max_completed_order)
                
                # Вычисляем прогресс с учетом уроков, тестов и заданий
                completed_materials = completed_lessons + completed_quizzes + completed_homeworks
                progress = int((completed_materials / total_materials) * 100) if total_materials > 0 else 0
                all_completed = (completed_lessons >= total_lessons and 
                                completed_quizzes >= total_quizzes and 
                                completed_homeworks >= total_homeworks)

                # Автоматическое завершение курса при выполнении условий
                if user_course.status == 'started' and all_completed:
                    # Если есть финальный тест, проверяем его прохождение в рамках этого курса
                    if course.final_quiz:
                        quiz_passed = QuizResult.objects.filter(
                            user=user,
                            course=course,
                            quiz_title=course.final_quiz.name,
                            passed=True
                        ).exists()
                        if quiz_passed:
                            # Проверяем, был ли курс уже завершен ранее
                            was_completed_before = user_course.status == 'completed'
                            if not was_completed_before:
                                user_course.status = 'completed'
                                user_course.save()
                                
                                # Начисляем очки только если курс только что завершен И курс не является инцидентом
                                if not course.is_incident:
                                    award_dascoin_points(user, course.points, f"Завершение курса {course.title}")
                                    award_course_badge(user, course)
                                    # Выдаем сертификат за курс (если настроено)
                                    issue_certificate(user, course=course)
                            else:
                                # Курс уже был завершен, просто обновляем статус
                                user_course.status = 'completed'
                                user_course.save()
                        else:
                            # Тест не пройден, но все уроки завершены - автоматически разблокируем тест
                            auto_unlock_quiz_if_lessons_completed(user, course)
                    else:
                        # Если теста нет - завершаем автоматически
                        # Проверяем, был ли курс уже завершен ранее
                        was_completed_before = user_course.status == 'completed'
                        if not was_completed_before:
                            user_course.status = 'completed'
                            user_course.save()
                            
                            # Начисляем очки только если курс только что завершен И курс не является инцидентом
                            if not course.is_incident:
                                award_dascoin_points(user, course.points, f"Завершение курса {course.title}")
                                award_course_badge(user, course)
                        else:
                            # Курс уже был завершен, просто обновляем статус
                            user_course.status = 'completed'
                            user_course.save()

                # Проверка для отображения финального теста
                final_quiz_status = None
                final_quiz_passed = False
                if course.final_quiz:
                    final_quiz_passed = QuizResult.objects.filter(
                        user=user,
                        course=course,
                        quiz_title=course.final_quiz.name,
                        passed=True
                    ).exists()
                    
                    # Если курс помечен как завершенный, но финальный тест не пройден - сбрасываем статус
                    if user_course.status == 'completed' and not final_quiz_passed:
                        user_course.status = 'started'
                        user_course.save(update_fields=['status'])
                        # Перезагружаем объект из базы данных для обновления в памяти
                        user_course.refresh_from_db()
                    if final_quiz_passed:
                        show_final_quiz = True
                    
                    # Получаем статус финального теста (pending/reviewed/completed)
                    latest_final_quiz_result = QuizResult.objects.filter(
                        user=user,
                        course=course,
                        quiz_title=course.final_quiz.name
                    ).order_by('-completed_at').first()
                    
                    if latest_final_quiz_result:
                        final_quiz_status = latest_final_quiz_result.status
                    
                    # Получаем информацию о попытках для финального теста в рамках этого курса (исключаем те, что помечены как исключенные из лимита)
                    failed_attempts = QuizResult.objects.filter(
                        user=user,
                        course=course,
                        quiz_title=course.final_quiz.name,
                        passed=False,
                        excluded_from_limit=False
                    ).count()
                    
                    # Проверяем реальное состояние блокировки теста
                    from quizzes.models import QuizLock
                    quiz_lock = QuizLock.objects.filter(user=user, quiz=course.final_quiz).first()
                    is_actually_locked = quiz_lock.is_locked if quiz_lock else False
                    
                    quiz_attempts_info = {
                        'failed_attempts': failed_attempts,
                        'attempt_limit': course.final_quiz.attempt_limit,
                        'attempts_left': course.final_quiz.attempt_limit - failed_attempts if course.final_quiz.attempt_limit > 0 else None,
                        'is_locked': is_actually_locked
                    }
                elif all_completed:
                    show_final_quiz = True

                # Логика анимации завершения
                if user_course.status == 'completed' and not user_course.course_complete_animation_shown:
                    show_completion_animation = True
                    user_course.course_complete_animation_shown = True
                    user_course.save(update_fields=['course_complete_animation_shown'])
            else:
                # Для статуса 'available' используем все уроки курса
                lessons = course.lessons.all()
        else:
            # Для неаутентифицированных пользователей
            lessons = course.lessons.all()

        # Получение информации о следующем курсе в траектории
        next_course_in_trajectory = None
        user_trajectories_info = []
        if user.is_authenticated:
            # Ищем траектории, в которых есть этот курс
            user_trajectories = UserCourseTrajectory.objects.filter(
                user=user, 
                trajectory__courses=course
            )
            
            for user_trajectory in user_trajectories:
                # Получаем текущий курс в траектории
                current_tc = TrajectoryCourse.objects.filter(
                    trajectory=user_trajectory.trajectory, 
                    course=course
                ).first()
                
                if current_tc:
                    trajectory_info = {
                        'trajectory': user_trajectory.trajectory,
                        'user_trajectory': user_trajectory,
                        'order': current_tc.order,
                        'total_courses': TrajectoryCourse.objects.filter(trajectory=user_trajectory.trajectory).count()
                    }
                    user_trajectories_info.append(trajectory_info)
                    
                    # Ищем следующий курс в траектории только если текущий завершен
                    if user_course and user_course.status == 'completed':
                        next_tc = TrajectoryCourse.objects.filter(
                            trajectory=user_trajectory.trajectory,
                            order=current_tc.order + 1
                        ).first()
                        
                        if next_tc:
                            next_course_in_trajectory = next_tc.course
                            break

        # Проверяем параметры подсветки кнопки "Начать курс"
        highlight_start = request.GET.get('highlight_start') == '1'
        lesson_blocked_id = request.GET.get('lesson_blocked')
        quiz_blocked_id = request.GET.get('quiz_blocked')


        # Определяем страну пользователя
        user_country = ''
        if hasattr(request.user, 'profile') and request.user.profile:
            user_country = request.user.profile.country or ''

        # Проверяем deadline.
        # Для обычных курсов блокируем курс при просрочке.
        # Для курсов-инцидентов доступ не блокируем — уведомление о просрочке
        # рассылается по расписанию командой send_incident_course_overdue_notifications.
        is_deadline_overdue = False
        if user_course and user_course.deadline:
            is_deadline_overdue = timezone.now() > user_course.deadline
            if is_deadline_overdue:
                if not course.is_incident and user_course.status not in ['completed', 'blocked']:
                    user_course.status = 'blocked'
                    user_course.save(update_fields=['status'])
        
        # Получаем инцидент, связанный с курсом (если курс-инцидент)
        incident = None
        if course.is_incident:
            from builder.models import Incident
            incident = Incident.objects.filter(course=course).first()
        
        # Получаем информацию о связанных тестах для каждого урока
        lesson_quizzes_info = {}
        if user.is_authenticated:
            for lesson in course.lessons.all():
                if lesson.final_quiz:
                    quiz_result = QuizResult.objects.filter(
                        user=user,
                        course=course,
                        quiz_title=lesson.final_quiz.name,
                        passed=True
                    ).first()
                    lesson_quizzes_info[lesson.id] = {
                        'quiz': lesson.final_quiz,
                        'passed': quiz_result is not None,
                        'status': quiz_result.status if quiz_result else None
                    }
                    # Если нет пройденного теста, проверяем последний результат
                    if not quiz_result:
                        latest_result = QuizResult.objects.filter(
                            user=user,
                            course=course,
                            quiz_title=lesson.final_quiz.name
                        ).order_by('-completed_at').first()
                        if latest_result:
                            lesson_quizzes_info[lesson.id]['status'] = latest_result.status
        
        # Формирование контекста
        context.update({
            'course_author': course.author.username,
            'user_course': user_course,
            'progress': progress,
            'completed_lessons': completed_lessons,
            'completed_quizzes': completed_quizzes,
            'completed_homeworks': completed_homeworks,
            'completed_lessons_ids': completed_lessons_ids or [],
            'completed_quizzes_ids': completed_quizzes_ids or [],
            'completed_homeworks_ids': completed_homeworks_ids or [],
            'total_lessons': total_lessons,
            'total_quizzes': total_quizzes,
            'total_homeworks': total_homeworks,
            'total_materials': total_materials,
            'next_lesson': next_lesson,
            'next_material': next_lesson,  # Для совместимости с шаблоном
            'all_completed': all_completed,
            'show_completion_animation': show_completion_animation,
            'lessons': lessons,
            'show_final_quiz': show_final_quiz,
            'next_course_in_trajectory': next_course_in_trajectory,
            'user_trajectories_info': user_trajectories_info,
            'quiz_attempts_info': locals().get('quiz_attempts_info'),
            'quiz_passed': locals().get('final_quiz_passed', False),
            'final_quiz_status': locals().get('final_quiz_status'),
            'is_dental_checkup_course': course.title == "Чек-ап стоматологической клиники",
            'user_country': user_country,
            'highlight_start_button': highlight_start,
            'lesson_blocked_id': lesson_blocked_id,
            'quiz_blocked_id': quiz_blocked_id,
            'quiz_statuses': locals().get('quiz_statuses', {}),
            'homework_statuses': self._get_homework_statuses(user, course) if user.is_authenticated else {},
            'is_deadline_overdue': is_deadline_overdue,
            'incident': incident,
            'lesson_quizzes_info': lesson_quizzes_info,
        })
        
        return context

    def _get_homework_statuses(self, user, course):
        """
        Получает статусы заданий для пользователя в курсе.
        Возвращает словарь {homework_id: status}
        """
        from quizzes.models import HomeworkSubmission
        
        homework_statuses = {}
        for homework in course.homeworks:
            submission = HomeworkSubmission.objects.filter(
                user=user,
                homework=homework,
                course=course
            ).order_by('-submitted_at').first()
            if submission:
                homework_statuses[homework.id] = submission.status
            else:
                homework_statuses[homework.id] = None
        return homework_statuses
    
    def _get_next_material(self, user, course, trajectory, max_completed_order):
        """
        Находит следующий материал (урок или тест) для продолжения обучения
        """
        # Получаем все материалы курса в порядке
        materials = course.get_course_materials()
        
        # Получаем завершенные уроки и тесты
        completed_lessons_ids = set(
            UserProgress.objects.filter(
                user=user,
                course=course,
                completed=True
            ).values_list('lesson_id', flat=True)
        )
        
        completed_quizzes_ids = set(
            QuizResult.objects.filter(
                user=user,
                course=course,
                quiz_title__in=[quiz.name for quiz in course.quizzes],
                passed=True
            ).values_list('quiz_title', flat=True)
        )
        
        # Если есть траектория, фильтруем материалы по траектории
        if trajectory:
            trajectory_lesson_ids = set(trajectory.lessons.values_list('id', flat=True))
            materials = [m for m in materials if m['type'] == 'quiz' or m['id'] in trajectory_lesson_ids]
        
        # Ищем первый незавершенный материал
        for material in materials:
            if material['type'] == 'lesson':
                if material['id'] not in completed_lessons_ids:
                    return material
            elif material['type'] == 'quiz':
                if material['title'] not in completed_quizzes_ids:
                    return material
        
        return None




class LessonDetailView(DetailView):
    """Деталка урока"""
    model = Lesson
    template_name = 'courses/lesson_detail.html'
    context_object_name = 'lesson'
    
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            return redirect('users:login')
        return super().dispatch(request, *args, **kwargs)


    def get_object(self, queryset=None):
        course_slug = self.kwargs.get('course_slug')
        lesson_id = self.kwargs.get('lesson_id')
        
        course = get_object_or_404(Course, slug=course_slug)
        lesson = get_object_or_404(Lesson, id=lesson_id, courses=course)
        
        # Проверка доступа к курсу через менеджер (пропускаем для админов)
        if not (self.request.user.is_staff or self.request.user.is_superuser):
            available_courses = Course.objects.available_for_user(self.request.user)
            if course not in available_courses:
                return redirect('courses:course_detail', slug=course.slug)
        
        return lesson


    def get(self, request, *args, **kwargs):
        """Переопределяем get для проверки доступа перед рендерингом"""
        self.object = self.get_object()
        lesson = self.object
        course_slug = self.kwargs.get('course_slug')
        course = get_object_or_404(Course, slug=course_slug)
        
        # Получаем UserCourse для проверки статуса
        user_course = UserCourse.objects.filter(user=request.user, course=course).first()
        if not user_course:
            # Курсы из траекторий назначаются только после завершения предыдущего — не создаём запись здесь
            if UserCourseTrajectory.objects.filter(user=request.user, trajectory__trajectorycourse__course=course).exists():
                return redirect('courses:course_detail', slug=course.slug)
            # Проверяем, не был ли курс отменен вручную
            from myapp.models import ManualCourseUnassignment
            manual_unassignment = ManualCourseUnassignment.objects.filter(
                user=request.user, 
                course=course
            ).first()
            
            if not manual_unassignment:
                user_course, created = UserCourse.objects.get_or_create(
                    user=request.user, 
                    course=course, 
                    defaults={'status': 'available'}
                )
            else:
                # Если курс был отменён вручную, перенаправляем на страницу курса
                from django.urls import reverse
                return redirect('courses:course_detail', slug=course.slug)

        # Блокируем доступ к уроку, если курс не начат (пропускаем для админов)
        if not (request.user.is_staff or request.user.is_superuser):
            if user_course.status not in ['started', 'completed']:
                from django.urls import reverse
                from urllib.parse import urlencode
                url = reverse('courses:course_detail', kwargs={'slug': course.slug})
                params = urlencode({'highlight_start': '1', 'lesson_blocked': lesson.id})
                return redirect(f'{url}?{params}')

        # Проверка траектории (пропускаем для админов)
        trajectory = UserLessonTrajectory.objects.filter(user=request.user, course=course).first()
        if trajectory and not (request.user.is_staff or request.user.is_superuser):
            lessons_in_trajectory = trajectory.lessons.all()
            if lesson not in lessons_in_trajectory:
                return render(request, 'courses/lesson_access_denied.html', {'course': course, 'lesson': lesson})
        
        # Проверяем, был ли только что пройден тест (параметр из GET-запроса)
        quiz_completed = request.GET.get('quiz_completed') == '1'
        
        # Если все проверки пройдены, вызываем стандартный метод get
        response = super().get(request, *args, **kwargs)
        
        # Добавляем информацию о завершении теста в контекст
        if quiz_completed:
            response.context_data['quiz_just_completed'] = True
        
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        lesson = self.object
        course_slug = self.kwargs.get('course_slug')
        course = get_object_or_404(Course, slug=course_slug)
        
        # Получаем траекторию для определения предыдущего и следующего уроков
        trajectory = UserLessonTrajectory.objects.filter(user=self.request.user, course=course).first()
        
        # Определяем страну пользователя (нужна для фильтрации уроков метрик)
        user_country = ''
        if hasattr(self.request.user, 'profile') and self.request.user.profile:
            user_country = self.request.user.profile.country or ''

        # Определяем предыдущий и следующий уроки
        if trajectory:
            trajectory_lessons = trajectory.lessons.all().order_by('order')
            previous_lesson = trajectory_lessons.filter(
                order__lt=lesson.order
            ).order_by('-order').first()
            next_lesson = trajectory_lessons.filter(
                order__gt=lesson.order
            ).order_by('order').first()
        else:
            previous_lesson = lesson.get_previous_lesson(course)
            next_lesson = lesson.get_next_lesson(course)

        # Специальная логика для курса "Чек-ап стоматологической клиники"
        is_dental_checkup_course = course.title == "Чек-ап стоматологической клиники"
        
        # Пропускаем скрытые уроки метрик в зависимости от страны
        if is_dental_checkup_course:
            # Для Казахстана скрываем обычный урок метрик
            if user_country == 'Казахстан':
                if previous_lesson and previous_lesson.title == 'Метрики эффективности стоматологической клиники':
                    previous_lesson = previous_lesson.get_previous_lesson(course)
                if next_lesson and next_lesson.title == 'Метрики эффективности стоматологической клиники':
                    next_lesson = next_lesson.get_next_lesson(course)
            # Для других стран скрываем KZ урок метрик
            else:
                if previous_lesson and previous_lesson.title == 'KZ Метрики эффективности стоматологической клиники':
                    previous_lesson = previous_lesson.get_previous_lesson(course)
                if next_lesson and next_lesson.title == 'KZ Метрики эффективности стоматологической клиники':
                    next_lesson = next_lesson.get_next_lesson(course)

        # Помечаем урок как просмотренный
        UserProgress.objects.get_or_create(
            user=self.request.user,
            course=course,
            lesson=lesson,
            defaults={'completed': False}
        )
        
        is_first_lesson = False
        if is_dental_checkup_course:
            course_lessons = course.lessons
            if len(course_lessons) >= 1:
                first_lesson = course_lessons.first()
                is_first_lesson = lesson.id == first_lesson.id

        # Специальная логика для урока "Метрики эффективности стоматологической клиники"
        is_metrics_lesson = lesson.title == "Метрики эффективности стоматологической клиники"
        
        # Специальная логика для урока "KZ Метрики эффективности стоматологической клиники"
        is_metrics_kz_lesson = lesson.title == "KZ Метрики эффективности стоматологической клиники"
        
        # Проверяем, является ли пользователь внешним пользователем
        is_external_user = self.request.user.groups.filter(name='Внешний пользователь').exists()

        # Проверяем, является ли урок последним
        is_last_lesson = next_lesson is None
        
        # Определяем следующий материал (урок или тест) после текущего урока
        next_material = None
        if not is_last_lesson:
            # Используем логику из _get_next_material для определения следующего элемента
            next_material = self._get_next_material_after_lesson(lesson, course, trajectory)
        
        # Проверяем наличие связанного теста и его статус
        lesson_quiz = None
        lesson_quiz_passed = False
        lesson_quiz_status = None
        if lesson.final_quiz:
            lesson_quiz = lesson.final_quiz
            # Проверяем, завершен ли тест
            quiz_result = QuizResult.objects.filter(
                user=self.request.user,
                course=course,
                quiz_title=lesson.final_quiz.name,
                passed=True
            ).first()
            if quiz_result:
                lesson_quiz_passed = True
                lesson_quiz_status = quiz_result.status
            else:
                # Проверяем, есть ли незавершенная попытка
                latest_result = QuizResult.objects.filter(
                    user=self.request.user,
                    course=course,
                    quiz_title=lesson.final_quiz.name
                ).order_by('-completed_at').first()
                if latest_result:
                    lesson_quiz_status = latest_result.status
        
        context.update({
            'course': course,
            'previous_lesson': previous_lesson,
            'next_lesson': next_lesson,
            'next_material': next_material,
            'is_dental_checkup_course': is_dental_checkup_course,
            'is_first_lesson': is_first_lesson,
            'is_last_lesson': is_last_lesson,
            'is_metrics_lesson': is_metrics_lesson,
            'is_metrics_kz_lesson': is_metrics_kz_lesson,
            'is_external_user': is_external_user,
            'user_country': user_country,
            'lesson_quiz': lesson_quiz,
            'lesson_quiz_passed': lesson_quiz_passed,
            'lesson_quiz_status': lesson_quiz_status,
        })
        
        return context

    def _get_next_material_after_lesson(self, lesson, course, trajectory):
        """
        Находит следующий материал (урок, тест или задание) после текущего урока
        """
        from quizzes.models import HomeworkSubmission
        
        # Получаем все материалы курса в порядке
        materials = course.get_course_materials()
        
        # Получаем завершенные уроки и тесты
        completed_lessons_ids = set(
            UserProgress.objects.filter(
                user=self.request.user,
                course=course,
                completed=True
            ).values_list('lesson_id', flat=True)
        )
        
        completed_quizzes_ids = set(
            QuizResult.objects.filter(
                user=self.request.user,
                course=course,
                quiz_title__in=[quiz.name for quiz in course.quizzes],
                passed=True
            ).values_list('quiz_title', flat=True)
        )
        
        # Получаем завершенные задания (correct)
        completed_homework_ids = set(
            HomeworkSubmission.objects.filter(
                user=self.request.user,
                course=course,
                status='correct'
            ).values_list('homework_id', flat=True)
        )
        
        # Если есть траектория, фильтруем материалы по траектории
        if trajectory:
            trajectory_lesson_ids = set(trajectory.lessons.values_list('id', flat=True))
            materials = [m for m in materials if m['type'] in ('quiz', 'homework') or m['id'] in trajectory_lesson_ids]
        
        # Находим текущий урок в списке материалов
        current_lesson_index = None
        for i, material in enumerate(materials):
            if material['type'] == 'lesson' and material['id'] == lesson.id:
                current_lesson_index = i
                break
        
        if current_lesson_index is None:
            return None
        
        # Ищем следующий незавершенный материал после текущего урока
        for i in range(current_lesson_index + 1, len(materials)):
            material = materials[i]
            if material['type'] == 'lesson':
                if material['id'] not in completed_lessons_ids:
                    return material
            elif material['type'] == 'quiz':
                if material['title'] not in completed_quizzes_ids:
                    return material
            elif material['type'] == 'homework':
                if material['id'] not in completed_homework_ids:
                    return material
        
        return None


class CreateCourseView(LoginRequiredMixin, UserPassesTestMixin, CreateView):
    """
    Создание нового курса.
    """
    model = Course
    form_class = CourseForm
    template_name = 'courses/create_course.html'
    success_url = reverse_lazy('courses:course_detail')

    def test_func(self):
        """
        Доступ только для администраторов.
        """
        return is_admin(self.request.user)

    def get_form_kwargs(self):
        """Передает is_incident в форму, если параметр передан в GET."""
        kwargs = super().get_form_kwargs()
        is_incident_param = self.request.GET.get('is_incident')
        if is_incident_param == '1':
            if 'initial' not in kwargs:
                kwargs['initial'] = {}
            kwargs['initial']['is_incident'] = True
            kwargs['initial']['is_incident_readonly'] = True
        return kwargs

    def form_valid(self, form):
        """
        Обработка успешной валидации формы
        """
        course = form.save(commit=False)
        course.author = self.request.user
        
        # Убеждаемся, что is_incident установлен, если форма была заблокирована
        if form.fields.get('is_incident') and form.fields['is_incident'].disabled:
            course.is_incident = True
        
        course.save()
        form.save_m2m()

        self._assign_course_to_staff(course)
        return redirect('courses:course_detail', slug=course.slug)

    def _assign_course_to_staff(self, course):
        """
        Назначаем курс всем staff\\superuser
        """
        User = get_user_model()
        staff_users = User.objects.filter(
            is_active=True
        ).filter(
            models.Q(is_staff=True) | models.Q(is_superuser=True)
        ).distinct()

        for user in staff_users:
            UserCourse.objects.get_or_create(
                user=user,
                course=course,
                defaults={'status': 'available'}
            )




class CreateLessonView(LoginRequiredMixin, CreateView):
    """
    Создание урока в курсе с автоматическим порядком
    """
    model = Lesson
    form_class = LessonForm
    template_name = 'courses/create_lesson.html'
    

    def dispatch(self, request, *args, **kwargs):
        """Получаем курс и сохраняем на self для использования в других методах."""
        self.course = get_object_or_404(Course, slug=kwargs['course_slug'])
        return super().dispatch(request, *args, **kwargs)


    def get_form_kwargs(self):
        """Передает initial-данные и признак скрытия `order` в форму."""
        kwargs = super().get_form_kwargs()
        max_order = self.course.lessons.aggregate(models.Max('order'))['order__max'] or 0
        
        kwargs['initial'] = {
            'order': max_order + 1, 
            'courses': [self.course]
        }
        kwargs['hide_order'] = True
        return kwargs


    def form_valid(self, form):
        """Сохраняет урок, назначает следующий `order`, привязывает к курсу."""
        lesson = form.save(commit=False)
        lesson.order = self.course.lessons.aggregate(models.Max('order'))['order__max'] or 0 + 1
        lesson.save()
        form.save_m2m()  # Сохраняем связь many-to-many с курсами
        return redirect('courses:course_detail', course_slug=self.course.slug)


    def get_context_data(self, **kwargs):
        """Добавляет выбранный курс в контекст шаблона."""
        context = super().get_context_data(**kwargs)
        context['course'] = self.course
        return context




def get_category_full_path(category):
    """Возвращает полный путь категории через `/`, включая родителей."""
    path = [category.name]
    parent = category.parent
    while parent:
        path.append(parent.name)
        parent = parent.parent
    return '/'.join(reversed(path))




class AddLessonView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    Добавление уроков/тестов в курс из существующих материалов
    """
    template_name = 'courses/add_lesson.html'
    
    def test_func(self):
        """Проверка прав: только админы"""
        return is_admin(self.request.user)
    
    def dispatch(self, request, *args, **kwargs):
        """Получаем курс и сохраняем на self"""
        self.course = get_object_or_404(Course, slug=kwargs['course_slug'])
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        """GET запрос - отображение формы выбора материалов"""
        context = self.get_context_data()
        return render(request, self.template_name, context)
    
    def post(self, request, *args, **kwargs):
        """POST запрос - обработка добавления материалов"""
        if 'add_selected' in request.POST:
            return self._handle_add_selected()
        elif 'create_new' in request.POST:
            return redirect('courses:create_lesson', course_slug=self.course.slug)
        
        # Если ни одна кнопка не нажата, возвращаем на форму
        return self.get(request, *args, **kwargs)
    
    def _handle_add_selected(self):
        """Обработка добавления выбранных материалов"""
        selected_items_str = self.request.POST.get('selected_items', '')
        selected_items = [item.strip() for item in selected_items_str.split(',') if item.strip()]
        
        max_order = self.course.lessons.aggregate(Max('order'))['order__max'] or 0
        current_order = max_order + 1
        
        for item_id in selected_items:
            if item_id.startswith('category_'):
                self._add_category_lessons(item_id, current_order)
                current_order += self._count_lessons_in_category(item_id)
            elif item_id.startswith('lesson_'):
                current_order = self._add_single_lesson(item_id, current_order)
            elif item_id.startswith('uncategorized_'):
                current_order = self._add_uncategorized_lesson(item_id, current_order)
            elif item_id.startswith('quiz_'):
                current_order = self._add_quiz(item_id, current_order)
            elif item_id.startswith('homework_'):
                current_order = self._add_homework(item_id, current_order)
        
        return redirect('courses:course_detail', slug=self.course.slug)
    
    def _add_category_lessons(self, item_id, start_order):
        """Добавление всех уроков из категории"""
        category_id = item_id.replace('category_', '')
        lessons = self._get_all_lessons_in_category(category_id)
        
        current_order = start_order
        for lesson in lessons:
            if self.course not in lesson.courses.all():
                lesson.courses.add(self.course)
                lesson.order = current_order
                lesson.save()
                current_order += 1
    
    def _add_single_lesson(self, item_id, current_order):
        """Добавление отдельного урока"""
        lesson_id = item_id.replace('lesson_', '')
        lesson = get_object_or_404(Lesson, id=lesson_id)
        
        if self.course not in lesson.courses.all():
            lesson.courses.add(self.course)
            lesson.order = current_order
            lesson.save()
            return current_order + 1
        return current_order
    
    def _add_uncategorized_lesson(self, item_id, current_order):
        """Добавление урока без категории"""
        lesson_id = item_id.replace('uncategorized_', '')
        lesson = get_object_or_404(Lesson, id=lesson_id)
        
        if self.course not in lesson.courses.all():
            lesson.courses.add(self.course)
            lesson.order = current_order
            lesson.save()
            return current_order + 1
        return current_order
    
    def _add_quiz(self, item_id, current_order):
        """Добавление теста"""
        quiz_id = item_id.replace('quiz_', '')
        quiz = get_object_or_404(Quiz, id=quiz_id)
        
        if self.course not in quiz.courses.all():
            quiz.courses.add(self.course)
            quiz.order = current_order
            quiz.save()
            return current_order + 1
        return current_order
    
    def _add_homework(self, item_id, current_order):
        """Добавление задания"""
        from quizzes.models import Homework
        homework_id = item_id.replace('homework_', '')
        homework = get_object_or_404(Homework, id=homework_id)
        
        if self.course not in homework.courses.all():
            homework.courses.add(self.course)
            homework.order = current_order
            homework.save()
            return current_order + 1
        return current_order
    
    def _count_lessons_in_category(self, item_id):
        """Подсчет уроков в категории для правильного порядка"""
        category_id = item_id.replace('category_', '')
        lessons = self._get_all_lessons_in_category(category_id)
        return len([l for l in lessons if self.course not in l.courses.all()])
    
    def _get_all_lessons_in_category(self, category_id):
        """Получение всех уроков категории (включая подкатегории)"""
        category = CategoryName.objects.get(id=category_id)
        lessons = set()
        
        # Добавляем уроки текущей категории
        lessons.update(category.lessons.all())
        lessons.update([mirror.lesson for mirror in category.mirrored_lessons.all()])
        
        # Рекурсивно добавляем уроки подкатегорий
        for subcat in category.subcategories.all():
            lessons.update(self._get_all_lessons_in_category(subcat.id))
        
        return list(lessons)
    
    def _get_categories_with_lessons(self):
        """Получение всех категорий с их подкатегориями и уроками"""
        categories = CategoryName.objects.filter(parent=None).prefetch_related(
            'subcategories', 'lessons', 'mirrored_lessons__lesson'
        ).order_by('order', 'name')
        
        def process_category(cat):
            # Получаем все уроки категории (включая зеркала)
            lessons = list(cat.lessons.all())
            for mirror in cat.mirrored_lessons.all():
                if mirror.lesson not in lessons:
                    lessons.append(mirror.lesson)
            
            # Обрабатываем подкатегории
            subcategories = []
            for subcat in cat.subcategories.all():
                subcategories.append(process_category(subcat))
            
            return {
                'id': cat.id,
                'name': cat.name,
                'lessons': lessons,
                'subcategories': subcategories
            }
        
        return [process_category(cat) for cat in categories]
    
    def get_context_data(self):
        """Формирование контекста для шаблона"""
        from quizzes.models import Homework
        return {
            'course': self.course,
            'categories_data': self._get_categories_with_lessons(),
            'uncategorized_lessons': Lesson.objects.filter(category__isnull=True).order_by('order', 'title'),
            'all_quizzes': Quiz.objects.all().order_by('name'),
            'all_homeworks': Homework.objects.all().order_by('name'),
        }




@login_required
@user_passes_test(is_admin, login_url='/')
def reorder_materials(request, course_slug):
    """Страница для изменения порядка материалов курса (уроков и тестов)"""
    course = get_object_or_404(Course, slug=course_slug)
    
    if request.method == 'POST':
        # Обработка AJAX запроса для сохранения нового порядка
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            import json
            try:
                data = json.loads(request.body)
                materials_order = data.get('materials_order', [])
                
                # Обновляем порядок для каждого материала
                for index, material_data in enumerate(materials_order):
                    material_type = material_data['type']
                    material_id = material_data['id']
                    new_order = index + 1
                    
                    if material_type == 'lesson':
                        lesson = Lesson.objects.get(id=material_id)
                        lesson.order = new_order
                        lesson.save()
                    elif material_type == 'quiz':
                        quiz = Quiz.objects.get(id=material_id)
                        quiz.order = new_order
                        quiz.save()
                
                return JsonResponse({'success': True})
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)})
        else:
            # Обычная форма - редирект обратно к курсу
            return redirect('courses:course_detail', slug=course.slug)
    
    # GET запрос - показываем страницу с интерфейсом
    materials = course.get_course_materials()
    
    return render(request, 'courses/reorder_materials.html', {
        'course': course,
        'materials': materials,
    })




class DeleteCourseView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    """
    Удаление курса с поддержкой AJAX
    """
    model = Course
    slug_url_kwarg = 'slug'
    success_url = reverse_lazy('home')


    def test_func(self):
        """Проверка прав: только админ"""
        return is_admin(self.request.user)


    def delete(self, request, *args, **kwargs):
        """Удаляем объект и возвращаем предсказуемый JSON для любых POST-запросов.

        На практике удаление вызывается из AJAX. Ранее при потере заголовка
        X-Requested-With происходил редирект и фронт получал HTML без поля
        success, из‑за чего показывалось сообщение об ошибке, хотя удаление
        выполнялось. Возвращаем JSON всегда для POST, чтобы поведение было стабильным.
        """
        self.object = self.get_object()
        self.object.delete()

        # Возвращаем единый успешный ответ в формате JSON
        return JsonResponse({"success": True})

    def get(self, request, *args, **kwargs):
        """GET запрос - редирект на страницу курса"""
        return redirect('courses:course_detail', slug=kwargs['slug'])




class DeleteLessonView(LoginRequiredMixin, UserPassesTestMixin, DeleteView):
    model = Lesson
    success_url = reverse_lazy('home')

    def test_func(self):
        """Проверка прав: только админ"""
        return is_admin(self.request.user)


    def dispatch(self, request, *args, **kwargs):
        """
        Получаем урок и сохраняем на self
        """
        self.lesson = get_object_or_404(Lesson, id=kwargs['lesson_id'])
        self.course = self.lesson.courses.first()
        return super().dispatch(request, *args, **kwargs)

    
    def get(self, request, *args, **kwargs):
        """Редирект на страницу курса или на главную"""
        if self.course:
            return redirect('courses:course_detail', slug=self.course.slug)
        else:
            return redirect('home')


    def post(self, request, *args, **kwargs):
        """Удаление свзяи урока с курсом"""
        if self.course:
            self.lesson.courses.remove(self.course)
            return redirect('courses:course_detail', slug=self.course.slug)
        else:
            # Если урок не связан ни с одним курсом, редиректим на главную
            return redirect('home')


@login_required
@user_passes_test(is_admin, login_url='/')
def remove_quiz_from_course(request, course_slug, quiz_id):
    """Удаление теста из курса"""
    course = get_object_or_404(Course, slug=course_slug)
    quiz = get_object_or_404(Quiz, id=quiz_id)

    if request.method == 'POST':
        # Удаляем связь между тестом и курсом
        course.course_quizzes.remove(quiz)
        return redirect('courses:course_detail', slug=course.slug)

    # GET запрос - редирект на страницу курса
    return redirect('courses:course_detail', slug=course.slug)


@login_required
@user_passes_test(is_admin, login_url='/')
def remove_homework_from_course(request, course_slug, homework_id):
    """Удаление задания из курса"""
    from quizzes.models import Homework
    course = get_object_or_404(Course, slug=course_slug)
    homework = get_object_or_404(Homework, id=homework_id)

    if request.method == 'POST':
        # Удаляем связь между заданием и курсом
        course.course_homeworks.remove(homework)
        return redirect('courses:course_detail', slug=course.slug)

    # GET запрос - редирект на страницу курса
    return redirect('courses:course_detail', slug=course.slug)




@login_required
@user_passes_test(lambda u: is_author_or_admin(u, Course), login_url='/')
def edit_course(request, slug):
    course = get_object_or_404(Course, slug=slug)
    
    if request.method == 'POST':
        form = CourseForm(request.POST, request.FILES, instance=course)
        if form.is_valid():
            form.save()
            return redirect('courses:course_detail', slug=course.slug)
    else:
        form = CourseForm(instance=course)
    
    return render(request, 'courses/edit_course.html', {
        'form': form,
        'course': course
    })




@login_required
@user_passes_test(is_admin, login_url='/')
def edit_lesson(request, lesson_id):
    lesson = get_object_or_404(Lesson, id=lesson_id)
    # Получаем первый курс для контекста (или можно изменить логику)
    course = lesson.courses.first()
    
    if request.method == 'POST':
        form = LessonForm(request.POST, instance=lesson)
        if form.is_valid():
            form.save()
            if course:
                return redirect('courses:lesson_detail', course_slug=course.slug, lesson_id=lesson.id)
            else:
                return redirect('home')
    else:
        form = LessonForm(instance=lesson)
    
    return render(request, 'courses/edit_lesson.html', {
        'form': form,
        'course': course,
        'lesson': lesson
    })




@require_http_methods(["GET", "POST"])
def redir_to_quiz(request, course_slug):
    course = get_object_or_404(Course, slug=course_slug)

    if request.method == 'POST':
        # Проверяем, какую кнопку нажал пользователь
        action = request.POST.get('action')
        if action == 'start_quiz':
            request.session['course_slug'] = course.slug
            from django.urls import reverse
            from urllib.parse import urlencode
            url = reverse('quizzes:quiz_start', kwargs={'quiz_id': course.final_quiz.id})
            params = urlencode({'course_slug': course.slug})
            return redirect(f'{url}?{params}')
        else:
            return redirect('courses:course_detail', slug=course.slug)

    # GET-запрос - показываем страницу с подтверждением
    return render(request, 'courses/redir_to_quiz.html', {'course': course})




@require_POST
def complete_lesson(request, course_slug, lesson_id):
    """Отмечает урок как завершенный и начисляет пользователю опыт"""
    if not request.user.is_authenticated:
        return redirect('users:login')
    
    course = get_object_or_404(Course, slug=course_slug)
    lesson = get_object_or_404(Lesson, id=lesson_id, courses=course)
    user = request.user
    
    # Проверка доступа через менеджер
    available_courses = Course.objects.available_for_user(user)
    if course not in available_courses:
        return redirect('courses:course_detail', slug=course.slug)
    
    # Получаем UserCourse
    user_course = UserCourse.objects.filter(user=user, course=course).first()
    if not user_course:
        # Курсы из траекторий назначаются только после завершения предыдущего — не создаём запись здесь
        if UserCourseTrajectory.objects.filter(user=user, trajectory__trajectorycourse__course=course).exists():
            return redirect('courses:course_detail', slug=course.slug)
        # Проверяем, не был ли курс отменен вручную
        from myapp.models import ManualCourseUnassignment
        manual_unassignment = ManualCourseUnassignment.objects.filter(
            user=user, 
            course=course
        ).first()
        
        if not manual_unassignment:
            user_course, created = UserCourse.objects.get_or_create(
                user=user, 
                course=course, 
                defaults={'status': 'available'}
            )
        else:
            # Если курс был отменён вручную, перенаправляем на страницу курса
            return redirect('courses:course_detail', slug=course.slug)
    
    # Получаем траекторию пользователя
    trajectory = UserLessonTrajectory.objects.filter(user=user, course=course).first()
    
    # Проверяем, что урок входит в траекторию пользователя (если траектория задана)
    if trajectory and lesson not in trajectory.lessons.all():
        return redirect('courses:course_detail', slug=course.slug)

    # Проверяем, хочет ли пользователь вернуться к курсу (кнопка "Отдохну")
    return_to_course = request.POST.get('return_to_course') == 'true'
    
    # Проверяем, есть ли у урока связанный тест, и если есть - завершен ли он
    if lesson.final_quiz:
        quiz_passed = QuizResult.objects.filter(
            user=user,
            course=course,
            quiz_title=lesson.final_quiz.name,
            passed=True
        ).exists()
        
        if not quiz_passed:
            # Тест не завершен
            if return_to_course:
                # Пользователь нажал "Отдохну" - возвращаемся к курсу БЕЗ засчитывания урока
                return redirect('courses:course_detail', slug=course.slug)
            else:
                # Пользователь хочет завершить урок - перенаправляем на тест
                from django.urls import reverse
                from urllib.parse import urlencode
                url = reverse('quizzes:quiz_start', kwargs={'quiz_id': lesson.final_quiz.id})
                params = urlencode({'course_slug': course.slug, 'lesson_id': lesson.id})
                return redirect(f'{url}?{params}')

    # Если мы дошли сюда, значит либо теста нет, либо тест пройден
    # Засчитываем урок как завершенный
    
    # Проверяем, был ли урок уже завершен ранее
    progress, created = UserProgress.objects.get_or_create(
        user=user,
        course=course,
        lesson=lesson,
        defaults={'completed': False}
    )
    
    # Проверяем, получал ли пользователь уже баллы за этот урок в рамках данного курса
    from gamification.models import DascoinTransaction
    lesson_reward_reason = f"Завершение урока {lesson.title}"
    already_rewarded = DascoinTransaction.objects.filter(
        user=user,
        reason=lesson_reward_reason,
        transaction_type='award'
    ).exists()
    
    # Начисляем очки только если урок завершается впервые И баллы не были начислены ранее И курс не является инцидентом
    if not progress.completed and not already_rewarded and not course.is_incident:
        award_dascoin_points(user, lesson.points, lesson_reward_reason)
    
    # Создаем или обновляем прогресс
    UserProgress.objects.update_or_create(
        user=user,
        course=course,
        lesson=lesson,
        defaults={'completed': True}
    )
    
    # Проверяем и выдаем бейдж "Первый шаг" за первый урок после регистрации
    award_first_lesson_badge(user)

    # Получаем общее количество материалов для пользователя
    if trajectory:
        total_lessons = trajectory.lessons.count()
        lesson_ids = trajectory.lessons.values_list('id', flat=True)
    else:
        total_lessons = course.lessons.count()
        lesson_ids = course.lessons.values_list('id', flat=True)
    
    total_quizzes = course.quizzes.count()
    total_homeworks = course.homeworks.count()

    # Считаем пройденные уроки
    completed_lessons = UserProgress.objects.filter(
        user=user,
        course=course,
        completed=True,
        lesson_id__in=lesson_ids
    ).count()
    
    # Считаем пройденные тесты в рамках этого курса
    completed_quizzes = QuizResult.objects.filter(
        user=user,
        course=course,
        quiz_title__in=[quiz.name for quiz in course.quizzes],
        passed=True
    ).count()
    
    # Считаем выполненные задания в рамках этого курса
    completed_homeworks = HomeworkSubmission.objects.filter(
        user=user,
        course=course,
        homework__in=course.homeworks,
        status='correct'
    ).values('homework_id').distinct().count()

    # Курс завершен только если пройдены ВСЕ уроки, ВСЕ тесты И ВСЕ задания
    all_completed = (completed_lessons >= total_lessons and 
                    completed_quizzes >= total_quizzes and 
                    completed_homeworks >= total_homeworks)

    user_course = UserCourse.objects.get(user=user, course=course)
    
    if all_completed:
        if course.final_quiz:
            # Проверяем, прошел ли пользователь уже финальный тест
            final_quiz_passed = QuizResult.objects.filter(
                user=user,
                course=course,
                quiz_title=course.final_quiz.name,
                passed=True
            ).exists()
            
            if not final_quiz_passed:
                # Финальный тест еще не пройден - показываем страницу с предложением пройти тест
                return redirect('courses:redir_to_quiz', course_slug=course_slug)
            else:
                # Финальный тест уже пройден - просто обновляем статус курса
                was_completed_before = user_course.status == 'completed'
                if not was_completed_before:
                    user_course.status = 'completed'
                    user_course.save()
                    # Начисляем баллы только если курс не является инцидентом
                    if not course.is_incident:
                        award_dascoin_points(user, course.points, f"Завершение курса {course.title}")
                        award_course_badge(user, course)
                        # Выдаем сертификат за курс (если настроено)
                        issue_certificate(user, course=course)
                else:
                    # Курс уже был завершен, просто обновляем статус
                    user_course.status = 'completed'
                    user_course.save()
        else:
            # Проверяем, был ли курс уже завершен ранее
            was_completed_before = user_course.status == 'completed'
            if not was_completed_before:
                user_course.status = 'completed'
                user_course.save()
                # Начисляем баллы только если курс не является инцидентом
                if not course.is_incident:
                    award_dascoin_points(user, course.points, f"Завершение курса {course.title}")
                    award_course_badge(user, course)
                    # Выдаем сертификат за курс (если настроено)
                    issue_certificate(user, course=course)
            else:
                # Курс уже был завершен, просто обновляем статус
                user_course.status = 'completed'
                user_course.save()
    
    # Проверяем параметры из модального окна
    # Если пользователь нажал "Отдохну" и урок засчитан (тест пройден или теста нет) - возвращаемся к курсу
    if return_to_course:
        return redirect('courses:course_detail', slug=course.slug)
    elif request.POST.get('go_to_quiz'):
        # Пользователь хочет перейти к тесту после завершения урока
        quiz_id = request.POST.get('go_to_quiz')
        from django.urls import reverse
        from urllib.parse import urlencode
        url = reverse('quizzes:quiz_start', kwargs={'quiz_id': quiz_id})
        params = urlencode({'course_slug': course.slug})
        return redirect(f'{url}?{params}')
    elif request.POST.get('go_to_homework'):
        # Пользователь хочет перейти к заданию после завершения урока
        homework_id = request.POST.get('go_to_homework')
        from django.urls import reverse
        from urllib.parse import urlencode
        url = reverse('quizzes:homework_submit', kwargs={'homework_id': homework_id})
        params = urlencode({'course_slug': course.slug})
        return redirect(f'{url}?{params}')
    elif request.POST.get('continue_learning'):
        # Пользователь хочет продолжить обучение - ищем следующий урок
        # Используем ту же логику, что и в LessonDetailView
        next_lesson = None
        
        if trajectory:
            # Получаем уроки в порядке траектории (по order, как в LessonDetailView)
            trajectory_lessons = trajectory.lessons.all().order_by('order')
            next_lesson = trajectory_lessons.filter(
                order__gt=lesson.order
            ).order_by('order').first()
        else:
            # Используем метод get_next_lesson, как в LessonDetailView
            next_lesson = lesson.get_next_lesson(course)
        
        # Если есть следующий урок - переходим к нему, иначе возвращаемся к курсу
        if next_lesson:
            return redirect('courses:lesson_detail', course_slug=course.slug, lesson_id=next_lesson.id)
    
    # По умолчанию или если нет следующего урока - возвращаемся к курсу
    return redirect('courses:course_detail', slug=course.slug)




def complete_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    user_course = UserCourse.objects.get(user=request.user, course=course)
    user = request.user
    
    if course.final_quiz:
        quiz_result = QuizResult.objects.filter(
            user=request.user,
            quiz=course.final_quiz,
            passed=True
        ).exists()
        
        if quiz_result:
            # Проверяем, был ли курс уже завершен ранее
            was_completed_before = user_course.status == 'completed'
            if not was_completed_before:
                user_course.status = 'completed'
                user_course.save()
                # Начисляем очки только если курс не является инцидентом
                if not course.is_incident:
                    award_dascoin_points(user, course.points, f"Завершение курса {course.title}") 
                    award_course_badge(user, course)
                    # Выдаем сертификат за курс (если настроено)
                    issue_certificate(user, course=course)
            else:
                # Курс уже был завершен, просто обновляем статус
                user_course.status = 'completed'
                user_course.save()
            return redirect('courses:course_detail', slug=course.slug)
        else:
            from django.urls import reverse
            from urllib.parse import urlencode
            url = reverse('quizzes:quiz_start', kwargs={'quiz_id': course.final_quiz.id})
            params = urlencode({'course_slug': course.slug})
            return redirect(f'{url}?{params}')
    else:
        # Проверяем, был ли курс уже завершен ранее
        was_completed_before = user_course.status == 'completed'
        if not was_completed_before:
            user_course.status = 'completed'
            user_course.save()
            # Начисляем очки только если курс не является инцидентом
            if not course.is_incident:
                award_dascoin_points(user, course.points, f"Завершение курса {course.title}") 
                award_course_badge(user, course)
                # Выдаем сертификат за курс (если настроено)
                issue_certificate(user, course=course)
        else:
            # Курс уже был завершен, просто обновляем статус
            user_course.status = 'completed'
            user_course.save()
        return redirect('courses:course_detail', slug=course.slug)
    



@method_decorator(login_required, name='dispatch')
class UserCourseTrajectoryListView(ListView):
    """
    Список всех траекторий пользователя и прогресс по курсам.
    """
    model = UserCourseTrajectory
    template_name = 'courses/user_course_trajectory_list.html'
    context_object_name = 'user_trajectories'

    def get(self, request, *args, **kwargs):
        """
        Кэширует полный HTML-ответ страницы со списком траекторий и прогрессом по курсам.
        Ключ кэша зависит от пользователя, версии кэша и полного URL (включая GET-параметры).
        """
        if request.user.is_authenticated:
            user_id = request.user.pk
            version = _get_user_cache_version(user_id)
            cache_key = f"user_course_trajectory_list_page:user_{user_id}:v{version}:{request.get_full_path()}"
        else:
            cache_key = f"user_course_trajectory_list_page:user_anon:{request.get_full_path()}"

        cached_content = cache.get(cache_key)
        if cached_content is not None:
            return HttpResponse(cached_content, content_type='text/html; charset=utf-8')

        response = super().get(request, *args, **kwargs)
        if response.status_code == 200:
            content_type = response.get('Content-Type', '')
            if content_type.startswith('text/html'):
                response.render()
                cache.set(cache_key, response.content, timeout=INCIDENTS_PAGE_CACHE_TIMEOUT)
        return response

    def get_queryset(self):
        user = self.request.user
        user_trajectories = UserCourseTrajectory.objects.filter(
            user=user,
            trajectory__isnull=False
        ).select_related('trajectory')
        
        # Проверяем, нужно ли скрывать специализированные траектории
        if self._should_hide_specialized_trajectories(user):
            # Специализированные группы для медсестер/ассистентов
            specialized_groups = [
                "Медицинская сестра/ассистент в хирургии",
                "Медицинская сестра/ассистент в терапии", 
                "Медицинская сестра/ассистент в ортопедии",
                "Медицинская сестра/ассистент в ортодонтии"
            ]
            
            # Исключаем траектории, доступные только для специализированных групп
            user_specialized_groups = user.groups.filter(name__in=specialized_groups)
            if user_specialized_groups.exists():
                specialized_trajectories = Trajectory.objects.filter(
                    groups__in=user_specialized_groups
                )
                
                user_trajectories = user_trajectories.exclude(
                    trajectory__in=specialized_trajectories
                )
        
        return user_trajectories
    
    def _should_hide_specialized_trajectories(self, user):
        """
        Проверяет, нужно ли скрывать специализированные траектории для пользователя.
        Скрывает, если пользователь состоит в группе "Медсестра/ассистент" И в специализированной группе,
        но еще не завершил курс "Внедрение м/с и асс. День 6."
        """
        # Проверяем, состоит ли пользователь в группе "Медсестра/ассистент"
        nurse_assistant_group = user.groups.filter(name="Медсестра/ассистент").first()
        if not nurse_assistant_group:
            return False
        
        # Специализированные группы для медсестер/ассистентов
        specialized_groups = [
            "Медицинская сестра/ассистент в хирургии",
            "Медицинская сестра/ассистент в терапии", 
            "Медицинская сестра/ассистент в ортопедии",
            "Медицинская сестра/ассистент в ортодонтии"
        ]
        
        # Проверяем, состоит ли пользователь в какой-либо из специализированных групп
        user_specialized_groups = user.groups.filter(name__in=specialized_groups)
        if not user_specialized_groups.exists():
            return False
        
        # Проверяем, завершил ли пользователь курс "Внедрение м/с и асс. День 6."
        from myapp.models import UserCourse
        intro_course_completed = UserCourse.objects.filter(
            user=user,
            course__title__icontains="Внедрение м/с и асс. День 6",
            status='completed'
        ).exists()
        
        # Если курс не завершен, скрываем специализированные траектории
        return not intro_course_completed

    def _is_course_available_in_trajectory(self, user, course):
        """
        Проверяет, доступен ли курс в траектории для пользователя.
        Курс доступен, если он первый в траектории или предыдущий завершён.
        """
        # Получаем все траектории пользователя, содержащие этот курс
        user_trajectories = UserCourseTrajectory.objects.filter(
            user=user,
            trajectory__courses=course
        )
        
        for ut in user_trajectories:
            # Получаем порядок курса в траектории
            tc = TrajectoryCourse.objects.filter(
                trajectory=ut.trajectory,
                course=course
            ).first()
            
            if not tc:
                continue
                
            # Если курс первый в траектории, он доступен
            if tc.order == 1:
                return True
                
            # Проверяем, завершен ли предыдущий курс
            prev_tc = TrajectoryCourse.objects.filter(
                trajectory=ut.trajectory,
                order=tc.order - 1
            ).first()
            
            if prev_tc:
                prev_uc = UserCourse.objects.filter(
                    user=user,
                    course=prev_tc.course
                ).first()
                
                if prev_uc and prev_uc.status == 'completed':
                    return True
        
        return False

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user'] = self.request.user
        
        # Получаем прогресс по курсам
        user = self.request.user
        available_courses = Course.objects.available_for_user(user)
        
        # Фильтруем курсы: исключаем курсы из траекторий, которые еще не доступны, и курсы-инциденты
        filtered_courses = []
        for course in available_courses:
            # Исключаем курсы-инциденты
            # if course.is_incident:
            #     continue
            
            # Проверяем, доступен ли курс через группы пользователя напрямую (не через траекторию)
            course_available_via_groups = course.allowed_groups.filter(id__in=user.groups.all()).exists()
            
            # Проверяем, есть ли курс в траекториях пользователя
            course_in_trajectories = TrajectoryCourse.objects.filter(
                trajectory__usercoursetrajectory__user=user,
                course=course
            ).exists()
            
            if course_available_via_groups:
                # Если курс доступен через группы напрямую, он всегда отображается
                filtered_courses.append(course)
            elif course_in_trajectories:
                # Если курс в траектории, проверяем его доступность
                if self._is_course_available_in_trajectory(user, course):
                    filtered_courses.append(course)
            else:
                # Если курс не в траектории и не через группы, он доступен (назначен напрямую пользователю)
                filtered_courses.append(course)
        
        # Получаем UserCourse для каждого отфильтрованного курса
        from myapp.models import ManualCourseUnassignment
        
        user_courses = []
        for course in filtered_courses:
            user_course = UserCourse.objects.filter(user=user, course=course).first()
            if user_course:
                user_courses.append(user_course)
                continue
            # Курсы из траекторий назначаются только после завершения предыдущего — не создаём запись здесь
            if UserCourseTrajectory.objects.filter(user=user, trajectory__trajectorycourse__course=course).exists():
                continue
            # Проверяем, не был ли курс отменен вручную
            manual_unassignment = ManualCourseUnassignment.objects.filter(
                user=user, 
                course=course
            ).first()
            
            if not manual_unassignment:
                user_course, created = UserCourse.objects.get_or_create(
                    user=user, 
                    course=course, 
                    defaults={'status': 'available'}
                )
                user_courses.append(user_course)
            # Если была ручная отмена, просто пропускаем этот курс
        
        # Подготавливаем данные для каждого курса
        courses_data = []
        for user_course in user_courses:
            course = user_course.course
            
            # Подсчет завершенных уроков
            completed_lessons = UserProgress.objects.filter(
                user=user,
                course=course,
                completed=True
            ).count()

            trajectory = UserLessonTrajectory.objects.filter(user=user, course=course).first()
            total_lessons = trajectory.lessons.count() if trajectory else course.lessons.count()
            
            # Подсчет завершенных тестов в рамках этого курса (только уникальные по quiz_title)
            completed_quizzes = QuizResult.objects.filter(
                user=user,
                course=course,
                quiz_title__in=[quiz.name for quiz in course.quizzes.all()],
                passed=True
            ).values('quiz_title').distinct().count()
            total_quizzes = course.quizzes.count()
            
            # Подсчет выполненных заданий в рамках этого курса
            completed_homeworks = HomeworkSubmission.objects.filter(
                user=user,
                course=course,
                homework__in=course.homeworks,
                status='correct'
            ).values('homework_id').distinct().count()
            total_homeworks = course.homeworks.count()
            
            # Общий подсчет материалов (задания считаются как тесты)
            completed_materials = completed_lessons + completed_quizzes + completed_homeworks
            total_materials = total_lessons + total_quizzes + total_homeworks
            
            # Если у курса нет материалов, считаем его доступным
            if total_materials == 0:
                percent = 0
            else:
                percent = int((completed_materials / total_materials) * 100)

            # Проверяем deadline и блокируем курс, если срок истек
            # Для курс-инцидентов блокировка не применяется
            deadline = user_course.deadline
            is_deadline_overdue = False
            if deadline:
                from django.utils import timezone
                is_deadline_overdue = timezone.now() > deadline
                if is_deadline_overdue and user_course.status not in ['completed', 'blocked']:
                    # Проверяем, является ли курс курс-инцидентом
                    if not user_course.course.is_incident:
                        user_course.status = 'blocked'
                        user_course.save(update_fields=['status'])

            # Определяем статус курса
            # Если курс заблокирован, используем статус blocked
            if user_course.status == 'blocked':
                status = 'blocked'
                final_quiz_status = None
                quiz_passed = False
            else:
                final_quiz_status = None
                quiz_passed = None
                if course.final_quiz:
                    quiz_passed = QuizResult.objects.filter(
                        user=user,
                        course=course,
                        quiz_title=course.final_quiz.name,
                        passed=True
                    ).exists()
                    
                    # Получаем статус финального теста (pending/reviewed/completed)
                    latest_final_quiz_result = QuizResult.objects.filter(
                        user=user,
                        course=course,
                        quiz_title=course.final_quiz.name
                    ).order_by('-completed_at').first()
                    
                    if latest_final_quiz_result:
                        final_quiz_status = latest_final_quiz_result.status
                
                    # Получаем статус финального теста (pending/reviewed/completed)
                    latest_final_quiz_result = QuizResult.objects.filter(
                        user=user,
                        course=course,
                        quiz_title=course.final_quiz.name
                    ).order_by('-completed_at').first()
                
                    if latest_final_quiz_result:
                        final_quiz_status = latest_final_quiz_result.status
                
                    # Курс считается завершенным только если все материалы пройдены И финальный тест пройден
                    if total_materials > 0 and completed_materials >= total_materials and quiz_passed:
                        status = 'completed'
                    elif completed_materials > 0 or user_course.status in ['started', 'in_progress']:
                        status = 'in_progress'
                    else:
                        status = 'available'
                else:
                    # Если нет финального теста, курс завершен когда все материалы пройдены
                    if total_materials > 0 and completed_materials >= total_materials:
                        status = 'completed'
                    elif completed_materials > 0 or user_course.status in ['started', 'in_progress']:
                        status = 'in_progress'
                    else:
                        status = 'available'
                
                # Создаем course_data для всех курсов (и с финальным тестом, и без)
                course_data = {
                    'course': course,
                    'user_course': user_course,
                    'completed_lessons': completed_lessons,
                    'completed_quizzes': completed_quizzes,
                    'completed_homeworks': completed_homeworks,
                    'total_lessons': total_lessons,
                    'total_quizzes': total_quizzes,
                    'total_homeworks': total_homeworks,
                    'completed_materials': completed_materials,
                    'total_materials': total_materials,
                    'percent': percent,
                    'status': status,
                    'quiz_passed': quiz_passed if course.final_quiz else None,
                    'final_quiz_status': final_quiz_status,
                    'deadline': deadline,
                    'is_deadline_overdue': is_deadline_overdue
                }
                
                courses_data.append(course_data)
        
        # Сохраняем общие значения ДО фильтрации
        total_courses_all = len(courses_data)
        completed_courses_all = len([c for c in courses_data if c['status'] == 'completed'])
        in_progress_courses_all = len([c for c in courses_data if c['status'] == 'in_progress'])
        available_courses_all = len([c for c in courses_data if c['status'] == 'available'])
        incident_courses_all = len([c for c in courses_data if c['course'].is_incident])
        
        # Поиск по названию курса
        search_query = self.request.GET.get('search', '').strip()
        if search_query:
            courses_data = [course for course in courses_data 
                          if search_query.lower() in course['course'].title.lower()]
        
        # Фильтрация по статусу
        status_filter = self.request.GET.get('status', 'all')
        if status_filter != 'all':
            courses_data = [course for course in courses_data if course['status'] == status_filter]
        
        # Фильтрация по инцидентам
        incident_filter = self.request.GET.get('incident', 'all')
        if incident_filter == 'true':
            courses_data = [course for course in courses_data if course['course'].is_incident]
        elif incident_filter == 'false':
            courses_data = [course for course in courses_data if not course['course'].is_incident]
        
        # Добавляем данные о курсах в контекст
        context.update({
            'courses_data': courses_data,
            'status_filter': status_filter,
            'incident_filter': incident_filter,
            'search_query': search_query,
            'total_courses': len(courses_data),
            'completed_courses': len([c for c in courses_data if c['status'] == 'completed']),
            'in_progress_courses': len([c for c in courses_data if c['status'] == 'in_progress']),
            'available_courses': len([c for c in courses_data if c['status'] == 'available']),
            # Общие значения (не изменяются при фильтрации)
            'total_courses_all': total_courses_all,
            'completed_courses_all': completed_courses_all,
            'in_progress_courses_all': in_progress_courses_all,
            'available_courses_all': available_courses_all,
            'incident_courses_all': incident_courses_all,
        })
        
        return context




def should_hide_specialized_trajectories(user):
    """
    Проверяет, нужно ли скрывать специализированные траектории для пользователя.
    Скрывает, если пользователь состоит в группе "Медсестра/ассистент" И в специализированной группе,
    но еще не завершил курс "Внедрение м/с и асс. День 6."
    """
    nurse_assistant_group = user.groups.filter(name="Медсестра/ассистент").first()
    if not nurse_assistant_group:
        return False
    specialized_groups = [
        "Медицинская сестра/ассистент в хирургии",
        "Медицинская сестра/ассистент в терапии",
        "Медицинская сестра/ассистент в ортопедии",
        "Медицинская сестра/ассистент в ортодонтии"
    ]
    user_specialized_groups = user.groups.filter(name__in=specialized_groups)
    if not user_specialized_groups.exists():
        return False
    intro_course_completed = UserCourse.objects.filter(
        user=user,
        course__title__icontains="Внедрение м/с и асс. День 6",
        status='completed'
    ).exists()
    return not intro_course_completed




def is_course_available_in_trajectory(user, course):
    """
    Проверяет, доступен ли курс в траектории для пользователя.
    Курс доступен, если он первый в траектории или предыдущий завершён.
    """
    user_trajectories = UserCourseTrajectory.objects.filter(
        user=user,
        trajectory__courses=course
    )
    for ut in user_trajectories:
        tc = TrajectoryCourse.objects.filter(
            trajectory=ut.trajectory,
            course=course
        ).first()
        if not tc:
            continue
        if tc.order == 1:
            return True
        prev_tc = TrajectoryCourse.objects.filter(
            trajectory=ut.trajectory,
            order=tc.order - 1
        ).first()
        if prev_tc:
            prev_uc = UserCourse.objects.filter(
                user=user,
                course=prev_tc.course
            ).first()
            if prev_uc and prev_uc.status == 'completed':
                return True
    return False




def get_user_trajectories_queryset(user):
    """Возвращает queryset траекторий пользователя для списка."""
    user_trajectories = UserCourseTrajectory.objects.filter(
        user=user,
        trajectory__isnull=False
    ).select_related('trajectory')
    if should_hide_specialized_trajectories(user):
        specialized_groups = [
            "Медицинская сестра/ассистент в хирургии",
            "Медицинская сестра/ассистент в терапии",
            "Медицинская сестра/ассистент в ортопедии",
            "Медицинская сестра/ассистент в ортодонтии"
        ]
        user_specialized_groups = user.groups.filter(name__in=specialized_groups)
        if user_specialized_groups.exists():
            specialized_trajectories = Trajectory.objects.filter(
                groups__in=user_specialized_groups
            )
            user_trajectories = user_trajectories.exclude(
                trajectory__in=specialized_trajectories
            )
    return user_trajectories


def get_trajectory_list_context(request, skip_filters=False):
    """
    Строит контекст для страницы списка траекторий и курсов.
    Используется представлением и API. Возвращает dict без 'user' и без 'user_trajectories'
    (user_trajectories добавляется из get_queryset в представлении).
    При skip_filters=True (для API) фильтры не применяются, возвращаются все курсы и all-статистика.
    """
    from myapp.models import ManualCourseUnassignment

    user = request.user
    available_courses = Course.objects.available_for_user(user)
    filtered_courses = []
    for course in available_courses:
        course_available_via_groups = course.allowed_groups.filter(id__in=user.groups.all()).exists()
        course_in_trajectories = TrajectoryCourse.objects.filter(
            trajectory__usercoursetrajectory__user=user,
            course=course
        ).exists()
        if course_available_via_groups:
            filtered_courses.append(course)
        elif course_in_trajectories:
            if is_course_available_in_trajectory(user, course):
                filtered_courses.append(course)
        else:
            filtered_courses.append(course)

    user_courses = []
    for course in filtered_courses:
        manual_unassignment = ManualCourseUnassignment.objects.filter(
            user=user,
            course=course
        ).first()
        if not manual_unassignment:
            user_course, _ = UserCourse.objects.get_or_create(
                user=user,
                course=course,
                defaults={'status': 'available'}
            )
            user_courses.append(user_course)

    courses_data = []
    for user_course in user_courses:
        course = user_course.course
        completed_lessons = UserProgress.objects.filter(
            user=user,
            course=course,
            completed=True
        ).count()
        trajectory = UserLessonTrajectory.objects.filter(user=user, course=course).first()
        total_lessons = trajectory.lessons.count() if trajectory else course.lessons.count()
        completed_quizzes = QuizResult.objects.filter(
            user=user,
            course=course,
            quiz_title__in=[quiz.name for quiz in course.quizzes.all()],
            passed=True
        ).values('quiz_title').distinct().count()
        total_quizzes = course.quizzes.count()
        completed_homeworks = HomeworkSubmission.objects.filter(
            user=user,
            course=course,
            homework__in=course.homeworks,
            status='correct'
        ).values('homework_id').distinct().count()
        total_homeworks = course.homeworks.count()
        completed_materials = completed_lessons + completed_quizzes + completed_homeworks
        total_materials = total_lessons + total_quizzes + total_homeworks
        percent = int((completed_materials / total_materials) * 100) if total_materials else 0

        deadline = user_course.deadline
        is_deadline_overdue = False
        if deadline:
            if timezone.now() > deadline and user_course.status not in ['completed', 'blocked']:
                user_course.status = 'blocked'
                user_course.save(update_fields=['status'])
            is_deadline_overdue = timezone.now() > deadline

        if user_course.status == 'blocked':
            status = 'blocked'
            final_quiz_status = None
            quiz_passed = False
        else:
            final_quiz_status = None
            quiz_passed = None
            if course.final_quiz:
                quiz_passed = QuizResult.objects.filter(
                    user=user,
                    course=course,
                    quiz_title=course.final_quiz.name,
                    passed=True
                ).exists()
                latest_final_quiz_result = QuizResult.objects.filter(
                    user=user,
                    course=course,
                    quiz_title=course.final_quiz.name
                ).order_by('-completed_at').first()
                if latest_final_quiz_result:
                    final_quiz_status = latest_final_quiz_result.status
                if total_materials > 0 and completed_materials >= total_materials and quiz_passed:
                    status = 'completed'
                elif completed_materials > 0 or user_course.status in ['started', 'in_progress']:
                    status = 'in_progress'
                else:
                    status = 'available'
            else:
                if total_materials > 0 and completed_materials >= total_materials:
                    status = 'completed'
                elif completed_materials > 0 or user_course.status in ['started', 'in_progress']:
                    status = 'in_progress'
                else:
                    status = 'available'

        course_data = {
            'course': course,
            'user_course': user_course,
            'completed_lessons': completed_lessons,
            'completed_quizzes': completed_quizzes,
            'completed_homeworks': completed_homeworks,
            'total_lessons': total_lessons,
            'total_quizzes': total_quizzes,
            'total_homeworks': total_homeworks,
            'completed_materials': completed_materials,
            'total_materials': total_materials,
            'percent': percent,
            'status': status,
            'quiz_passed': quiz_passed if course.final_quiz else None,
            'final_quiz_status': final_quiz_status,
            'deadline': deadline,
            'is_deadline_overdue': is_deadline_overdue
        }
        courses_data.append(course_data)

    total_courses_all = len(courses_data)
    completed_courses_all = len([c for c in courses_data if c['status'] == 'completed'])
    in_progress_courses_all = len([c for c in courses_data if c['status'] == 'in_progress'])
    available_courses_all = len([c for c in courses_data if c['status'] == 'available'])
    incident_courses_all = len([c for c in courses_data if c['course'].is_incident])

    if not skip_filters:
        search_query = request.GET.get('search', '').strip()
        if search_query:
            courses_data = [c for c in courses_data if search_query.lower() in c['course'].title.lower()]
        status_filter = request.GET.get('status', 'all')
        if status_filter != 'all':
            courses_data = [c for c in courses_data if c['status'] == status_filter]
        incident_filter = request.GET.get('incident', 'all')
        if incident_filter == 'true':
            courses_data = [c for c in courses_data if c['course'].is_incident]
        elif incident_filter == 'false':
            courses_data = [c for c in courses_data if not c['course'].is_incident]
    else:
        search_query = ''
        status_filter = 'all'
        incident_filter = 'all'

    return {
        'courses_data': courses_data,
        'status_filter': status_filter,
        'incident_filter': incident_filter,
        'search_query': search_query,
        'total_courses': len(courses_data),
        'completed_courses': len([c for c in courses_data if c['status'] == 'completed']),
        'in_progress_courses': len([c for c in courses_data if c['status'] == 'in_progress']),
        'available_courses': len([c for c in courses_data if c['status'] == 'available']),
        'total_courses_all': total_courses_all,
        'completed_courses_all': completed_courses_all,
        'in_progress_courses_all': in_progress_courses_all,
        'available_courses_all': available_courses_all,
        'incident_courses_all': incident_courses_all,
    }




class IncidentCoursesListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    Список всех курсов-инцидентов для администраторов.
    """
    model = Course
    template_name = 'courses/incident_courses_list.html'
    context_object_name = 'courses'
    paginate_by = 20

    def test_func(self):
        """Доступ только для администраторов."""
        return is_admin(self.request.user)

    def get_queryset(self):
        """Возвращает только курсы-инциденты."""
        return Course.objects.filter(is_incident=True).order_by('-created_at')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        courses = self.get_queryset()
        
        # Поиск по названию курса
        search_query = self.request.GET.get('search', '').strip()
        if search_query:
            courses = courses.filter(title__icontains=search_query)
        
        # Фильтрация по статусу (не используется для инцидентов, но для совместимости)
        status_filter = self.request.GET.get('status', 'all')
        
        # Подготавливаем данные для каждого курса
        courses_data = []
        for course in courses:
            courses_data.append({
                'course': course,
                'author': course.author.get_full_name() or course.author.username,
                'created_at': course.created_at,
                'lessons_count': course.lessons.count(),
                'quizzes_count': course.quizzes.count(),
                'total_materials': course.lessons.count() + course.quizzes.count(),
            })
        
        context.update({
            'courses_data': courses_data,
            'status_filter': status_filter,
            'search_query': search_query,
            'total_courses': len(courses_data),
        })
        
        return context


class TrajectoryCreateView(UserPassesTestMixin, CreateView):
    """
    Создание новой траектории курсов.
    """
    model = Trajectory
    fields = ['name', 'description', 'groups', 'certificate']
    template_name = 'courses/create_trajectory.html'

    def test_func(self):
        return self.request.user.is_staff or self.request.user.is_superuser

    def get_success_url(self):
        return reverse('builder:trajectory_courses', kwargs={'trajectory_id': self.object.id})

    def form_valid(self, form):
        response = super().form_valid(form)
        if self.request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'success': True, 'id': self.object.id, 'name': self.object.name})
        return response




@method_decorator(login_required, name='dispatch')
class CertificateListView(TemplateView):
    """
    Представление для отображения сертификатов пользователя.
    """
    template_name = 'courses/user_certificates.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        certificates = get_user_certificates(self.request.user)
        context.update(certificates)
        return context




class ViewCertificatePdfView(TemplateView):
    template_name = 'courses/certificate_pdf.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        certificate = get_object_or_404(Certificate, certificate_id=kwargs['certificate_id'], user=self.request.user)
        context['certificate'] = certificate
        return context




@login_required
def download_certificate_pdf(request, certificate_id):
    """
    Скачивание сертификата в формате PDF.
    """
    # import os
    # BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(file)))
    # STATIC_ROOT = os.path.join(BASE_DIR, 'static')
    # Получаем сертификат и проверяем права доступа
    certificate = get_object_or_404(Certificate, certificate_id=certificate_id, user=request.user)
    
    # Генерируем HTML из шаблона
    html_string = render_to_string('courses/certificate_pdf.html', {
        'certificate': certificate,
        'generated_at': datetime.now(),
    })
    from myproject.settings import base as base_settings
    # import os
    # Создаем PDF с помощью WeasyPrint
    # static_dir = os.path.join(base_settings.BASE_DIR, 'static')
    # bg_path = os.path.join(static_dir, 'courses', 'imgs', 'cert_srcs', 'background.png')
    from django.conf import settings
    import os

    from django.contrib.staticfiles import finders
    static_path = 'courses/imgs/cert_srcs/background.png'
    found_path = finders.find(static_path)

    if found_path:
        base_dir = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(found_path))))
    else:
        base_dir = base_settings.STATIC_ROOT

    html = HTML(string=html_string, base_url=base_dir)
    pdf = html.write_pdf()
    
    # Формируем имя файла
    if certificate.certificate_type == 'course':
        filename = f"certificate_course_{certificate.course.slug}_{certificate.certificate_id}.pdf"
    else:
        filename = f"certificate_trajectory_{certificate.trajectory.id}_{certificate.certificate_id}.pdf"
    
    # Возвращаем PDF как HttpResponse
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Логируем скачивание
    logger.info(
        f'Скачан сертификат {certificate.certificate_id} пользователем {request.user.username}',
        extra={'user': request.user.username, 'certificate_id': certificate.certificate_id}
    )
    
    return response
    



class MetricsFormView(LoginRequiredMixin, UserPassesTestMixin, View):
    """
    CBV для формы "передача данных – Метрики эффективности стомклиники"
    Доступ только для пользователей группы "Внешний пользователь" или superuser/is_staff
    """
    template_name = 'courses/metrics_form.html'
    
    def test_func(self):
        """Проверка доступа: superuser/is_staff или группа "Внешний пользователь" """
        user = self.request.user
        if user.is_superuser or user.is_staff:
            return True
        return user.groups.filter(name='Внешний пользователь').exists()
    
    def get(self, request, *args, **kwargs):
        """GET запрос - отображение формы"""
        user_country = ''
        if hasattr(request.user, 'profile') and request.user.profile:
            user_country = request.user.profile.country or ''
        
        context = {
            'title': 'передача данных – Метрики эффективности стомклиники',
            'user_country': user_country
        }
        return render(request, self.template_name, context)
    
    def post(self, request, *args, **kwargs):
        """POST запрос - обработка AJAX данных"""
        import json
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            logger.error('Invalid JSON in metrics form submission')
            return JsonResponse({'success': False, 'error': 'Invalid JSON'}, status=400)
        
        try:
            # Сохраняем данные в базу
            from .models import MetricsSubmission
            
            # Извлекаем основные данные с безопасным преобразованием типов
            clinic_name = data.get('clinicName', '').strip()
            if not clinic_name:
                return JsonResponse({'success': False, 'error': 'Название клиники обязательно'}, status=400)
            
            initial_month = data.get('startMonth', '').strip()
            if not initial_month:
                return JsonResponse({'success': False, 'error': 'Начальный месяц обязателен'}, status=400)
            
            try:
                doctors_count = int(data.get('docCount', 1))
                if doctors_count < 1:
                    doctors_count = 1
            except (ValueError, TypeError):
                doctors_count = 1
            
            try:
                chairs_count = int(data.get('chairs', 0))
                if chairs_count < 0:
                    chairs_count = 0
            except (ValueError, TypeError):
                chairs_count = 0
            
            try:
                hours_weekdays = float(data.get('hoursWeekdays', 0))
                if hours_weekdays < 0:
                    hours_weekdays = 0
            except (ValueError, TypeError):
                hours_weekdays = 0
            
            try:
                hours_saturday = float(data.get('hoursSaturday', 0))
                if hours_saturday < 0:
                    hours_saturday = 0
            except (ValueError, TypeError):
                hours_saturday = 0
            
            try:
                hours_sunday = float(data.get('hoursSunday', 0))
                if hours_sunday < 0:
                    hours_sunday = 0
            except (ValueError, TypeError):
                hours_sunday = 0
            
            # Дни в месяце
            days = data.get('days', [])
            if not isinstance(days, list):
                days = []
            
            # Валюта
            currency = data.get('currency', 'rub').strip()
            if currency not in ['rub', 'kzt']:
                currency = 'rub'
            
            # Данные врачей
            doctors_data = {
                'doctors': data.get('doctors', []),
                'months': data.get('months', [])
            }
            
            # Создаем запись
            submission = MetricsSubmission.objects.create(
                user=request.user,
                clinic_name=clinic_name,
                initial_month=initial_month,
                doctors_count=doctors_count,
                chairs_count=chairs_count,
                hours_weekdays=hours_weekdays,
                hours_saturday=hours_saturday,
                hours_sunday=hours_sunday,
                days_month_1=days[0] if len(days) > 0 else 0,
                days_month_2=days[1] if len(days) > 1 else 0,
                days_month_3=days[2] if len(days) > 2 else 0,
                days_month_4=days[3] if len(days) > 3 else 0,
                days_month_5=days[4] if len(days) > 4 else 0,
                days_month_6=days[5] if len(days) > 5 else 0,
                currency=currency,
                doctors_data=doctors_data
            )
            
            logger.info(f'Metrics submission created successfully: ID {submission.id}, user {request.user.username}')
            return JsonResponse({'success': True})
            
        except Exception as e:
            logger.error(f'Error processing metrics form submission: {str(e)}', exc_info=True)
            return JsonResponse({
                'success': False, 
                'error': f'Ошибка при сохранении данных: {str(e)}'
            }, status=500)




class MetricsSuccessView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """
    CBV для страницы успешной отправки формы метрик
    Доступ только для пользователей группы "Внешний пользователь" или superuser/is_staff
    """
    template_name = 'courses/metrics_success.html'
    
    def test_func(self):
        """Проверка доступа: superuser/is_staff или группа "Внешний пользователь" """
        user = self.request.user
        if user.is_superuser or user.is_staff:
            return True
        return user.groups.filter(name='Внешний пользователь').exists()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Форма отправлена успешно'
        return context




class MetricsAdminListView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    CBV для административной страницы со списком всех заполненных форм метрик
    """
    template_name = 'courses/metrics_admin_list.html'
    context_object_name = 'submissions'
    
    def test_func(self):
        """Проверка доступа: только superuser и staff"""
        return self.request.user.is_superuser or self.request.user.is_staff
    
    def get_queryset(self):
        from .models import MetricsSubmission
        return MetricsSubmission.objects.select_related('user').all()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Администрирование форм метрик'
        return context




class MetricsAdminDetailView(LoginRequiredMixin, UserPassesTestMixin, DetailView):
    """
    CBV для детального просмотра заполненной формы метрик
    """
    template_name = 'courses/metrics_admin_detail.html'
    context_object_name = 'submission'
    pk_url_kwarg = 'submission_id'
    
    def test_func(self):
        """Проверка доступа: только superuser"""
        return self.request.user.is_superuser or self.request.user.is_staff
    
    def get_queryset(self):
        from .models import MetricsSubmission
        return MetricsSubmission.objects.all()
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Метрики {self.object.clinic_name}'
        return context


@login_required
def export_metrics_to_excel(request, submission_id):
    """
    Экспорт формы метрик в Excel формат согласно шаблону Google Sheets
    """
    # Проверяем права доступа
    if not (request.user.is_superuser or request.user.is_staff):
        return HttpResponseForbidden("Доступ запрещен")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from .models import MetricsSubmission
        from datetime import datetime
        import calendar
        
        # Получаем данные формы
        submission = MetricsSubmission.objects.get(id=submission_id)
        
        # Создаем workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Метрики"
        
        # Определяем стили
        header_font = Font(bold=True, size=11)
        normal_font = Font(size=10)
        fill_yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        fill_light_blue = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
        
        # Убираем общие границы - будем добавлять только для отдельных таблиц
        # border = Border(
        #     left=Side(border_style="thin"),
        #     right=Side(border_style="thin"),
        #     top=Side(border_style="thin"),
        #     bottom=Side(border_style="thin")
        # )
        
        # Пунктирная граница для таблиц
        dashed_border = Border(
            left=Side(border_style="dashed"),
            right=Side(border_style="dashed"),
            top=Side(border_style="dashed"),
            bottom=Side(border_style="dashed")
        )
        
        # Функция для получения названий месяцев на русском (только первые 3 месяца)
        def get_month_names_ru(initial_month_str):
            year, month = map(int, initial_month_str.split('-'))
            month_names_ru = [
                "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь",
                "Июль", "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"
            ]
            months = []
            for i in range(3):  # Только первые 3 месяца
                current_month = month + i - 1
                current_year = year
                if current_month >= 12:
                    current_month -= 12
                    current_year += 1
                months.append(f"{month_names_ru[current_month]} {current_year}")
            return months
        
        month_names = get_month_names_ru(submission.initial_month)
        doctors_data = submission.doctors_data.get('doctors', [])
        months_data = submission.doctors_data.get('months', [])
        
        # Заполняем заголовок
        # ws['A1'] = "Правила заполнения:"
        # ws['A1'].font = header_font
        # ws['A2'] = "Ячейки только этого цвета: К ЗАПОЛНЕНИЮ"
        # ws['A2'].fill = fill_yellow
        
        # Основные параметры клиники (строка 4)
        ws['A4'] = "Кол-во кресел, загрузка клиники"
        ws['A4'].font = header_font
        
        # Строка 5 - числовые данные (начинаем с C)
        # Формула: количество кресел * среднее количество часов работы в месяц
        # Расчет: (будни * 22 + суббота * 4 + воскресенье * 4) / 30 * кресла * 30
        # Упрощенно: будни * 22 + суббота * 4 + воскресенье * 4
        monthly_hours = submission.hours_weekdays * 22 + submission.hours_saturday * 4 + submission.hours_sunday * 4
        ws['A5'] = f"={submission.chairs_count}*{monthly_hours}"
        
        # Загрузка клиники по месяцам (формулы, только первые 3 месяца)
        # Нужно будет добавить формулы после создания итоговых строк
        ws['D5'] = "5%"
        ws['E5'] = "0%"
        ws['F5'] = "0%"
        
        # Строка 6 - заголовки месяцев (начинаем с D)
        ws['A6'] = "Месяц"
        for i, month_name in enumerate(month_names):
            ws[f'{chr(68+i)}6'] = month_name  # D6, E6, F6, etc.
        
        # Строка 7 - заголовок таблицы часов по графику
        ws['A7'] = "Кол-во часов по графику, Т,раб"
        ws['B7'] = "ФИО"
        ws['C7'] = "Специализация"
        
        # Рабочие дни по месяцам в строке 7 (начинаем с D, только первые 3 месяца)
        days_list = [
            submission.days_month_1, submission.days_month_2, submission.days_month_3
        ]
        for i, days in enumerate(days_list):
            ws[f'{chr(68+i)}7'] = days
        
        # Карта специализаций
        spec_map = {
            "hygienist": "Гигиенист",
            "implantologist": "Имплантолог", 
            "orthodontist": "Ортодонт",
            "orthopedist": "Ортопед",
            "periodontist": "Пародонтолог",
            "therapist": "Терапевт",
            "surgeon": "Хирург",
            "therapist_surgeon": "Терапевт-Хирург",
            "orthopedist_surgeon": "Ортопед-Хирург",
            "universal": "Универсал",
            "custom": "Свой вариант"
        }
        
        # Начинаем заполнение данных врачей с 8 строки
        current_row = 8
        schedule_hours_start_row = current_row
        
        # БЛОК 1: Часы по графику (строки 8-26)
        # Используем реальное количество врачей из формы
        max_doctors = len(doctors_data) if doctors_data else 0
        for i in range(max_doctors):
            ws[f'A{current_row}'] = i + 1
            
            if i < len(doctors_data):
                doctor = doctors_data[i]
                ws[f'B{current_row}'] = doctor.get('name', f'ФИО доктора {i+1}')
                ws[f'C{current_row}'] = spec_map.get(doctor.get('specialization', ''), 'заполнить специализацию доктора')
                
                # Часы по графику по месяцам (начинаем с D, только первые 3 месяца)
                for month_idx in range(3):
                    if month_idx < len(months_data) and i < len(months_data[month_idx].get('doctors', [])):
                        schedule_hours = months_data[month_idx]['doctors'][i].get('scheduleHours', 0)
                        ws[f'{chr(68+month_idx)}{current_row}'] = schedule_hours or 0
                    else:
                        ws[f'{chr(68+month_idx)}{current_row}'] = 0
            else:
                ws[f'B{current_row}'] = "ФИО"
                ws[f'C{current_row}'] = "заполнить"
                for month_idx in range(3):
                    ws[f'{chr(68+month_idx)}{current_row}'] = "заполнить"
            
            # Добавляем формулу среднего значения в столбец H для каждой строки врача
            ws[f'H{current_row}'] = f"=(D{current_row}+E{current_row}+F{current_row})/3"
            ws[f'H{current_row}'].number_format = '0.00'
                    
            current_row += 1
        
        
        # Пропускаем 2 строки
        current_row += 2
        
        # БЛОК 2: Часы с пациентами (строки 29-47)
        ws[f'A{current_row}'] = "Кол-во часов с пациентами, Т,заг"
        ws[f'B{current_row}'] = "ФИО"
        ws[f'C{current_row}'] = "Специализация"
        current_row += 1
        patient_hours_start_row = current_row
        
        for i in range(max_doctors):
            ws[f'A{current_row}'] = i + 1
            
            if i < len(doctors_data):
                doctor = doctors_data[i]
                ws[f'B{current_row}'] = doctor.get('name', f'ФИО доктора {i+1}')
                ws[f'C{current_row}'] = spec_map.get(doctor.get('specialization', ''), 'заполнить специализацию доктора')
                
                # Часы с пациентами по месяцам (начинаем с D)
                for month_idx in range(3):
                    if month_idx < len(months_data) and i < len(months_data[month_idx].get('doctors', [])):
                        patient_hours = months_data[month_idx]['doctors'][i].get('patientHours', 0)
                        ws[f'{chr(68+month_idx)}{current_row}'] = patient_hours or 0
                    else:
                        ws[f'{chr(68+month_idx)}{current_row}'] = 0
            else:
                ws[f'B{current_row}'] = "ФИО"
                ws[f'C{current_row}'] = "заполнить"
                for month_idx in range(3):
                    ws[f'{chr(68+month_idx)}{current_row}'] = "заполнить"
            
            # Добавляем формулу среднего значения в столбец H для каждой строки врача
            ws[f'H{current_row}'] = f"=(D{current_row}+E{current_row}+F{current_row})/3"
            ws[f'H{current_row}'].number_format = '0.00'
                    
            current_row += 1
        
        # Вычисляем позиции для таблиц
        schedule_table_start = 7
        schedule_table_end = schedule_table_start + max_doctors + 1  # +1 для заголовка
        
        # Вычисляем позиции для таблицы "Часы с пациентами"
        patient_hours_table_start = schedule_table_end + 2 + 1  # +1 для заголовка
        patient_hours_table_end = patient_hours_table_start + max_doctors
        
        # Добавляем суммарные значения для блока 2 (Часы с пациентами) в строку заголовка
        patient_header_row = patient_hours_table_start - 1  # поднимаем на одну строку выше
        for month_idx in range(3):
            # Создаем формулу с операторами плюс для каждой ячейки врача
            cell_formula_parts = []
            for i in range(max_doctors):
                cell_address = f'{chr(68+month_idx)}{patient_hours_start_row + i}'
                cell_formula_parts.append(cell_address)
            formula = f"=({' + '.join(cell_formula_parts)})"
            ws[f'{chr(68+month_idx)}{patient_header_row}'] = formula
        
        # Добавляем формулу среднего значения в столбец H для таблицы "Кол-во часов с пациентами"
        ws[f'H{patient_header_row}'] = "=AVERAGE(D{0}:F{0})".format(patient_header_row)
        ws[f'H{patient_header_row}'].number_format = '0.00'
        
        # Пропускаем 2 строки
        current_row += 2
        
        # БЛОК 3: Выручка (строки 50-68)
        ws[f'A{current_row}'] = "Выручка, ВВ"
        ws[f'B{current_row}'] = "ФИО"
        ws[f'C{current_row}'] = "Специализация"
        current_row += 1
        revenue_start_row = current_row
        
        for i in range(max_doctors):
            ws[f'A{current_row}'] = i + 1
            
            if i < len(doctors_data):
                doctor = doctors_data[i]
                ws[f'B{current_row}'] = doctor.get('name', f'ФИО доктора {i+1}')
                ws[f'C{current_row}'] = spec_map.get(doctor.get('specialization', ''), 'заполнить специализацию доктора')
                
                # Выручка по месяцам (начинаем с D)
                for month_idx in range(3):
                    if month_idx < len(months_data) and i < len(months_data[month_idx].get('doctors', [])):
                        revenue = months_data[month_idx]['doctors'][i].get('revenue', 0)
                        ws[f'{chr(68+month_idx)}{current_row}'] = revenue or 0
                    else:
                        ws[f'{chr(68+month_idx)}{current_row}'] = 0
            else:
                ws[f'B{current_row}'] = "ФИО"
                ws[f'C{current_row}'] = "заполнить"
                for month_idx in range(3):
                    ws[f'{chr(68+month_idx)}{current_row}'] = "заполнить"
            
            # Добавляем формулу среднего значения в столбец H для каждой строки врача
            ws[f'H{current_row}'] = f"=(D{current_row}+E{current_row}+F{current_row})/3"
            ws[f'H{current_row}'].number_format = '0.00'
                    
            current_row += 1
        
        # Вычисляем позиции для таблицы "Выручка"
        revenue_table_start = patient_hours_table_end + 2 + 1  # +1 для заголовка
        revenue_table_end = revenue_table_start + max_doctors
        
        # Добавляем суммарные значения для блока 3 (Выручка) в строку заголовка
        revenue_header_row = revenue_table_start - 1  # поднимаем на одну строку выше
        for month_idx in range(3):
            # Создаем формулу с операторами плюс для каждой ячейки врача
            cell_formula_parts = []
            for i in range(max_doctors):
                cell_address = f'{chr(68+month_idx)}{revenue_start_row + i}'
                cell_formula_parts.append(cell_address)
            formula = f"=({' + '.join(cell_formula_parts)})"
            ws[f'{chr(68+month_idx)}{revenue_header_row}'] = formula
        
        # Добавляем формулу среднего значения в столбец H для таблицы "Выручка, ВВ"
        ws[f'H{revenue_header_row}'] = "=AVERAGE(D{0}:F{0})".format(revenue_header_row)
        ws[f'H{revenue_header_row}'].number_format = '0.00'
        
        # Пропускаем 2 строки
        current_row += 2
        
        # БЛОК 4: Загрузка доктора пациентами (формулы) - строки 73-91
        ws[f'A{current_row}'] = "Загрузка доктора пациентами"
        ws[f'B{current_row}'] = "ФИО"
        ws[f'C{current_row}'] = "Специализация"
        current_row += 1
        
        for i in range(max_doctors):
            ws[f'A{current_row}'] = i + 1
            
            if i < len(doctors_data):
                doctor = doctors_data[i]
                ws[f'B{current_row}'] = doctor.get('name', f'ФИО доктора {i+1}')
                ws[f'C{current_row}'] = spec_map.get(doctor.get('specialization', ''), 'заполнить специализацию доктора')
                
                # Формулы для расчета загрузки (часы с пациентами / часы по графику) (начинаем с D)
                for month_idx in range(3):
                    patient_cell = f'{chr(68+month_idx)}{patient_hours_start_row + i}'
                    schedule_cell = f'{chr(68+month_idx)}{schedule_hours_start_row + i}'
                    ws[f'{chr(68+month_idx)}{current_row}'] = f'=IF({schedule_cell}=0,0,{patient_cell}/{schedule_cell})'
            else:
                ws[f'B{current_row}'] = "ФИО"
                ws[f'C{current_row}'] = "заполнить"
                for month_idx in range(3):
                    ws[f'{chr(68+month_idx)}{current_row}'] = "#VALUE!"
            
            # Добавляем формулу в столбец H для каждой строки врача
            # Формула: соответствующая строка из столбца H таблицы "Кол-во часов с пациентами" / соответствующая строка из столбца H таблицы "Кол-во часов по графику"
            patient_h_row = patient_hours_start_row + i
            schedule_h_row = schedule_hours_start_row + i
            ws[f'H{current_row}'] = f"=IF(H{schedule_h_row}=0,0,H{patient_h_row}/H{schedule_h_row})"
            ws[f'H{current_row}'].number_format = '0%'
                    
            current_row += 1
        
        
        # Применяем процентный формат к блоку 4 (Загрузка доктора пациентами)
        for i in range(max_doctors):
            row_num = current_row - max_doctors + i
            ws[f'D{row_num}'].number_format = '0%'
            ws[f'E{row_num}'].number_format = '0%'
            ws[f'F{row_num}'].number_format = '0%'
        
        # Добавляем суммарные значения для блока 4 (Загрузка доктора пациентами) в строку заголовка
        load_header_row = current_row - max_doctors - 1  # строка заголовка таблицы
        for month_idx in range(3):
            # Формула: закрашенная ячейка из таблицы "Кол-во часов с пациентами" / закрашенная ячейка из таблицы "Кол-во часов по графику"
            patient_sum_cell = f'{chr(68+month_idx)}{patient_header_row}'  # суммарная ячейка из таблицы "Кол-во часов с пациентами"
            schedule_sum_cell = f'{chr(68+month_idx)}7'  # суммарная ячейка из таблицы "Кол-во часов по графику" (строка 7)
            formula = f"=IF({schedule_sum_cell}=0,0,{patient_sum_cell}/{schedule_sum_cell})"
            ws[f'{chr(68+month_idx)}{load_header_row}'] = formula
            # Применяем процентный формат к суммарным ячейкам
            ws[f'{chr(68+month_idx)}{load_header_row}'].number_format = '0%'
        
        # Добавляем формулу среднего значения в столбец H для таблицы "Загрузка доктора пациентами"
        # Формула: среднее значение из таблицы "Кол-во часов с пациентами" / среднее значение из таблицы "Кол-во часов по графику" (H7)
        patient_avg_cell = f'H{patient_header_row}'  # среднее из таблицы "Кол-во часов с пациентами"
        schedule_avg_cell = 'H7'  # среднее из таблицы "Кол-во часов по графику"
        ws[f'H{load_header_row}'] = f"=IF({schedule_avg_cell}=0,0,{patient_avg_cell}/{schedule_avg_cell})"
        ws[f'H{load_header_row}'].number_format = '0%'
        
        # Пропускаем 2 строки
        current_row += 2
        
        # БЛОК 5: Средний час (формулы) - строки 95-113
        ws[f'A{current_row}'] = "Средний час"
        ws[f'B{current_row}'] = "ФИО"
        ws[f'C{current_row}'] = "Специализация"
        current_row += 1
        
        for i in range(max_doctors):
            ws[f'A{current_row}'] = i + 1
            
            if i < len(doctors_data):
                doctor = doctors_data[i]
                ws[f'B{current_row}'] = doctor.get('name', f'ФИО доктора {i+1}')
                ws[f'C{current_row}'] = spec_map.get(doctor.get('specialization', ''), 'заполнить специализацию доктора')
                
                # Формулы для расчета среднего часа (выручка / часы с пациентами) (начинаем с D)
                for month_idx in range(3):
                    revenue_cell = f'{chr(68+month_idx)}{revenue_start_row + i}'
                    patient_cell = f'{chr(68+month_idx)}{patient_hours_start_row + i}'
                    ws[f'{chr(68+month_idx)}{current_row}'] = f'=IF({patient_cell}=0,0,{revenue_cell}/{patient_cell})'
            else:
                ws[f'B{current_row}'] = "ФИО"
                ws[f'C{current_row}'] = "заполнить"
                for month_idx in range(3):
                    ws[f'{chr(68+month_idx)}{current_row}'] = "#VALUE!"
            
            # Добавляем формулу в столбец H для каждой строки врача
            # Формула: соответствующая строка из столбца H таблицы "Выручка" / соответствующая строка из столбца H таблицы "Кол-во часов с пациентами"
            revenue_h_row = revenue_start_row + i
            patient_h_row = patient_hours_start_row + i
            ws[f'H{current_row}'] = f"=IF(H{patient_h_row}=0,0,H{revenue_h_row}/H{patient_h_row})"
            ws[f'H{current_row}'].number_format = '0.00'
                    
            current_row += 1
        
        
        # Применяем формат с 2 знаками после запятой к блоку 5 (Средний час)
        for i in range(max_doctors):
            row_num = current_row - max_doctors + i
            ws[f'D{row_num}'].number_format = '0.00'
            ws[f'E{row_num}'].number_format = '0.00'
            ws[f'F{row_num}'].number_format = '0.00'
        
        # Добавляем суммарные значения для блока 5 (Средний час) в строку заголовка
        avg_hour_header_row = current_row - max_doctors - 1  # строка заголовка таблицы
        for month_idx in range(3):
            # Формула: закрашенная ячейка из таблицы "Выручка" / закрашенная ячейка из таблицы "Кол-во часов с пациентами"
            revenue_sum_cell = f'{chr(68+month_idx)}{revenue_header_row}'  # суммарная ячейка из таблицы "Выручка"
            patient_sum_cell = f'{chr(68+month_idx)}{patient_header_row}'  # суммарная ячейка из таблицы "Кол-во часов с пациентами"
            formula = f"=IF({patient_sum_cell}=0,0,{revenue_sum_cell}/{patient_sum_cell})"
            ws[f'{chr(68+month_idx)}{avg_hour_header_row}'] = formula
            # Применяем формат с 2 знаками после запятой к суммарным ячейкам
            ws[f'{chr(68+month_idx)}{avg_hour_header_row}'].number_format = '0.00'
        
        # Добавляем формулу среднего значения в столбец H для таблицы "Средний час"
        ws[f'H{avg_hour_header_row}'] = "=AVERAGE(D{0}:F{0})".format(avg_hour_header_row)
        ws[f'H{avg_hour_header_row}'].number_format = '0.00'
        
        # Обновляем формулы загрузки клиники по месяцам (только первые 3 месяца)
        for month_idx in range(3):
            # Загрузка клиники = D8+D9+...+DN/A5, где N зависит от количества врачей
            max_hours_cell = 'A5'  # ячейка с максимальными часами
            # Создаем формулу с операторами плюс для каждой ячейки врача
            cell_formula_parts = []
            for i in range(max_doctors):
                cell_address = f'{chr(68+month_idx)}{schedule_hours_start_row + i}'
                cell_formula_parts.append(cell_address)
            formula = f"=({' + '.join(cell_formula_parts)})/{max_hours_cell}"
            ws[f'{chr(68+month_idx)}5'] = formula
        
        # Применяем стили (без границ)
        for row in ws.iter_rows():
            for cell in row:
                # cell.border = border  # Убираем общие границы
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Устанавливаем процентный формат для ячеек загрузки клиники
        from openpyxl.styles import NamedStyle
        percentage_style = NamedStyle(name="percentage")
        percentage_style.number_format = '0%'
        
        # Применяем процентный формат к ячейкам D5, E5, F5
        ws['D5'].number_format = '0%'
        ws['E5'].number_format = '0%'
        ws['F5'].number_format = '0%'
        
        # Вычисляем динамические позиции для границ таблиц в зависимости от количества врачей
        
        # 1. Таблица "Кол-во кресел, загрузка клиники" (A6:H7) - фиксированная позиция
        for row in range(6, 8):  # строки 6-7
            for col in range(1, 9):  # колонки A-H
                cell = ws.cell(row=row, column=col)
                cell.border = dashed_border
        
        # 2. Таблица "Кол-во часов по графику" - поднимаем на одну строку выше
        for row in range(schedule_table_start - 1, schedule_table_end + 1):
            for col in range(1, 9):  # колонки A-H
                cell = ws.cell(row=row, column=col)
                cell.border = dashed_border
        
        # 3. Таблица "Кол-во часов с пациентами" - поднимаем на одну строку выше
        for row in range(patient_hours_table_start - 1, patient_hours_table_end + 1):
            for col in range(1, 9):  # колонки A-H
                cell = ws.cell(row=row, column=col)
                cell.border = dashed_border
        
        # 4. Таблица "Выручка" - поднимаем на одну строку выше
        for row in range(revenue_table_start - 1, revenue_table_end + 1):
            for col in range(1, 9):  # колонки A-H
                cell = ws.cell(row=row, column=col)
                cell.border = dashed_border
        
        # 5. Таблица "Загрузка доктора пациентами" - поднимаем на одну строку выше
        load_table_start = revenue_table_end + 2 + 1  # +1 для заголовка
        load_table_end = load_table_start + max_doctors
        for row in range(load_table_start - 1, load_table_end + 1):
            for col in range(1, 9):  # колонки A-H
                cell = ws.cell(row=row, column=col)
                cell.border = dashed_border
        
        # 6. Таблица "Средний час" - поднимаем на одну строку выше
        avg_hour_table_start = load_table_end + 2 + 1  # +1 для заголовка
        avg_hour_table_end = avg_hour_table_start + max_doctors
        for row in range(avg_hour_table_start - 1, avg_hour_table_end + 1):
            for col in range(1, 9):  # колонки A-H
                cell = ws.cell(row=row, column=col)
                cell.border = dashed_border
        
        # Красим ячейки в 7-й строке (кроме A и G) в голубой цвет
        for col in range(2, 9):  # колонки B, C, D, E, F, G, H
            if col != 7:  # пропускаем G (7-я колонка)
                cell = ws.cell(row=7, column=col)
                cell.fill = fill_light_blue
        
        # Красим суммарные ячейки в голубой цвет (динамические позиции)
        # Ячейки для "Часы с пациентами" - в строке заголовка таблицы
        patient_header_row = patient_hours_table_start - 1  # поднимаем на одну строку выше
        for col in range(2, 9):  # колонки B, C, D, E, F, H (пропускаем A и G)
            if col != 7:  # пропускаем G (7-я колонка)
                ws.cell(row=patient_header_row, column=col).fill = fill_light_blue
        
        # Ячейки для "Выручка" - в строке заголовка таблицы
        revenue_header_row = revenue_table_start - 1  # поднимаем на одну строку выше
        for col in range(2, 9):  # колонки B, C, D, E, F, H (пропускаем A и G)
            if col != 7:  # пропускаем G (7-я колонка)
                ws.cell(row=revenue_header_row, column=col).fill = fill_light_blue
        
        # Ячейки для "Загрузка доктора пациентами" - в строке заголовка таблицы
        load_header_row = load_table_start - 1  # поднимаем на одну строку выше
        for col in range(2, 9):  # колонки B, C, D, E, F, H (пропускаем A и G)
            if col != 7:  # пропускаем G (7-я колонка)
                ws.cell(row=load_header_row, column=col).fill = fill_light_blue
        
        # Ячейки для "Средний час" - в строке заголовка таблицы
        avg_hour_header_row = avg_hour_table_start - 1  # поднимаем на одну строку выше
        for col in range(2, 9):  # колонки B, C, D, E, F, H (пропускаем A и G)
            if col != 7:  # пропускаем G (7-я колонка)
                ws.cell(row=avg_hour_header_row, column=col).fill = fill_light_blue
        
        
        # Добавляем формулы для ячеек D7, E7, F7 - сумма часов по графику
        for month_idx in range(3):
            # Создаем формулу с операторами плюс для каждой ячейки врача
            cell_formula_parts = []
            for i in range(max_doctors):
                cell_address = f'{chr(68+month_idx)}{schedule_hours_start_row + i}'
                cell_formula_parts.append(cell_address)
            formula = f"=({' + '.join(cell_formula_parts)})"
            ws[f'{chr(68+month_idx)}7'] = formula

            ws['H7'] = "=(D7+E7+F7)/3"
            ws['H7'].number_format = '0.00'
            
            # Добавляем формулу в ячейку H5: H7/C5 в процентах
            ws['H5'] = "=H7/A5"
            ws['H5'].number_format = '0%'
        

        
        
        # Автоподбор ширины колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max(max_length + 2, 10), 30)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Подготавливаем ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
        # Создаем безопасное имя файла из названия клиники
        import re
        import urllib.parse
        
        # Убираем все нелатинские символы и заменяем на латинские аналоги
        clinic_name_safe = submission.clinic_name
        # Заменяем кириллицу на латиницу для совместимости
        cyrillic_to_latin = {
            'а': 'a', 'б': 'b', 'в': 'v', 'г': 'g', 'д': 'd', 'е': 'e', 'ё': 'e',
            'ж': 'zh', 'з': 'z', 'и': 'i', 'й': 'y', 'к': 'k', 'л': 'l', 'м': 'm',
            'н': 'n', 'о': 'o', 'п': 'p', 'р': 'r', 'с': 's', 'т': 't', 'у': 'u',
            'ф': 'f', 'х': 'h', 'ц': 'ts', 'ч': 'ch', 'ш': 'sh', 'щ': 'sch',
            'ъ': '', 'ы': 'y', 'ь': '', 'э': 'e', 'ю': 'yu', 'я': 'ya',
            'А': 'A', 'Б': 'B', 'В': 'V', 'Г': 'G', 'Д': 'D', 'Е': 'E', 'Ё': 'E',
            'Ж': 'Zh', 'З': 'Z', 'И': 'I', 'Й': 'Y', 'К': 'K', 'Л': 'L', 'М': 'M',
            'Н': 'N', 'О': 'O', 'П': 'P', 'Р': 'R', 'С': 'S', 'Т': 'T', 'У': 'U',
            'Ф': 'F', 'Х': 'H', 'Ц': 'Ts', 'Ч': 'Ch', 'Ш': 'Sh', 'Щ': 'Sch',
            'Ъ': '', 'Ы': 'Y', 'Ь': '', 'Э': 'E', 'Ю': 'Yu', 'Я': 'Ya'
        }
        
        for cyr, lat in cyrillic_to_latin.items():
            clinic_name_safe = clinic_name_safe.replace(cyr, lat)
        
        # Убираем все символы кроме букв, цифр, пробелов и дефисов
        clinic_name_safe = re.sub(r'[^\w\s-]', '', clinic_name_safe).strip()
        clinic_name_safe = re.sub(r'[-\s]+', '_', clinic_name_safe)
        
        # Если название пустое, используем fallback
        if not clinic_name_safe:
            clinic_name_safe = f"clinic_{submission.id}"
        
        # Форматируем дату заполнения
        date_str = submission.submitted_at.strftime('%d-%m-%Y')
        
        # Создаем имя файла: название_клиники_дата.xlsx
        # Если название клиники слишком длинное или содержит проблемные символы, используем упрощенный вариант
        if len(clinic_name_safe) > 20 or not clinic_name_safe.replace('_', '').isalnum():
            filename = f"metrics_{submission.id}_{date_str}.xlsx"
        else:
            filename = f"{clinic_name_safe}_{date_str}.xlsx"

        # Пробуем разные варианты заголовков
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
        wb.save(response)
        return response
        
    except ImportError:
        return HttpResponse("Библиотека openpyxl не установлена. Установите её командой: pip install openpyxl", status=500)
    except MetricsSubmission.DoesNotExist:
        return HttpResponse("Форма метрик не найдена", status=404)
    except Exception as e:
        return HttpResponse(f"Ошибка при создании Excel файла: {str(e)}", status=500)
    


class ExternalUsersActivityControlView(ListView):
    template_name = 'courses/external_users_activity_control.html'
    context_object_name = 'activity_logs'
    paginate_by = 30

    def get_queryset(self):
        from builder.models import AuditLog
        from django.contrib.auth.models import User
        
        # Получаем всех внешних пользователей
        external_users = User.objects.filter(groups__name='Внешний пользователь')
        
        # Получаем все логи активности внешних пользователей
        logs = AuditLog.objects.filter(
            user__in=external_users
        ).select_related('user').order_by('-timestamp')
        
        return logs

    def get_context_data(self, **kwargs):
        from django.contrib.auth.models import User
        
        context = super().get_context_data(**kwargs)
        context['title'] = 'Отслеживание активности внешних пользователей'
        
        # Получаем статистику
        external_users = User.objects.filter(groups__name='Внешний пользователь')
        context['total_users'] = external_users.count()
        context['total_actions'] = self.get_queryset().count()
        
        # Получаем последние действия по пользователям
        context['recent_activities'] = self.get_queryset()[:10]
        
        return context


@method_decorator(login_required, name='dispatch')
class AssignCourseToExpertView(View):
    """
    Назначение курса-инцидента руководителю (expert) из связанного инцидента.
    """
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            return render(request, '403.html', status=403)
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        course_slug = kwargs.get('slug')
        try:
            course = get_object_or_404(Course, slug=course_slug, is_incident=True)
            incident = Incident.objects.filter(course=course).first()
            
            if not incident:
                return JsonResponse({'success': False, 'error': 'Инцидент не найден'}, status=404)
            
            if not incident.expert:
                return JsonResponse({'success': False, 'error': 'Руководитель не назначен в инциденте'}, status=400)
                        
                        # Определяем дедлайн:
            # 1) приоритет у incident.expert_time_to_complete (если задано и > 0),
            # 2) затем используем course.default_deadline_days (если задано и > 0),
            # 3) иначе берём значение по умолчанию (3 дня).
            time_to_complete_for_experts = incident.expert_time_to_complete
            if not time_to_complete_for_experts or time_to_complete_for_experts <= 0:
                if course.default_deadline_days and course.default_deadline_days > 0:
                    time_to_complete_for_experts = course.default_deadline_days
                else:
                    time_to_complete_for_experts = 3
            deadline = timezone.now() + timedelta(days=time_to_complete_for_experts)

            # Назначаем курс руководителю
            user_course, created = UserCourse.objects.get_or_create(
                user=incident.expert,
                course=course,
                defaults={'status': 'available', 'deadline': deadline}
            )
            
            if created:
                # Создаем внутреннее уведомление
                try:
                    from notifications.models import Notification
                    Notification.create_course_assignment_notification(incident.expert, course)
                except Exception as e:
                    logger.error(f"Ошибка создания внутреннего уведомления о курсе-инциденте {course.title}: {e}")
                
                # Отправляем email уведомление
                try:
                    from user_management.utils import send_course_assignment_email
                    send_course_assignment_email(incident.expert, course)
                    logger.info(f"Отправлено email уведомление о курсе-инциденте {course.title} руководителю {incident.expert.email}")
                except Exception as e:
                    logger.error(f"Ошибка отправки email уведомления о курсе-инциденте {course.title}: {e}")
            
            # При назначении руководителю статус должен быть 'accepted' (если еще не 'assigned')
            # Статус 'assigned' устанавливается только когда курс назначен сотрудникам
            if incident.status not in ['assigned', 'studies_completed', 'resolved', 'declined']:
                incident.status = 'accepted'
                incident.save(update_fields=['status', 'updated_at'])
            
            return JsonResponse({
                'success': True, 
                'message': f'Курс назначен руководителю {incident.expert.get_full_name()}'
            })
        except Exception as e:
            logger.error(f"Ошибка назначения курса руководителю: {e}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)


@method_decorator(login_required, name='dispatch')
class AssignCourseToAssignedView(View):
    """
    Назначение курса-инцидента назначенным пользователям (assigned_to) из связанного инцидента.
    """
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            return render(request, '403.html', status=403)
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        course_slug = kwargs.get('slug')
        try:
            course = get_object_or_404(Course, slug=course_slug, is_incident=True)
            incident = Incident.objects.filter(course=course).first()
            
            if not incident:
                return JsonResponse({'success': False, 'error': 'Инцидент не найден'}, status=404)
            
            if not incident.assigned_to.exists():
                return JsonResponse({'success': False, 'error': 'Нет назначенных пользователей в инциденте'}, status=400)
            
            assigned_count = 0
            # Определяем дедлайн:
            # 1) приоритет у incident.assigned_to_time_to_complete (если задано и > 0),
            # 2) затем используем course.default_deadline_days (если задано и > 0),
            # 3) иначе берём значение по умолчанию (3 дня).
            time_to_complete = incident.assigned_to_time_to_complete
            if not time_to_complete or time_to_complete <= 0:
                if course.default_deadline_days and course.default_deadline_days > 0:
                    time_to_complete = course.default_deadline_days
                else:
                    time_to_complete = 3
            deadline = timezone.now() + timedelta(days=time_to_complete)
            

            
            # Назначаем курс всем назначенным пользователям
            for user in incident.assigned_to.all():
                user_course, created = UserCourse.objects.get_or_create(
                    user=user,
                    course=course,
                    defaults={'status': 'available', 'deadline': deadline}
                )
                
                if created:
                    assigned_count += 1
                    # Создаем внутреннее уведомление
                    try:
                        from notifications.models import Notification
                        Notification.create_course_assignment_notification(user, course)
                    except Exception as e:
                        logger.error(f"Ошибка создания внутреннего уведомления о курсе-инциденте {course.title}: {e}")
                    
                    # Отправляем email уведомление
                    try:
                        from user_management.utils import send_course_assignment_email
                        send_course_assignment_email(user, course)
                        logger.info(f"Отправлено email уведомление о курсе-инциденте {course.title} пользователю {user.email}")
                    except Exception as e:
                        logger.error(f"Ошибка отправки email уведомления о курсе-инциденте {course.title}: {e}")
            
            # Меняем статус инцидента на 'assigned', так как курс назначен сотрудникам
            if assigned_count > 0 and incident.status != 'assigned':
                incident.status = 'assigned'
                incident.save(update_fields=['status', 'updated_at'])
            
            return JsonResponse({
                'success': True, 
                'message': f'Курс назначен {assigned_count} пользователям'
            })
        except Exception as e:
            logger.error(f"Ошибка назначения курса назначенным пользователям: {e}")
            return JsonResponse({'success': False, 'error': str(e)}, status=500)
