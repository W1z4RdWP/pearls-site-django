from datetime import datetime, timedelta

from django.shortcuts import get_object_or_404
from django.views.generic import TemplateView, ListView
from django.contrib.auth.models import User, Group
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.db.models import Q
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, HttpResponseForbidden
from django.utils import timezone




class HomeworkCheckDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """
    Страница проверки заданий для наставников - простая статистика
    """
    template_name = 'reports/homework_check_dashboard.html'
    
    def test_func(self):
        """Проверяет права доступа"""
        if not self.request.user.is_authenticated:
            return False
        
        # Суперпользователи и персонал имеют доступ
        if self.request.user.is_superuser or self.request.user.is_staff:
            return True
        
        # Наставники имеют доступ
        try:
            return self.request.user.profile.is_mentor_user
        except:
            return False
    
    def get_context_data(self, **kwargs):
        """Добавляет статистику в контекст"""
        context = super().get_context_data(**kwargs)
        
        # Импортируем модели
        from courses.models import Lesson, Course
        from quizzes.models import Quiz
        
        # Для последних завершений
        from myapp.models import UserProgress, QuizResult
        
        # Проверяем, является ли пользователь суперпользователем или стафом
        is_admin = self.request.user.is_superuser or self.request.user.is_staff
        
        if is_admin:
            # Для администраторов - общая статистика по всей платформе
            total_lessons = Lesson.objects.count()
            total_quizzes = Quiz.objects.count()
            total_materials = total_lessons + total_quizzes
            active_users = User.objects.filter(profile__is_approved=True).count()
            total_groups = Group.objects.count()
            # Последние завершения по всей платформе
            recent_lesson_progress = list(
                UserProgress.objects.select_related('user', 'course', 'lesson')
                .filter(completed=True)
                .exclude(completed_at__isnull=True)
                .order_by('-completed_at')[:20]
            )
            recent_quiz_results = list(
                QuizResult.objects.select_related('user', 'course')
                .filter(passed=True)
                .exclude(completed_at__isnull=True)
                .order_by('-completed_at')[:20]
            )
        else:
            # Для наставников - статистика только по их группам
            mentor_groups = self.request.user.groups.all()
            
            if mentor_groups.exists():
                # Получаем пользователей из групп наставника
                mentor_group_users = User.objects.filter(groups__in=mentor_groups).distinct()
                
                # Получаем курсы, на которые записаны пользователи из групп наставника
                mentor_courses = Course.objects.filter(usercourse__user__groups__in=mentor_groups).distinct()
                
                # Уроки из курсов наставника
                total_lessons = Lesson.objects.filter(courses__in=mentor_courses).distinct().count()
                
                # Тесты из курсов наставника
                total_quizzes = Quiz.objects.filter(courses__in=mentor_courses).distinct().count()
                
                total_materials = total_lessons + total_quizzes
                
                # Активные пользователи из групп наставника
                active_users = mentor_group_users.filter(profile__is_approved=True).count()
                
                # Количество групп наставника
                total_groups = mentor_groups.count()

                # Последние завершения только пользователей из групп наставника
                recent_lesson_progress = list(
                    UserProgress.objects.select_related('user', 'course', 'lesson')
                    .filter(completed=True, user__in=mentor_group_users)
                    .exclude(completed_at__isnull=True)
                    .order_by('-completed_at')[:20]
                )
                recent_quiz_results = list(
                    QuizResult.objects.select_related('user', 'course')
                    .filter(passed=True, user__in=mentor_group_users)
                    .exclude(completed_at__isnull=True)
                    .order_by('-completed_at')[:20]
                )
            else:
                # Если у наставника нет групп, показываем нули
                total_lessons = 0
                total_quizzes = 0
                total_materials = 0
                active_users = 0
                total_groups = 0
                recent_lesson_progress = []
                recent_quiz_results = []
        
        # Объединяем и сортируем последние завершения, берем топ-10
        def fio_short(user):
            last_name = (user.last_name or '').strip()
            first_initial = (user.first_name[:1] + '.') if user.first_name else ''
            middle_initial = ''
            try:
                middle_name = getattr(user, 'profile', None) and getattr(user.profile, 'middle_name', '')
                if middle_name:
                    middle_initial = middle_name[:1] + '.'
            except Exception:
                middle_initial = ''
            parts = [p for p in [last_name, first_initial + middle_initial] if p]
            return ' '.join(parts) if parts else (user.get_username() or user.email)

        combined = []
        for lp in recent_lesson_progress:
            combined.append({
                'type': 'lesson',
                'user': lp.user,
                'fio_short': fio_short(lp.user),
                'course_title': getattr(lp.course, 'title', getattr(lp.course, 'name', '')) if lp.course else '',
                'material_title': getattr(lp.lesson, 'title', getattr(lp.lesson, 'name', '')) if lp.lesson else '',
                'completed_at': lp.completed_at,
            })
        for qr in recent_quiz_results:
            combined.append({
                'type': 'quiz',
                'user': qr.user,
                'fio_short': fio_short(qr.user),
                'course_title': getattr(qr.course, 'title', getattr(qr.course, 'name', '')) if qr.course else '',
                'material_title': qr.quiz_title,
                'completed_at': qr.completed_at,
            })

        combined.sort(key=lambda x: x['completed_at'] or timezone.make_aware(datetime.min), reverse=True)
        recent_completions = combined[:10]

        # Подсчитываем количество тестов, ожидающих проверки (pending)
        # Учитываем только лучшие попытки для каждой комбинации (user, quiz_title, course)
        from myapp.models import QuizResult
        
        if is_admin:
            # Для администраторов - все pending тесты
            base_query = QuizResult.objects.filter(status='pending')
        else:
            # Для наставников - только pending тесты студентов из их групп
            mentor_groups = self.request.user.groups.all()
            if mentor_groups.exists():
                mentor_group_users = User.objects.filter(groups__in=mentor_groups).distinct()
                base_query = QuizResult.objects.filter(
                    status='pending',
                    user__in=mentor_group_users
                )
            else:
                base_query = QuizResult.objects.none()
        
        # Получаем только лучшие попытки для каждой комбинации (user, quiz_title, course)
        # Сортируем по user, quiz_title, затем по course (с учетом NULL), затем по percent и completed_at
        all_results = base_query.order_by(
            'user_id', 'quiz_title', 'course_id', '-percent', '-completed_at'
        )
        
        # Группируем по (user, quiz_title, course) и берем первую (лучшую) попытку
        seen = set()
        best_result_ids = []
        for result in all_results:
            # Используем None для course_id, если курс не указан
            course_id = result.course_id if result.course_id else None
            key = (result.user_id, result.quiz_title, course_id)
            if key not in seen:
                seen.add(key)
                best_result_ids.append(result.id)
        
        # Считаем только лучшие попытки
        pending_tests_count = len(best_result_ids)

        context.update({
            'total_materials': total_materials,
            'total_lessons': total_lessons,
            'total_quizzes': total_quizzes,
            'active_users': active_users,
            'total_groups': total_groups,
            'is_admin': is_admin,
            'recent_completions': recent_completions,
            'pending_tests_count': pending_tests_count,
        })
        
        return context




class UsersWithLearningView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    Страница списка пользователей с назначенным обучением
    """
    model = User
    template_name = 'reports/users_with_learning.html'
    context_object_name = 'users'
    paginate_by = 20
    
    def test_func(self):
        """Проверяет права доступа"""
        if not self.request.user.is_authenticated:
            return False
        
        # Суперпользователи и персонал имеют доступ
        if self.request.user.is_superuser or self.request.user.is_staff:
            return True
        
        # Наставники имеют доступ
        try:
            return self.request.user.profile.is_mentor_user
        except:
            return False
    
    def get_queryset(self):
        """Возвращает пользователей с назначенным обучением"""
        from courses.models import Course
        from myapp.models import UserCourse
        
        # Проверяем, является ли пользователь суперпользователем или стафом
        is_admin = self.request.user.is_superuser or self.request.user.is_staff
        
        if is_admin:
            # Для администраторов - все пользователи с назначенными курсами (исключая админов и суперюзеров)
            queryset = User.objects.filter(
                started_courses__isnull=False,
                profile__is_approved=True
            ).exclude(
                Q(is_superuser=True) | Q(is_staff=True)
            ).distinct().select_related('profile').prefetch_related('groups')
        else:
            # Для наставников - только пользователи из их групп с назначенными курсами (исключая админов и суперюзеров)
            mentor_groups = self.request.user.groups.all()
            if mentor_groups.exists():
                queryset = User.objects.filter(
                    groups__in=mentor_groups,
                    started_courses__isnull=False,
                    profile__is_approved=True
                ).exclude(
                    Q(is_superuser=True) | Q(is_staff=True)
                ).distinct().select_related('profile').prefetch_related('groups')
            else:
                queryset = User.objects.none()
        
        # Поиск по ФИО
        search_query = self.request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(profile__middle_name__icontains=search_query)
            )
        
        # Фильтр по группе
        group_filter = self.request.GET.get('group')
        if group_filter:
            queryset = queryset.filter(groups__id=group_filter)
        
        queryset = queryset.order_by('last_name', 'first_name')
        
        # Добавляем информацию о статусе курсов для каждого пользователя
        for user in queryset:
            user_courses = user.started_courses.all()
            total_courses = user_courses.count()
            completed_courses = user_courses.filter(status='completed').count()
            
            # Добавляем атрибуты для использования в template
            user.total_courses = total_courses
            user.completed_courses = completed_courses
            user.is_fully_completed = completed_courses == total_courses if total_courses > 0 else False
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Добавляет дополнительный контекст"""
        context = super().get_context_data(**kwargs)
        
        # Импортируем модели
        from courses.models import Course
        from myapp.models import UserCourse, UserProgress, QuizResult
        
        # Проверяем, является ли пользователь суперпользователем или стафом
        is_admin = self.request.user.is_superuser or self.request.user.is_staff
        
        # Группы для фильтра
        if is_admin:
            context['groups'] = Group.objects.all().order_by('name')
        else:
            context['groups'] = self.request.user.groups.all().order_by('name')
        
        # Сортируем пользователей по проценту завершенных курсов (от большего к меньшему)
        users_list = list(context['users'])
        users_list.sort(key=lambda u: (u.completed_courses / u.total_courses * 100 if u.total_courses > 0 else 0), reverse=True)
        context['users'] = users_list
        
        # Статистика обученности - считаем по курсам, а не по пользователям
        users_with_learning = self.get_queryset()
        
        # Получаем все назначенные курсы для этих пользователей
        from myapp.models import UserCourse
        user_courses = UserCourse.objects.filter(user__in=users_with_learning)
        
        total_courses = user_courses.count()
        
        if total_courses > 0:
            # Подсчитываем курсы по статусам
            completed_courses = user_courses.filter(status='completed').count()
            in_progress_courses = user_courses.filter(status='started').count()
            available_courses = user_courses.filter(status='available').count()
            
            # Процент обученности (завершенные курсы от общего количества)
            learning_percentage = round((completed_courses / total_courses) * 100, 1) if total_courses > 0 else 0
            
            context.update({
                'total_courses': total_courses,
                'completed_courses': completed_courses,
                'in_progress_courses': in_progress_courses,
                'available_courses': available_courses,
                'learning_percentage': learning_percentage,
                'learning_data': [
                    {'label': 'Завершено', 'value': completed_courses, 'color': '#28a745'},
                    {'label': 'В процессе', 'value': in_progress_courses, 'color': '#ffc107'},
                    {'label': 'Не начато', 'value': available_courses, 'color': '#6c757d'}
                ]
            })
        else:
            context.update({
                'total_courses': 0,
                'completed_courses': 0,
                'in_progress_courses': 0,
                'available_courses': 0,
                'learning_percentage': 0,
                'learning_data': []
            })
        
        # Параметры фильтрации
        context['search_query'] = self.request.GET.get('search', '')
        context['selected_group'] = self.request.GET.get('group', '')
        context['is_admin'] = is_admin
        
        # Сериализуем данные для JavaScript
        import json
        context['learning_data_json'] = json.dumps(context.get('learning_data', []), ensure_ascii=False)
        
        return context


@login_required
def export_users_learning_excel(request):
    """
    Экспорт пользователей с обучением в Excel
    """
    # Проверяем права доступа
    if not request.user.is_authenticated:
        return HttpResponseForbidden("Доступ запрещен")
    
    is_admin = request.user.is_superuser or request.user.is_staff
    if not is_admin:
        try:
            if not request.user.profile.is_mentor_user:
                return HttpResponseForbidden("Доступ запрещен")
        except:
            return HttpResponseForbidden("Доступ запрещен")
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        from myapp.models import UserCourse
        
        # Получаем queryset с теми же фильтрами, что и в UsersWithLearningView
        if is_admin:
            queryset = User.objects.filter(
                started_courses__isnull=False,
                profile__is_approved=True
            ).exclude(
                Q(is_superuser=True) | Q(is_staff=True)
            ).distinct().select_related('profile').prefetch_related('groups', 'started_courses')
        else:
            mentor_groups = request.user.groups.all()
            if mentor_groups.exists():
                queryset = User.objects.filter(
                    groups__in=mentor_groups,
                    started_courses__isnull=False,
                    profile__is_approved=True
                ).exclude(
                    Q(is_superuser=True) | Q(is_staff=True)
                ).distinct().select_related('profile').prefetch_related('groups', 'started_courses')
            else:
                queryset = User.objects.none()
        
        # Поиск по ФИО
        search_query = request.GET.get('search')
        if search_query:
            queryset = queryset.filter(
                Q(first_name__icontains=search_query) |
                Q(last_name__icontains=search_query) |
                Q(profile__middle_name__icontains=search_query)
            )
        
        # Фильтр по группе
        group_filter = request.GET.get('group')
        if group_filter:
            queryset = queryset.filter(groups__id=group_filter)
        
        queryset = queryset.order_by('last_name', 'first_name')
        
        # Создаем workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Прогресс учащихся"
        
        # Заголовок отчета
        ws.merge_cells('A1:K1')
        header_cell = ws['A1']
        header_cell.value = "Отчёт Прогресс учащихся"
        header_cell.font = Font(bold=True, size=14)
        header_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        ws.merge_cells('A2:K2')
        date_cell = ws['A2']
        date_cell.value = f"Дата создания отчёта: {timezone.now().strftime('%Y-%m-%d')}"
        date_cell.font = Font(size=10)
        date_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Пустая строка
        ws.append([])
        ws.append([])
        
        # Заголовки столбцов
        headers = [
            'Имя пользователя',
            'Подразделение',
            'Обученность (%)',
            'Назначений',
            'Завершено',
            'Заблокированно',
            'В процессе',
            'Не начато',
            'Просрочено',
            'Общее установленное время',
            'Общее затраченное время'
        ]
        ws.append(headers)
        
        # Стили для заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )
        
        for cell in ws[5]:  # Заголовки в строке 5
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Текущее время для расчетов
        now = timezone.now()
        
        # Заполняем данные
        for user in queryset:
            user_courses = user.started_courses.all()
            
            # Подсчитываем курсы по статусам
            total_courses = user_courses.count()
            completed = user_courses.filter(status='completed').count()
            blocked = user_courses.filter(status='blocked').count()
            in_progress = user_courses.filter(status='started').count()
            not_started = user_courses.filter(status='available').count()
            
            # Просроченные курсы (deadline < now и status != 'completed')
            overdue = user_courses.filter(
                deadline__lt=now
            ).exclude(status='completed').count()
            
            # Обученность (%)
            learning_percentage = round((completed / total_courses * 100), 2) if total_courses > 0 else 0
            
            # Общее установленное время (сумма времени из deadline для всех курсов с deadline)
            total_allocated_time = timedelta(0)
            for uc in user_courses.filter(deadline__isnull=False):
                if uc.deadline:
                    # Извлекаем только время из deadline (часы, минуты, секунды)
                    deadline_time = uc.deadline.time()
                    total_allocated_time += timedelta(
                        hours=deadline_time.hour,
                        minutes=deadline_time.minute,
                        seconds=deadline_time.second
                    )
            
            # Общее затраченное время
            # Сумма времени курсов с момента начала до текущего момента (если не завершен) или до завершения
            # Но только для курсов С deadline (если deadline не задан, то +0:00:00)
            total_spent_time = timedelta(0)
            for uc in user_courses:
                # Если для курса НЕ задан deadline, не учитываем (добавляем 0:00:00)
                if not uc.deadline:
                    continue  # Пропускаем курсы без deadline
                
                # Для курсов с deadline считаем время
                if uc.status == 'completed' and uc.end_date and uc.start_date:
                    total_spent_time += (uc.end_date - uc.start_date)
                elif uc.status != 'completed' and uc.start_date:
                    total_spent_time += (now - uc.start_date)
            
            # Форматируем время
            def format_timedelta(td):
                """Форматирует timedelta в читаемый формат"""
                total_seconds = int(td.total_seconds())
                if total_seconds == 0:
                    return "0:00:00"
                days = td.days
                hours, remainder = divmod(total_seconds, 3600)
                minutes, seconds = divmod(remainder, 60)
                if days > 0:
                    return f"{days}д {hours:02d}:{minutes:02d}:{seconds:02d}"
                else:
                    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            # Группы пользователя
            groups = ', '.join([group.name for group in user.groups.all()]) if user.groups.exists() else ''
            
            # Имя пользователя
            user_name = user.get_full_name() or user.email
            
            # Добавляем строку данных
            row = [
                user_name,
                groups,
                learning_percentage,
                total_courses,
                completed,
                blocked,
                in_progress,
                not_started,
                overdue,
                format_timedelta(total_allocated_time),
                format_timedelta(total_spent_time)
            ]
            ws.append(row)
            
            # Применяем стили к строке данных
            for cell in ws[ws.max_row]:
                cell.border = border
                if cell.column == 3:  # Обученность (%)
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif cell.column >= 4 and cell.column <= 9:  # Числовые столбцы
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Автоматическая ширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Создаем ответ
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"progress_report_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response
        
    except ImportError:
        return HttpResponse("Библиотека openpyxl не установлена. Установите её командой: pip install openpyxl", status=500)
    except Exception as e:
        return HttpResponse(f"Ошибка при создании Excel файла: {str(e)}", status=500)


class GroupsProgressView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    Страница прогресса групп
    """
    model = Group
    template_name = 'reports/groups_progress.html'
    context_object_name = 'groups'
    
    def test_func(self):
        """Проверяет права доступа"""
        if not self.request.user.is_authenticated:
            return False
        
        # Суперпользователи и персонал имеют доступ
        if self.request.user.is_superuser or self.request.user.is_staff:
            return True
        
        # Наставники имеют доступ
        try:
            return self.request.user.profile.is_mentor_user
        except:
            return False
    
    def get_queryset(self):
        """Возвращает все группы"""
        from myapp.models import UserCourse
        
        # Проверяем, является ли пользователь суперпользователем или стафом
        is_admin = self.request.user.is_superuser or self.request.user.is_staff
        
        if is_admin:
            # Для администраторов - все группы
            queryset = Group.objects.all().prefetch_related('user_set')
        else:
            # Для наставников - только их группы
            mentor_groups = self.request.user.groups.all()
            if mentor_groups.exists():
                queryset = Group.objects.filter(
                    id__in=mentor_groups
                ).prefetch_related('user_set')
            else:
                queryset = Group.objects.none()
        
        # Фильтр по группе
        group_filter = self.request.GET.get('group')
        if group_filter:
            queryset = queryset.filter(id=group_filter)
        
        return queryset.order_by('name')
    
    def get_context_data(self, **kwargs):
        """Добавляет дополнительный контекст"""
        context = super().get_context_data(**kwargs)
        
        # Импортируем модели
        from myapp.models import UserCourse
        
        # Проверяем, является ли пользователь суперпользователем или стафом
        is_admin = self.request.user.is_superuser or self.request.user.is_staff
        
        # Добавляем информацию о прогрессе для каждой группы на текущей странице
        groups_list = list(context['groups'])
        for group in groups_list:
            # Получаем пользователей группы с назначенными курсами (исключая админов и суперюзеров)
            group_users = group.user_set.filter(
                started_courses__isnull=False,
                profile__is_approved=True
            ).exclude(
                Q(is_superuser=True) | Q(is_staff=True)
            ).distinct()
            
            # Получаем все курсы для пользователей этой группы
            group_courses = UserCourse.objects.filter(user__in=group_users)
            
            total_courses = group_courses.count()
            completed_courses = group_courses.filter(status='completed').count()
            in_progress_courses = group_courses.filter(status='started').count()
            available_courses = group_courses.filter(status='available').count()
            
            # Процент обученности группы
            learning_percentage = round((completed_courses / total_courses) * 100, 1) if total_courses > 0 else 0
            
            # Добавляем атрибуты для использования в template
            group.total_users = group_users.count()
            group.total_courses = total_courses
            group.completed_courses = completed_courses
            group.in_progress_courses = in_progress_courses
            group.available_courses = available_courses
            group.learning_percentage = learning_percentage
        
        # Сортируем группы по проценту обученности (от большего к меньшему)
        groups_list.sort(key=lambda x: x.learning_percentage, reverse=True)
        context['groups'] = groups_list
        
        # Получаем все доступные группы для фильтра (без применения фильтра)
        if is_admin:
            all_available_groups = Group.objects.all().order_by('name')
        else:
            mentor_groups = self.request.user.groups.all()
            if mentor_groups.exists():
                all_available_groups = Group.objects.filter(
                    id__in=mentor_groups
                ).order_by('name')
            else:
                all_available_groups = Group.objects.none()
        
        # Общая статистика по всем отображаемым группам (с учетом фильтра)
        all_groups = context['groups']
        total_groups = len(all_groups)
        
        # Подсчитываем общую статистику по всем отображаемым группам
        all_users = []
        for group in all_groups:
            group_users = group.user_set.filter(
                started_courses__isnull=False,
                profile__is_approved=True
            ).exclude(
                Q(is_superuser=True) | Q(is_staff=True)
            ).distinct()
            all_users.extend(group_users)
        
        all_courses = UserCourse.objects.filter(user__in=all_users)
        total_courses = all_courses.count()
        completed_courses = all_courses.filter(status='completed').count()
        in_progress_courses = all_courses.filter(status='started').count()
        available_courses = all_courses.filter(status='available').count()
        
        overall_learning_percentage = round((completed_courses / total_courses) * 100, 1) if total_courses > 0 else 0
        
        context.update({
            'total_groups': total_groups,
            'total_courses': total_courses,
            'completed_courses': completed_courses,
            'in_progress_courses': in_progress_courses,
            'available_courses': available_courses,
            'overall_learning_percentage': overall_learning_percentage,
            'is_admin': is_admin,
            'all_available_groups': all_available_groups,
            'selected_group': self.request.GET.get('group', ''),
            'learning_data': [
                {'label': 'Завершено', 'value': completed_courses, 'color': '#28a745'},
                {'label': 'В процессе', 'value': in_progress_courses, 'color': '#ffc107'},
                {'label': 'Не начато', 'value': available_courses, 'color': '#6c757d'}
            ]
        })
        
        # Сериализуем данные для JavaScript
        import json
        context['learning_data_json'] = json.dumps(context.get('learning_data', []), ensure_ascii=False)
        
        return context




class GroupStudentsProgressView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """
    Страница прогресса студентов конкретной группы
    """
    model = User
    template_name = 'reports/group_students_progress.html'
    context_object_name = 'students'
    paginate_by = 20
    
    def test_func(self):
        """Проверяет права доступа"""
        if not self.request.user.is_authenticated:
            return False
        
        # Суперпользователи и персонал имеют доступ
        if self.request.user.is_superuser or self.request.user.is_staff:
            return True
        
        # Наставники имеют доступ только к своим группам
        try:
            if self.request.user.profile.is_mentor_user:
                group_id = self.kwargs.get('group_id')
                return self.request.user.groups.filter(id=group_id).exists()
            return False
        except:
            return False
    
    def get_queryset(self):
        """Возвращает студентов группы с назначенным обучением"""
        from myapp.models import UserCourse
        from django.db.models import Count, Case, When, FloatField, F, Q
        
        group_id = self.kwargs.get('group_id')
        self.group = get_object_or_404(Group, id=group_id)
        
        # Получаем пользователей группы с назначенными курсами (исключая админов и суперюзеров)
        queryset = User.objects.filter(
            groups=self.group,
            started_courses__isnull=False,
            profile__is_approved=True
        ).exclude(
            Q(is_superuser=True) | Q(is_staff=True)
        ).annotate(
            total_courses=Count('started_courses'),
            completed_courses=Count('started_courses', filter=Q(started_courses__status='completed')),
            in_progress_courses=Count('started_courses', filter=Q(started_courses__status='started')),
            learning_percentage=Case(
                When(total_courses=0, then=0),
                default=F('completed_courses') * 100.0 / F('total_courses'),
                output_field=FloatField()
            )
        ).distinct().order_by('-learning_percentage', 'last_name', 'first_name')
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Добавляет дополнительный контекст"""
        context = super().get_context_data(**kwargs)
        
        context['group'] = self.group
        
        # Данные уже рассчитаны в queryset, дополнительная обработка не нужна
        return context
