from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from django.views.generic import ListView, CreateView, UpdateView, DetailView, FormView, TemplateView
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User, Group
from django.db.models import Q, Count, Max, F, Sum
from users.models import Profile, Role
from users.permissions import MentorRequiredMixin
from django import forms
from django.urls import reverse_lazy
from django.core.exceptions import PermissionDenied
from myapp.models import UserProgress, UserCourse, QuizResult, UserAnswer
from quizzes.models import Quiz, QuizLock
from courses.models import Course, Lesson, UserLessonTrajectory
from django.core.paginator import Paginator, PageNotAnInteger, EmptyPage
from django.utils.decorators import method_decorator
from django.contrib.admin.views.decorators import staff_member_required
from .forms import UserProfileForm
from users.forms import UserRegisterNoCaptchaForm
from django.contrib.auth.forms import SetPasswordForm
from django.views.decorators.http import require_POST
from django.urls import reverse
from django.contrib import messages
from .utils import send_user_credentials_email
from gamification.models import DascoinTransaction
import logging
from datetime import datetime
from django.utils import timezone
from django.template.loader import render_to_string
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from urllib.parse import urlencode
from weasyprint import HTML


# Получаем логгер для записи в журнал аудита
audit_logger = logging.getLogger('audit')


class UserListView(ListView):
    model = User
    template_name = 'user_management/user_list.html'
    context_object_name = 'users'
    paginate_by = 20

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.is_mentor_user)):
            raise PermissionDenied("У вас нет доступа к управлению пользователями.")
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        queryset = super().get_queryset().order_by('email')
        
        # Если пользователь - наставник (но не superuser и не staff), показываем только его группу
        if (hasattr(self.request.user, 'profile') and 
            self.request.user.profile.is_mentor_user and 
            not self.request.user.is_superuser and 
            not self.request.user.is_staff):
            # Получаем группы наставника
            mentor_groups = self.request.user.groups.all()
            if mentor_groups.exists():
                # Показываем только пользователей из групп наставника
                queryset = queryset.filter(groups__in=mentor_groups).distinct()
            else:
                # Если у наставника нет групп, показываем пустой список
                queryset = queryset.none()
        
        q = self.request.GET.get('q')
        if q:
            queryset = queryset.filter(
                Q(email__icontains=q) |
                Q(first_name__icontains=q) |
                Q(last_name__icontains=q)
            )
            
        filter_val = self.request.GET.get('filter')
        # По умолчанию применяем фильтр "approved", если filter не задан
        if filter_val is None:
            filter_val = 'approved'

        if filter_val == 'approved':
            queryset = queryset.filter(profile__is_approved=True)
        elif filter_val == 'not_approved':
            queryset = queryset.filter(profile__is_approved=False)
        elif filter_val == 'responsible':
            queryset = queryset.filter(profile__role__responsible_user=F('id'))
        elif filter_val == 'not_responsible':
            queryset = queryset.filter(
                Q(profile__role__responsible_user__isnull=True) |
                ~Q(profile__role__responsible_user=F('id'))
            )
        
        # Фильтрация по группе (только для не-наставников)
        if not (hasattr(self.request.user, 'profile') and 
                self.request.user.profile.is_mentor_user and 
                not self.request.user.is_superuser and 
                not self.request.user.is_staff):
            group_filter = self.request.GET.get('group')
            if group_filter:
                queryset = queryset.filter(groups__id=group_filter)
        
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Добавляем список групп для фильтра (только для не-наставников)
        from django.contrib.auth.models import Group
        if not (hasattr(self.request.user, 'profile') and 
                self.request.user.profile.is_mentor_user and 
                not self.request.user.is_superuser and 
                not self.request.user.is_staff):
            context['groups'] = Group.objects.all().order_by('name')
        else:
            # Для наставников показываем только их группы
            context['groups'] = self.request.user.groups.all().order_by('name')
        return context


class UserCreateStep1View(CreateView):
    template_name = 'user_management/user_create_step1.html'
    form_class = UserRegisterNoCaptchaForm
    success_url = reverse_lazy('user_management:user_create_step2')

    def form_valid(self, form):
        user = form.save()
        self.request.session['user_create_step1_user_id'] = user.id
        self.request.session['user_password'] = form.cleaned_data['password1']
        return redirect(self.success_url)




class UserCreateStep2View(CreateView):
    template_name = 'user_management/user_create_step2.html'
    form_class = UserProfileForm
    success_url = reverse_lazy('user_management:user_list')


    def dispatch(self, request, *args, **kwargs):
        if 'user_create_step1_user_id' not in request.session:
            return redirect('user_management:user_create_step1')
        return super().dispatch(request, *args, **kwargs)


    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        user_id = self.request.session.get('user_create_step1_user_id')
        user = User.objects.get(id=user_id)
        kwargs['instance'] = user.profile
        kwargs['user_instance'] = user
        return kwargs


    def form_valid(self, form):
        form.save()
        user_id = self.request.session.get('user_create_step1_user_id')
        user = User.objects.get(id=user_id)
        
        # Получаем пароль из сессии (нужно сохранить его в step1)
        password = self.request.session.get('user_password')
        
        # Отправляем email с данными для входа
        if password and send_user_credentials_email(user, password):
            messages.success(self.request, f'Пользователь {user.email} создан. Email с данными для входа отправлен на {user.email}')
        else:
            messages.warning(self.request, f'Пользователь {user.email} создан, но не удалось отправить email с данными для входа')
        
        del self.request.session['user_create_step1_user_id']
        if 'user_password' in self.request.session:
            del self.request.session['user_password']
        
        return redirect(self.success_url)




def get_user_privilege_level(user):
    if user.is_superuser:
        return 3
    if user.is_staff:
        return 2
    return 1




def role_manage(request):
    if not request.user.is_staff:
        raise PermissionDenied
    if request.method == 'POST':
        name = request.POST.get('new_role', '').strip()
        if name:
            Role.objects.get_or_create(name=name)
            messages.success(request, f'Должность "{name}" добавлена.')
        return redirect(request.META.get('HTTP_REFERER', reverse('user_management:user_list')))
    return redirect('user_management:user_list')




@require_POST
def role_delete(request, role_id):
    if not request.user.is_staff:
        raise PermissionDenied
    Role.objects.filter(id=role_id).delete()
    messages.success(request, 'Должность удалена.')
    return redirect(request.META.get('HTTP_REFERER', reverse('user_management:user_list')))

@require_POST
def role_edit(request, role_id):
    if not request.user.is_staff:
        raise PermissionDenied
    new_name = request.POST.get('new_name', '').strip()
    if new_name:
        Role.objects.filter(id=role_id).update(name=new_name)
        messages.success(request, f'Должность переименована в "{new_name}".')
    else:
        messages.error(request, 'Название не может быть пустым.')
    return redirect(request.META.get('HTTP_REFERER', reverse('user_management:user_list')))

@require_POST
def role_responsible_manage(request, role_id):
    """
    Управление ответственным для роли
    """
    if not request.user.is_staff:
        raise PermissionDenied
    
    try:
        role = Role.objects.get(id=role_id)
    except Role.DoesNotExist:
        messages.error(request, 'Должность не найдена.')
        return redirect(request.META.get('HTTP_REFERER', reverse('user_management:user_list')))
    
    responsible_id = request.POST.get('responsible_id')
    
    if responsible_id:
        try:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            user = User.objects.get(id=responsible_id)
            
            # Проверяем, что у пользователя эта роль
            if user.profile.role != role:
                messages.error(request, f'Пользователь {user.get_full_name()} не имеет должности "{role.name}"')
                return redirect(request.META.get('HTTP_REFERER', reverse('user_management:user_list')))
            
            # Проверяем, что у другой роли этот пользователь не назначен ответственным
            other_role = Role.objects.filter(responsible_user=user).exclude(id=role.id).first()
            if other_role:
                messages.error(request, f'Пользователь {user.get_full_name()} уже назначен ответственным за должность "{other_role.name}"')
                return redirect(request.META.get('HTTP_REFERER', reverse('user_management:user_list')))
            
            role.responsible_user = user
            role.save()
            messages.success(request, f'Пользователь {user.get_full_name()} назначен ответственным за должность "{role.name}"')
            
        except User.DoesNotExist:
            messages.error(request, 'Пользователь не найден.')
    else:
        # Убираем ответственного
        if role.responsible_user:
            old_responsible = role.responsible_user.get_full_name()
            role.responsible_user = None
            role.save()
            messages.success(request, f'Ответственный {old_responsible} снят с должности "{role.name}"')
        else:
            messages.info(request, 'Ответственный не был назначен.')
    
    return redirect(request.META.get('HTTP_REFERER', reverse('user_management:user_list')))


def role_users_json(request, role_id):
    """
    Возвращает список пользователей с данной ролью в JSON формате
    """
    if not request.user.is_staff:
        from django.http import JsonResponse
        return JsonResponse({'error': 'forbidden'}, status=403)
    
    try:
        role = Role.objects.get(id=role_id)
    except Role.DoesNotExist:
        from django.http import JsonResponse
        return JsonResponse({'error': 'role not found'}, status=404)
    
    from django.contrib.auth import get_user_model
    from django.http import JsonResponse
    User = get_user_model()
    
    users = User.objects.filter(profile__role=role)
    users_data = []
    
    for user in users:
        users_data.append({
            'id': user.id,
            'full_name': user.get_full_name(),
            'is_responsible': role.responsible_user == user
        })
    
    return JsonResponse({'users': users_data})


def roles_all_json(request):
    """
    Возвращает список всех ролей с их ответственными в JSON формате
    """
    if not request.user.is_staff:
        from django.http import JsonResponse
        return JsonResponse({'error': 'forbidden'}, status=403)
    
    from django.http import JsonResponse
    from users.models import Role
    
    roles = Role.objects.all().prefetch_related('responsible_user')
    roles_data = []
    
    for role in roles:
        role_data = {
            'id': role.id,
            'name': role.name,
            'responsible_user': None
        }
        
        if role.responsible_user:
            role_data['responsible_user'] = {
                'id': role.responsible_user.id,
                'full_name': role.responsible_user.get_full_name()
            }
        
        roles_data.append(role_data)
    
    return JsonResponse({'roles': roles_data})


def lesson_allowed_roles_json(request, lesson_id):
    """
    Возвращает список разрешенных должностей для урока в JSON формате
    """
    if not request.user.is_staff:
        from django.http import JsonResponse
        return JsonResponse({'error': 'forbidden'}, status=403)
    
    from django.http import JsonResponse
    from builder.models import LessonAllowedRole
    from courses.models import Lesson
    
    try:
        lesson = Lesson.objects.get(id=lesson_id)
    except Lesson.DoesNotExist:
        return JsonResponse({'error': 'lesson not found'}, status=404)
    
    allowed_roles = LessonAllowedRole.objects.filter(lesson=lesson).select_related('role', 'role__responsible_user')
    roles_data = []
    
    for allowed_role in allowed_roles:
        role_data = {
            'id': allowed_role.role.id,
            'name': allowed_role.role.name,
            'responsible_fio': allowed_role.responsible_fio,
            'added_at': allowed_role.added_at.strftime('%d.%m.%Y')
        }
        roles_data.append(role_data)
    
    return JsonResponse({'allowed_roles': roles_data})


@require_POST
def lesson_add_allowed_role(request, lesson_id):
    """
    Добавляет должность в список разрешенных для урока
    """
    if not request.user.is_staff:
        from django.http import JsonResponse
        return JsonResponse({'error': 'forbidden'}, status=403)
    
    from django.http import JsonResponse
    from builder.models import LessonAllowedRole
    from courses.models import Lesson
    from users.models import Role
    
    try:
        lesson = Lesson.objects.get(id=lesson_id)
        role_id = request.POST.get('role_id')
        if not role_id:
            return JsonResponse({'error': 'role_id is required'}, status=400)
        
        role = Role.objects.get(id=role_id)
        
        # Проверяем, не добавлена ли уже эта должность
        if LessonAllowedRole.objects.filter(lesson=lesson, role=role).exists():
            return JsonResponse({'error': 'Должность уже добавлена'}, status=400)
        
        # Добавляем должность
        LessonAllowedRole.objects.create(lesson=lesson, role=role)
        
        return JsonResponse({'success': True})
        
    except Lesson.DoesNotExist:
        return JsonResponse({'error': 'lesson not found'}, status=404)
    except Role.DoesNotExist:
        return JsonResponse({'error': 'role not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)


@require_POST
def lesson_remove_allowed_role(request, lesson_id, role_id):
    """
    Удаляет должность из списка разрешенных для урока
    """
    if not request.user.is_staff:
        from django.http import JsonResponse
        return JsonResponse({'error': 'forbidden'}, status=403)
    
    from django.http import JsonResponse
    from builder.models import LessonAllowedRole
    from courses.models import Lesson
    
    try:
        lesson = Lesson.objects.get(id=lesson_id)
        allowed_role = LessonAllowedRole.objects.get(lesson=lesson, role_id=role_id)
        allowed_role.delete()
        
        return JsonResponse({'success': True})
        
    except Lesson.DoesNotExist:
        return JsonResponse({'error': 'lesson not found'}, status=404)
    except LessonAllowedRole.DoesNotExist:
        return JsonResponse({'error': 'allowed role not found'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

class UserUpdateView(UpdateView):
    model = User
    template_name = 'user_management/user_form.html'
    fields = ['email', 'first_name', 'last_name', 'groups', 'is_active']
    success_url = reverse_lazy('user_management:user_list')

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied("У вас нет доступа к управлению пользователями.")
        user_to_edit = self.get_object()
        if get_user_privilege_level(request.user) < get_user_privilege_level(user_to_edit):
            self.readonly = True
        else:
            self.readonly = False
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.POST:
            context['profile_form'] = UserProfileForm(self.request.POST, self.request.FILES, instance=self.object.profile, user_instance=self.object)
        else:
            context['profile_form'] = UserProfileForm(instance=self.object.profile, user_instance=self.object)
        context['readonly'] = getattr(self, 'readonly', False)
        context['roles'] = Role.objects.all()
        return context

    def form_valid(self, form):
        if getattr(self, 'readonly', False):
            raise PermissionDenied("Недостаточно прав для редактирования этого пользователя.")
        response = super().form_valid(form)
        profile_form = UserProfileForm(self.request.POST, self.request.FILES, instance=self.object.profile, user_instance=self.object)
        if profile_form.is_valid():
            profile_form.save()
        return response

class UserProgressDashboardView(DetailView):
    model = User
    template_name = 'user_management/user_progress_dashboard.html'
    context_object_name = 'target_user'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser or (hasattr(request.user, 'profile') and request.user.profile.is_mentor_user)):
            raise PermissionDenied("У вас нет доступа к управлению пользователями.")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.get_object()
        profile = user.profile
        
        # Получаем все доступные курсы через менеджер
        available_courses = Course.objects.available_for_user(user)
        # Получаем UserCourse для каждого доступного курса
        user_courses = []
        for course in available_courses:
            user_course = UserCourse.objects.filter(user=user, course=course).first()
            if user_course:
                user_courses.append(user_course)
            else:
                # Создаем UserCourse если его нет (для курсов из траекторий)
                user_course = UserCourse.objects.create(user=user, course=course, status='available')
                user_courses.append(user_course)
        
        # Получаем все результаты тестирования пользователя ДО цикла по курсам
        quiz_results = list(QuizResult.objects.filter(user=user).order_by('-completed_at'))
        
        # Подготавливаем данные о прогрессе для каждого курса
        courses_progress = []
        
        for user_course in user_courses:
            course = user_course.course
            
            # Получаем траекторию пользователя для этого курса
            trajectory = UserLessonTrajectory.objects.filter(user=user, course=course).first()
            
            if trajectory:
                # Используем уроки из траектории
                lessons = trajectory.lessons.all().order_by('order')
                total_lessons = lessons.count()
                lesson_ids = lessons.values_list('id', flat=True)
                
                completed_lessons = UserProgress.objects.filter(
                    user=user,
                    course=course,
                    completed=True,
                    lesson_id__in=lesson_ids
                ).count()
            else:
                # Используем все уроки курса
                lessons = course.lessons.all().order_by('order')
                total_lessons = lessons.count()
                completed_lessons = UserProgress.objects.filter(
                    user=user,
                    course=course,
                    completed=True
                ).count()
            
            # Подсчитываем завершенные тесты в рамках этого курса
            completed_quizzes = QuizResult.objects.filter(
                user=user,
                course=course,
                quiz_title__in=[quiz.name for quiz in course.quizzes.all()],
                passed=True
            ).count()
            total_quizzes = course.quizzes.count()
            
            # Вычисляем процент прогресса с учетом уроков и тестов
            total_materials = total_lessons + total_quizzes
            completed_materials = completed_lessons + completed_quizzes
            progress_percent = int((completed_materials / total_materials) * 100) if total_materials > 0 else 0
            
            # Проверяем прохождение финального теста в рамках этого курса
            quiz_passed = False
            if course.final_quiz:
                quiz_passed = QuizResult.objects.filter(
                    user=user,
                    course=course,
                    quiz_title=course.final_quiz.name,
                    passed=True
                ).exists()
            
            # Получаем детальную информацию об уроках и тестах
            materials_detail = []
            
            # Добавляем уроки
            for lesson in lessons:
                progress = UserProgress.objects.filter(
                    user=user,
                    lesson=lesson,
                    completed=True
                ).first()
                
                materials_detail.append({
                    'type': 'lesson',
                    'lesson': lesson,
                    'completed': progress is not None,
                    'completed_at': progress.completed_at if progress else None,
                    'order': lesson.order,
                    'title': lesson.title
                })
            
            # Добавляем тесты курса
            for quiz in course.quizzes.all():
                # Получаем все попытки теста
                quiz_attempts = QuizResult.objects.filter(
                    user=user,
                    course=course,
                    quiz_title=quiz.name
                ).order_by('-completed_at')
                
                # Проверяем, есть ли успешная попытка
                quiz_result = quiz_attempts.filter(passed=True).first()
                
                # Получаем лучшую попытку (по проценту, затем по дате)
                best_attempt = None
                if quiz_attempts.exists():
                    best_attempt = sorted(quiz_attempts, key=lambda x: (x.percent, x.completed_at), reverse=True)[0]
                
                materials_detail.append({
                    'type': 'quiz',
                    'quiz': quiz,
                    'completed': quiz_result is not None,
                    'completed_at': quiz_result.completed_at if quiz_result else None,
                    'order': quiz.order,
                    'title': quiz.name,
                    'attempts_count': quiz_attempts.count(),
                    'best_attempt': best_attempt
                })
            
            # Сортируем материалы по порядку
            materials_detail.sort(key=lambda x: x['order'])
            
            # Для совместимости с шаблоном сохраняем старое название
            lessons_detail = materials_detail
            
            best_attempt = None
            if course.final_quiz:
                attempts = [qr for qr in quiz_results if qr.quiz_title == course.final_quiz.name]
                if attempts:
                    best_attempt = sorted(attempts, key=lambda x: (x.percent, x.completed_at), reverse=True)[0]
            
            courses_progress.append({
                'course': course,
                'user_course': user_course,
                'total_lessons': total_lessons,
                'completed_lessons': completed_lessons,
                'total_quizzes': total_quizzes,
                'completed_quizzes': completed_quizzes,
                'total_materials': total_materials,
                'completed_materials': completed_materials,
                'progress_percent': progress_percent,
                'quiz_passed': quiz_passed,
                'lessons_detail': lessons_detail,
                'best_attempt': best_attempt,
            })
        
        # Общая статистика
        total_courses = len(courses_progress)
        completed_courses = len([cp for cp in courses_progress if cp['user_course'].status == 'completed'])
        started_courses = len([cp for cp in courses_progress if cp['user_course'].status == 'started'])
        available_courses = len([cp for cp in courses_progress if cp['user_course'].status == 'available'])
        
        total_materials_completed = sum(cp['completed_materials'] for cp in courses_progress)
        total_materials_available = sum(cp['total_materials'] for cp in courses_progress)
        overall_progress = int((total_materials_completed / total_materials_available) * 100) if total_materials_available > 0 else 0
        
        # Для совместимости с шаблоном сохраняем старые переменные
        total_lessons_completed = sum(cp['completed_lessons'] for cp in courses_progress)
        total_lessons_available = sum(cp['total_lessons'] for cp in courses_progress)
        


        # Детальная информация о результатах тестов
        detailed_quiz_results = []
        for quiz_result in quiz_results:
            # Получаем ответы пользователя для этого теста
            user_answers = UserAnswer.objects.filter(
                user=user,
                quiz_result=quiz_result
            ).select_related('question', 'selected_answer')
            
            # Группируем ответы по вопросам
            questions_data = []
            for answer in user_answers:
                question = answer.question
                
                # Получаем все варианты ответов для вопроса
                all_answers = question.answer_set.all()
                
                # Определяем правильные ответы
                correct_answers = all_answers.filter(is_correct=True)
                
                questions_data.append({
                    'question': question,
                    'user_answer': answer,
                    'all_answers': all_answers,
                    'correct_answers': correct_answers,
                    'is_correct': answer.is_correct,
                    'question_type': question.question_type
                })
            
            detailed_quiz_results.append({
                'quiz_result': quiz_result,
                'questions_data': questions_data,
                'total_questions': len(questions_data),
                'correct_answers_count': len([qd for qd in questions_data if qd['is_correct']])
            })
        
        # Пагинация по попыткам тестов
        paginator = Paginator(quiz_results, 4)
        page_number = self.request.GET.get('page', 1)
        try:
            page_obj = paginator.page(page_number)
        except PageNotAnInteger:
            page_obj = paginator.page(1)
        except EmptyPage:
            page_obj = paginator.page(paginator.num_pages)
        
        # Фильтрация курсов по статусу
        course_filter = self.request.GET.get('course_filter', 'all')
        if course_filter == 'completed':
            courses_progress = [cp for cp in courses_progress if cp['user_course'].status == 'completed']
        elif course_filter == 'started':
            courses_progress = [cp for cp in courses_progress if cp['user_course'].status == 'started']
        # Для 'all' и других значений показываем все курсы
        
        # Пагинация по курсам
        paginator_courses = Paginator(courses_progress, 4)
        page_number_courses = self.request.GET.get('courses_page', 1)
        try:
            page_obj_courses = paginator_courses.page(page_number_courses)
        except PageNotAnInteger:
            page_obj_courses = paginator_courses.page(1)
        except EmptyPage:
            page_obj_courses = paginator_courses.page(paginator_courses.num_pages)
        
        context.update({
            'courses_progress': courses_progress,
            'total_courses': total_courses,
            'completed_courses': completed_courses,
            'started_courses': started_courses,
            'available_courses': available_courses,
            'total_lessons_completed': total_lessons_completed,
            'total_lessons_available': total_lessons_available,
            'overall_progress': overall_progress,
            'detailed_quiz_results': detailed_quiz_results,
            'quiz_results': quiz_results,
            'page_obj': page_obj,
            'page_obj_courses': page_obj_courses,
            'course_filter': course_filter,
        })
        
        return context


class UserQuizReportView(MentorRequiredMixin, DetailView):
    model = QuizResult
    template_name = 'user_management/user_quiz_report.html'
    context_object_name = 'quiz_result'

    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_object(self):
        return QuizResult.objects.get(id=self.kwargs['quiz_id'])

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        quiz_result = self.get_object()
        answers = quiz_result.answers.select_related('question', 'selected_answer').all()
        # Группируем ответы по вопросам
        grouped = {}
        for ans in answers:
            grouped.setdefault(ans.question, []).append(ans)
        
        # Обработка типа "match" - парсим answer_text и подготавливаем данные для отображения
        from quizzes.models import Answer
        
        # Добавляем match_results непосредственно к каждому вопросу в grouped
        for question, user_answers in grouped.items():
            if question.question_type == 'match':
                user_answer = user_answers[0]  # Для match всегда один UserAnswer
                # Парсим answer_text вида "question_id:answer_id; question_id:answer_id"
                user_matches = {}
                if user_answer.answer_text:
                    for pair in user_answer.answer_text.split('; '):
                        if ':' in pair:
                            q_id, a_id = pair.split(':')
                            user_matches[q_id] = a_id
                
                # Получаем все ответы для этого вопроса (question и answer пары)
                all_answers = Answer.objects.filter(question=question).order_by('id')
                answers_list = list(all_answers)
                
                # Вопросы (неперетаскиваемые элементы справа)
                questions_dict = {}
                # Ответы (перетаскиваемые элементы слева)
                answers_dict = {}
                
                # Правильные соответствия
                correct_matches = {}
                
                # Группируем ответы по парам (каждые 2 ответа - это пара вопрос-ответ)
                for i in range(0, len(answers_list), 2):
                    if i < len(answers_list):
                        # Четные элементы - это вопросы
                        question_answer = answers_list[i]
                        question_key = str(question_answer.id)
                        questions_dict[question_key] = {
                            'text': question_answer.text,
                            'image': question_answer.image.url if question_answer.image else None
                        }
                    
                    if i + 1 < len(answers_list):
                        # Нечетные элементы - это ответы
                        answer_answer = answers_list[i + 1]
                        answers_dict[str(answer_answer.id)] = {
                            'text': answer_answer.text,
                            'image': answer_answer.image.url if answer_answer.image else None
                        }
                        
                        # Если оба отмечены как правильные, это правильная пара
                        if i < len(answers_list):
                            question_answer = answers_list[i]
                            if question_answer.is_correct and answer_answer.is_correct:
                                correct_matches[str(question_answer.id)] = str(answer_answer.id)
                
                # Если не нашли пары через is_correct, используем простую логику
                if not correct_matches:
                    for i in range(0, len(answers_list), 2):
                        if i + 1 < len(answers_list):
                            question_answer = answers_list[i]
                            answer_answer = answers_list[i + 1]
                            correct_matches[str(question_answer.id)] = str(answer_answer.id)
                
                # Формируем результаты для каждого вопроса
                match_results = []
                for question_id, question_data in questions_dict.items():
                    user_answer_id = user_matches.get(question_id, '')
                    correct_answer_id = correct_matches.get(question_id, '')
                    user_answer_data = answers_dict.get(user_answer_id, {'text': 'Неизвестный ответ', 'image': None})
                    correct_answer_data = answers_dict.get(correct_answer_id, {'text': 'Неизвестный ответ', 'image': None})
                    is_question_correct = (user_answer_id == correct_answer_id)
                    
                    match_results.append({
                        'question_text': question_data['text'],
                        'question_image': question_data['image'],
                        'user_answer_text': user_answer_data['text'],
                        'user_answer_image': user_answer_data['image'],
                        'correct_answer_text': correct_answer_data['text'],
                        'correct_answer_image': correct_answer_data['image'],
                        'is_correct': is_question_correct,
                    })
                
                # Добавляем match_results непосредственно к user_answer для удобства доступа в шаблоне
                user_answer.match_results = match_results
        
        context['grouped_answers'] = grouped
        return context


    def post(self, request, *args, **kwargs):
        quiz_result = self.get_object()
        for key, val in request.POST.items():
            if not key.startswith('text_eval_'):
                continue
            try:
                ua_id = int(key.replace('text_eval_', ''))
            except ValueError:
                continue
            if val == '':
                new_val = None
            elif val == 'true':
                new_val = True
            elif val == 'false':
                new_val = False
            else:
                continue
            quiz_result.answers.filter(id=ua_id, question__question_type='text').update(is_correct=new_val)
        return redirect(request.path)


class UserPasswordChangeView(FormView):
    template_name = 'user_management/user_password_change.html'
    form_class = SetPasswordForm
    success_url = reverse_lazy('user_management:user_list')

    def dispatch(self, request, *args, **kwargs):
        # Только staff или superuser могут менять пароли другим
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            raise PermissionDenied('У вас нет прав для смены пароля других пользователей.')
        # Можно добавить: запрет staff менять пароль superuser, если нужно
        return super().dispatch(request, *args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = User.objects.get(pk=self.kwargs['pk'])
        return kwargs

    def form_valid(self, form):
        form.save()  # set_password + save
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['object'] = User.objects.get(pk=self.kwargs['pk'])
        return context


class UserQuizAttemptsView(DetailView):
    """
    Страница управления доступом к тестам для пользователя.
    """
    model = User
    template_name = 'user_management/user_quiz_attempts.html'
    context_object_name = 'target_user'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser or request.user.profile.is_mentor_user):
            raise PermissionDenied("У вас нет доступа к управлению пользователями.")
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        user = self.get_object()
        
        # Получаем все тесты
        # quizzes = Quiz.objects.all().order_by('name') # TODO: Фильтровать по доступности
        
        # Получаем все курсы, доступные пользователю
        available_courses = Course.objects.available_for_user(user)

        # Собираем тесты из доступных курсов
        quizzes = set()
        for course in available_courses:
            if course.final_quiz:
                quizzes.add(course.final_quiz)
        
        # Добавляем тесты, которые напрямую доступны пользователю (если есть такая логика)
        # На данный момент считаем, что все тесты привязаны к курсам или доступны глобально.
        # Если в будущем появится логика для отдельно доступных тестов, нужно добавить ее сюда.
        # Пример: quizzes.update(Quiz.objects.filter(is_globally_available=True))

        quizzes = sorted(list(quizzes), key=lambda q: q.name) # Сортируем для единообразия
        
        quiz_data = {}
        for quiz in quizzes:
            # Получаем результаты тестов
            results = QuizResult.objects.filter(
                user=user,
                quiz_title=quiz.name
            ).order_by('-completed_at')
            
            # Просто показываем все результаты
            attempts_with_results = []
            for i, result in enumerate(results, 1):
                attempts_with_results.append({
                    'result': result,
                    'attempt_number': i,  # Порядковый номер результата
                    'started_at': result.completed_at,
                    'completed_at': result.completed_at
                })
            
            # Статистика
            total_attempts = results.count()  # Общее количество результатов
            passed_attempts = results.filter(passed=True).count()
            failed_attempts = results.filter(passed=False).count()
            best_result = results.order_by('-percent').first()
            best_score = best_result.percent if best_result else None
            
            # Проверяем блокировку теста
            quiz_lock = QuizLock.objects.filter(user=user, quiz=quiz).first()
            is_blocked = quiz_lock.is_locked if quiz_lock else False
            
            quiz_data[quiz] = {
                'attempts': attempts_with_results,
                'total_attempts': total_attempts,
                'passed_attempts': passed_attempts,
                'failed_attempts': failed_attempts,
                'best_score': best_score,
                'is_blocked': is_blocked
            }
        
        context['quiz_data'] = quiz_data
        return context



@require_POST  
def unlock_quiz_access(request, user_id, quiz_id):
    """
    Разблокирует доступ к тесту для пользователя.
    Позволяет пройти тест еще 1 раз сверх ограничения попыток.
    """
    if not request.user.is_staff or not request.user.is_superuser:
        raise PermissionDenied("У вас нет доступа к этому действию.")
    
    try:
        user = User.objects.get(id=user_id)
        quiz = Quiz.objects.get(id=quiz_id)
        
        # Разблокируем тест
        quiz_lock, created = QuizLock.objects.get_or_create(
            user=user,
            quiz=quiz,
            defaults={'is_locked': False}
        )
        
        if quiz_lock.is_locked:
            quiz_lock.is_locked = False
            quiz_lock.locked_at = None
            quiz_lock.save()
            
            # Восстанавливаем прогресс курса, если тест является финальным
            course = Course.objects.filter(final_quiz=quiz).first()
            if course:
                user_course = UserCourse.objects.filter(user=user, course=course).first()
                if user_course:
                    # Отмечаем все уроки курса как завершенные
                    from myapp.models import UserProgress
                    for lesson in course.lessons.all():
                        UserProgress.objects.update_or_create(
                            user=user,
                            course=course,
                            lesson=lesson,
                            defaults={'completed': True}
                        )
                    
                    # Устанавливаем статус курса как "начат" (не завершен, так как тест еще не пройден)
                    user_course.status = 'started'
                    user_course.save()
            
            messages.success(
                request,
                f'Тест "{quiz.name}" разблокирован для пользователя {user.get_full_name()}. '
                f'Прогресс курса восстановлен. Пользователь может пройти еще одну попытку.'
            )
        else:
            messages.info(
                request,
                f'Тест "{quiz.name}" уже разблокирован для пользователя {user.get_full_name()}.'
            )
        
    except (User.DoesNotExist, Quiz.DoesNotExist):
        messages.error(request, 'Пользователь или тест не найден.')
    except Exception as e:
        messages.error(request, f'Ошибка при разблокировке: {str(e)}')
    
    return redirect('user_management:user_quiz_attempts', pk=user_id)




# TODO: Сделать CBV для homework_check_dashboard.html



class AdminDashboardView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """CBV для отображения административной панели статистики пользователей по баллам DASCOIN"""
    model = User
    template_name = 'user_management/admin_dascoin_dashboard.html'
    context_object_name = 'users'
    paginate_by = 25
    ordering = ['-profile__dascoin_points', 'email']
    
    def test_func(self):
        """Проверяет, что пользователь является staff или superuser"""
        return self.request.user.is_staff or self.request.user.is_superuser or (hasattr(self.request.user, 'profile') and self.request.user.profile.is_mentor_user)
    
    def get_queryset(self):
        """Возвращает пользователей с фильтрацией"""
        queryset = User.objects.select_related('profile', 'profile__role').prefetch_related('groups').order_by('-profile__dascoin_points', 'email')
        
        # Если пользователь - наставник (но не superuser и не staff), показываем только его группу
        if (hasattr(self.request.user, 'profile') and 
            self.request.user.profile.is_mentor_user and 
            not self.request.user.is_superuser and 
            not self.request.user.is_staff):
            # Получаем группы наставника
            mentor_groups = self.request.user.groups.all()
            if mentor_groups.exists():
                # Показываем только пользователей из групп наставника
                queryset = queryset.filter(groups__in=mentor_groups).distinct()
            else:
                # Если у наставника нет групп, показываем пустой список
                queryset = queryset.none()
        
        # Фильтрация по группе (только для не-наставников)
        if not (hasattr(self.request.user, 'profile') and 
                self.request.user.profile.is_mentor_user and 
                not self.request.user.is_superuser and 
                not self.request.user.is_staff):
            group_id = self.request.GET.get('group')
            if group_id:
                queryset = queryset.filter(groups__id=group_id)
        
        # Фильтрация по должности
        role_id = self.request.GET.get('role')
        if role_id:
            queryset = queryset.filter(profile__role__id=role_id)
        
        # Фильтрация по минимальному количеству баллов
        points_min = self.request.GET.get('points_min')
        if points_min and points_min.isdigit():
            queryset = queryset.filter(profile__dascoin_points__gte=int(points_min))
        
        # Фильтрация по максимальному количеству баллов
        points_max = self.request.GET.get('points_max')
        if points_max and points_max.isdigit():
            queryset = queryset.filter(profile__dascoin_points__lte=int(points_max))
        
        # Быстрые фильтры
        zero_points = self.request.GET.get('zero_points')
        if zero_points:
            queryset = queryset.filter(profile__dascoin_points=0)
        
        approved_only = self.request.GET.get('approved')
        show_all = self.request.GET.get('show_all')
        # По умолчанию показываем только подтвержденных пользователей, если нет никаких параметров
        has_any_params = bool(self.request.GET)
        
        if approved_only == '1' or (not has_any_params and not show_all):
            queryset = queryset.filter(profile__is_approved=True)
        
        # Применяем distinct() до среза
        queryset = queryset.distinct()
        
        # Быстрый фильтр топ-N применяется после distinct()
        top_users = self.request.GET.get('top')
        if top_users and top_users.isdigit():
            queryset = queryset.order_by('-profile__dascoin_points')[:int(top_users)]
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Добавляет дополнительный контекст"""
        context = super().get_context_data(**kwargs)
        
        # Общая статистика
        if not (hasattr(self.request.user, 'profile') and 
                self.request.user.profile.is_mentor_user and 
                not self.request.user.is_superuser and 
                not self.request.user.is_staff):
            # Для обычных пользователей - общая статистика
            all_users = User.objects.select_related('profile')
            context['total_users'] = all_users.count()
            context['total_dascoin_points'] = all_users.aggregate(total=Sum('profile__dascoin_points'))['total'] or 0
            context['active_users'] = all_users.filter(is_active=True).count()
        else:
            # Для наставников - статистика только по их группам
            mentor_groups = self.request.user.groups.all()
            if mentor_groups.exists():
                all_users = User.objects.filter(groups__in=mentor_groups).select_related('profile').distinct()
                context['total_users'] = all_users.count()
                context['total_dascoin_points'] = all_users.aggregate(total=Sum('profile__dascoin_points'))['total'] or 0
                context['active_users'] = all_users.filter(is_active=True).count()
            else:
                context['total_users'] = 0
                context['total_dascoin_points'] = 0
                context['active_users'] = 0
        
        # Статистика по баллам DASCOIN
        from gamification.models import DascoinTransaction
        
        # Общее количество потраченных баллов (все списания)
        total_spent_points = DascoinTransaction.objects.filter(
            transaction_type='deduct'
        ).aggregate(total=Sum('points_change'))['total'] or 0
        context['total_spent_points'] = abs(total_spent_points)  # Берем абсолютное значение
        
        # Время последнего начисления баллов
        last_award_transaction = DascoinTransaction.objects.filter(
            transaction_type='award'
        ).order_by('-created_at').first()
        
        if last_award_transaction:
            context['last_award_date'] = last_award_transaction.created_at
            context['last_award_user'] = last_award_transaction.user
            context['last_award_points'] = last_award_transaction.points_change
        else:
            context['last_award_date'] = None
            context['last_award_user'] = None
            context['last_award_points'] = None
        
        # Группы и должности для фильтров
        if not (hasattr(self.request.user, 'profile') and 
                self.request.user.profile.is_mentor_user and 
                not self.request.user.is_superuser and 
                not self.request.user.is_staff):
            context['groups'] = Group.objects.all().order_by('name')
        else:
            # Для наставников показываем только их группы
            context['groups'] = self.request.user.groups.all().order_by('name')
        context['roles'] = Role.objects.all().order_by('name')
        
        # Параметры фильтрации
        context['selected_group'] = self.request.GET.get('group')
        context['selected_role'] = self.request.GET.get('role')
        context['points_min'] = self.request.GET.get('points_min')
        context['points_max'] = self.request.GET.get('points_max')
        
        # Флаги быстрых фильтров
        has_any_params = bool(self.request.GET)
        show_all = self.request.GET.get('show_all')
        
        context['any_filter'] = has_any_params
        context['top_users'] = bool(self.request.GET.get('top'))
        context['zero_points'] = bool(self.request.GET.get('zero_points'))
        # approved_only активен по умолчанию (нет параметров) или когда явно передан approved=1
        context['approved_only'] = (self.request.GET.get('approved') == '1') or (not has_any_params and not show_all)
        context['show_all'] = bool(show_all)
        
        # Параметры для пагинации
        query_params = self.request.GET.copy()
        if 'page' in query_params:
            del query_params['page']
        context['query_params'] = '&' + urlencode(query_params) if query_params else ''
        
        # Логирование действия
        audit_logger.info(
            'Просматривает административную панель статистики DASCOIN', 
            extra={
                'user': self.request.user.email if self.request.user.is_authenticated else 'Anonymous'
            }
        )
        
        return context


@login_required
def export_admin_stats_excel(request):
    """Экспорт статистики администратора в Excel"""
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponse("Доступ запрещен", status=403)
    
    # Получаем данные с теми же фильтрами, что и в dashboard
    queryset = User.objects.select_related('profile', 'profile__role').prefetch_related('groups')
    
    # Применяем фильтры
    group_id = request.GET.get('group')
    if group_id:
        queryset = queryset.filter(groups__id=group_id)
    
    role_id = request.GET.get('role')
    if role_id:
        queryset = queryset.filter(profile__role__id=role_id)
    
    points_min = request.GET.get('points_min')
    if points_min and points_min.isdigit():
        queryset = queryset.filter(profile__dascoin_points__gte=int(points_min))
    
    points_max = request.GET.get('points_max')
    if points_max and points_max.isdigit():
        queryset = queryset.filter(profile__dascoin_points__lte=int(points_max))
    
    # Быстрые фильтры
    zero_points = request.GET.get('zero_points')
    if zero_points:
        queryset = queryset.filter(profile__dascoin_points=0)
    
    approved_only = request.GET.get('approved')
    show_all = request.GET.get('show_all')
    # По умолчанию показываем только подтвержденных пользователей, если нет никаких параметров
    has_any_params = bool(request.GET)
    
    if approved_only == '1' or (not has_any_params and not show_all):
        queryset = queryset.filter(profile__is_approved=True)
    
    # Применяем distinct() до среза
    queryset = queryset.distinct()
    
    # Быстрый фильтр топ-N применяется после distinct()
    top_users = request.GET.get('top')
    if top_users and top_users.isdigit():
        queryset = queryset.order_by('-profile__dascoin_points')[:int(top_users)]
    else:
        queryset = queryset.order_by('-profile__dascoin_points', 'email')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Статистика DASCOIN"
    
    headers = ['Пользователь', 'Email', 'Группы', 'Должность', 'DASCOIN', 'Статус', 'Дата регистрации']
    ws.append(headers)
    
    # Стили для заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for user in queryset:
        groups = ', '.join([group.name for group in user.groups.all()])
        role = user.profile.role.name if user.profile.role else ''
        
        if user.is_active:
            if user.profile.is_approved:
                status = 'Активен'
            else:
                status = 'Ожидает подтверждения'
        else:
            status = 'Неактивен'
        
        ws.append([
            user.get_full_name() or user.email,
            user.email,
            groups,
            role,
            user.profile.dascoin_points,
            status,
            user.date_joined.strftime("%d.%m.%Y %H:%M")
        ])
    
    # Автоматическая ширина колонок
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"admin_stats_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    
    audit_logger.info(
        'Экспортировал статистику администратора в Excel', 
        extra={
            'user': request.user.email if request.user.is_authenticated else 'Anonymous'
        }
    )
    return response


@login_required
def export_admin_stats_pdf(request):
    """Экспорт статистики администратора в PDF"""
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponse("Доступ запрещен", status=403)
    
    # Получаем данные с теми же фильтрами
    queryset = User.objects.select_related('profile', 'profile__role').prefetch_related('groups')
    
    # Применяем фильтры
    group_id = request.GET.get('group')
    if group_id:
        queryset = queryset.filter(groups__id=group_id)
    
    role_id = request.GET.get('role')
    if role_id:
        queryset = queryset.filter(profile__role__id=role_id)
    
    points_min = request.GET.get('points_min')
    if points_min and points_min.isdigit():
        queryset = queryset.filter(profile__dascoin_points__gte=int(points_min))
    
    points_max = request.GET.get('points_max')
    if points_max and points_max.isdigit():
        queryset = queryset.filter(profile__dascoin_points__lte=int(points_max))
    
    # Быстрые фильтры
    zero_points = request.GET.get('zero_points')
    if zero_points:
        queryset = queryset.filter(profile__dascoin_points=0)
    
    approved_only = request.GET.get('approved')
    show_all = request.GET.get('show_all')
    # По умолчанию показываем только подтвержденных пользователей, если нет никаких параметров
    has_any_params = bool(request.GET)
    
    if approved_only == '1' or (not has_any_params and not show_all):
        queryset = queryset.filter(profile__is_approved=True)
    
    # Применяем distinct() до среза
    queryset = queryset.distinct()
    
    # Быстрый фильтр топ-N применяется после distinct()
    top_users = request.GET.get('top')
    if top_users and top_users.isdigit():
        queryset = queryset.order_by('-profile__dascoin_points')[:int(top_users)]
    else:
        queryset = queryset.order_by('-profile__dascoin_points', 'email')
    
    # Общая статистика
    all_users = User.objects.select_related('profile')
    total_users = all_users.count()
    total_dascoin_points = all_users.aggregate(total=Sum('profile__dascoin_points'))['total'] or 0
    active_users = all_users.filter(is_active=True).count()
    
    # Статистика по баллам DASCOIN
    total_spent_points = DascoinTransaction.objects.filter(
        transaction_type='deduct'
    ).aggregate(total=Sum('points_change'))['total'] or 0
    total_spent_points = abs(total_spent_points)
    
    # Время последнего начисления баллов
    last_award_transaction = DascoinTransaction.objects.filter(
        transaction_type='award'
    ).order_by('-created_at').first()
    
    last_award_date = None
    if last_award_transaction:
        last_award_date = last_award_transaction.created_at
    
    html_string = render_to_string('user_management/admin_stats_pdf.html', {
        'users': queryset,
        'total_users': total_users,
        'total_dascoin_points': total_dascoin_points,
        'active_users': active_users,
        'total_spent_points': total_spent_points,
        'last_award_date': last_award_date,
        'generated_at': datetime.now(),
        'generated_by': request.user.get_full_name() or request.user.email,
    })
    
    html = HTML(string=html_string)
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = f"admin_stats_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    audit_logger.info(
        'Экспортировал статистику администратора в PDF', 
        extra={
            'user': request.user.email if request.user.is_authenticated else 'Anonymous'
        }
    )
    return response


class AdminUserTransactionsView(LoginRequiredMixin, UserPassesTestMixin, ListView):
    """CBV для отображения истории транзакций DASCOIN конкретного пользователя администратором"""
    model = DascoinTransaction
    template_name = 'user_management/admin_user_transactions.html'
    context_object_name = 'transactions'
    paginate_by = 20
    ordering = ['-created_at']
    
    def test_func(self):
        """Проверяет, что пользователь является staff или superuser"""
        return self.request.user.is_staff or self.request.user.is_superuser
    
    def get_queryset(self):
        """Возвращает транзакции конкретного пользователя с возможностью фильтрации"""
        user_id = self.kwargs.get('user_id')
        self.user = get_object_or_404(User, id=user_id)
        
        queryset = DascoinTransaction.objects.filter(user=self.user).order_by('-created_at')
        
        # Фильтрация по типу транзакции
        transaction_type = self.request.GET.get('type')
        if transaction_type and transaction_type in ['award', 'deduct', 'set', 'correction']:
            queryset = queryset.filter(transaction_type=transaction_type)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        """Добавляет дополнительный контекст"""
        context = super().get_context_data(**kwargs)
        
        context['user'] = self.user
        context['total_transactions'] = self.get_queryset().count()
        context['current_filter'] = self.request.GET.get('type', '')
        
        # Статистика по типам транзакций
        all_transactions = DascoinTransaction.objects.filter(user=self.user)
        context['stats'] = {
            'award': all_transactions.filter(transaction_type='award').count(),
            'deduct': all_transactions.filter(transaction_type='deduct').count(),
            'set': all_transactions.filter(transaction_type='set').count(),
            'correction': all_transactions.filter(transaction_type='correction').count(),
        }
        
        # Логирование действия
        audit_logger.info(
            f'Просматривает историю транзакций пользователя {self.user.email}', 
            extra={
                'user': self.request.user.email if self.request.user.is_authenticated else 'Anonymous',
                'target_user': self.user.email
            }
        )
        
        return context


@login_required
def export_admin_user_transactions_excel(request, user_id):
    """Экспорт транзакций конкретного пользователя в Excel"""
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponse("Доступ запрещен", status=403)
    
    user = get_object_or_404(User, id=user_id)
    transactions = DascoinTransaction.objects.filter(user=user).order_by('-created_at')
    
    # Применяем фильтры
    transaction_type = request.GET.get('type')
    if transaction_type and transaction_type in ['award', 'deduct', 'set', 'correction']:
        transactions = transactions.filter(transaction_type=transaction_type)
    
    wb = Workbook()
    ws = wb.active
    ws.title = f"Транзакции {user.email}"
    
    headers = ['Дата', 'Тип', 'Изменение', 'До', 'После', 'Причина', 'Администратор']
    ws.append(headers)
    
    # Стили для заголовков
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    for tx in transactions:
        ws.append([
            tx.created_at.strftime("%d.%m.%Y %H:%M"),
            tx.get_transaction_type_display(),
            tx.points_change,
            tx.points_before,
            tx.points_after,
            tx.reason or "Не указана",
            tx.admin_user.get_full_name() if tx.admin_user else "Система"
        ])
    
    # Автоматическая ширина колонок
    for column in ws.columns:
        max_length = max(len(str(cell.value)) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max_length + 2, 50)
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"transactions_{user.email}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    
    audit_logger.info(
        f'Экспортировал транзакции пользователя {user.email} в Excel', 
        extra={
            'user': request.user.email if request.user.is_authenticated else 'Anonymous',
            'target_user': user.email
        }
    )
    return response


@login_required
def export_admin_user_transactions_pdf(request, user_id):
    """Экспорт транзакций конкретного пользователя в PDF"""
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponse("Доступ запрещен", status=403)
    
    user = get_object_or_404(User, id=user_id)
    transactions = DascoinTransaction.objects.filter(user=user).order_by('-created_at')
    
    # Применяем фильтры
    transaction_type = request.GET.get('type')
    if transaction_type and transaction_type in ['award', 'deduct', 'set', 'correction']:
        transactions = transactions.filter(transaction_type=transaction_type)
    
    html_string = render_to_string('users/transactions_pdf.html', {
        'transactions': transactions,
        'user': user,
        'generated_at': datetime.now(),
        'total_transactions': transactions.count(),
        'is_admin_view': True,
        'generated_by': request.user.get_full_name() or request.user.email,
    })
    
    html = HTML(string=html_string)
    pdf = html.write_pdf()
    response = HttpResponse(pdf, content_type='application/pdf')
    filename = f"transactions_{user.email}_{datetime.now().strftime('%Y%m%d_%H%M')}.pdf"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    
    audit_logger.info(
        f'Экспортировал транзакции пользователя {user.email} в PDF', 
        extra={
            'user': request.user.email if request.user.is_authenticated else 'Anonymous',
            'target_user': user.email
        }
    )
    return response


class HomeworkCheckDashboardView(LoginRequiredMixin, UserPassesTestMixin, TemplateView):
    """
    Страница проверки заданий для наставников - простая статистика
    """
    template_name = 'user_management/homework_check_dashboard.html'
    
    def test_func(self):
        """Проверяет права доступа"""
        if not self.request.user.is_authenticated:
            return False
        
        # Суперпользователи и персонал имеют доступ
        if self.request.user.is_superuser or self.request.user.is_staff:
            return True
        
        # Наставники имеют доступ
        try:
            return self.request.user.profile.is_mentor_user
        except:
            return False
    
    def get_context_data(self, **kwargs):
        """Добавляет статистику в контекст"""
        context = super().get_context_data(**kwargs)
        
        # Импортируем модели
        from courses.models import Lesson
        from quizzes.models import Quiz
        from django.contrib.auth.models import Group
        
        # Проверяем, является ли пользователь суперпользователем или стафом
        is_admin = self.request.user.is_superuser or self.request.user.is_staff
        
        if is_admin:
            # Для администраторов - общая статистика по всей платформе
            total_lessons = Lesson.objects.count()
            total_quizzes = Quiz.objects.count()
            total_materials = total_lessons + total_quizzes
            active_users = User.objects.filter(profile__is_approved=True).count()
            total_groups = Group.objects.count()
        else:
            # Для наставников - статистика только по их группам
            mentor_groups = self.request.user.groups.all()
            
            if mentor_groups.exists():
                # Получаем пользователей из групп наставника
                mentor_group_users = User.objects.filter(groups__in=mentor_groups).distinct()
                
                # Получаем курсы, на которые записаны пользователи из групп наставника
                mentor_courses = Course.objects.filter(usercourse__user__groups__in=mentor_groups).distinct()
                
                # Уроки из курсов наставника
                total_lessons = Lesson.objects.filter(courses__in=mentor_courses).distinct().count()
                
                # Тесты из курсов наставника
                total_quizzes = Quiz.objects.filter(courses__in=mentor_courses).distinct().count()
                
                total_materials = total_lessons + total_quizzes
                
                # Активные пользователи из групп наставника
                active_users = mentor_group_users.filter(profile__is_approved=True).count()
                
                # Количество групп наставника
                total_groups = mentor_groups.count()
            else:
                # Если у наставника нет групп, показываем нули
                total_lessons = 0
                total_quizzes = 0
                total_materials = 0
                active_users = 0
                total_groups = 0
        
        context.update({
            'total_materials': total_materials,
            'total_lessons': total_lessons,
            'total_quizzes': total_quizzes,
            'active_users': active_users,
            'total_groups': total_groups,
            'is_admin': is_admin,
        })
        
        return context

