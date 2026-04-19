from datetime import datetime, timedelta
from django.contrib import messages
from django.contrib.auth import login
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import Group, User
from django.core.cache import cache
from django.db.models import Count, Q, F
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse, reverse_lazy
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.generic import CreateView, ListView, UpdateView, View
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
from openpyxl.chart.label import DataLabelList
from weasyprint.css.validation.properties import word_break

from builder.audit_logger import AuditLoggerMixin, serialize_model_data
from builder.forms import IncidentForm
from builder.models import Incident
from builder.utils import get_total_incidents_students, PageCacheMixin
from courses.models import Course, UserLessonTrajectory
from myapp.models import UserCourse, UserProgress, ManualCourseUnassignment, QuizResult
from myapp.views import is_admin
from users.models import Department
from quizzes.models import HomeworkSubmission

import json
import logging

logger = logging.getLogger(__name__)
audit_logger = logging.getLogger('audit')




class IncidentListView(ListView):
    """
    Список инцидентов с фильтрацией и быстрым просмотром.
    Ответ кэшируется по пользователю и полному URL (включая GET-параметры).
    """
    model = Incident
    template_name = 'builder/incidents/incidents.html'
    context_object_name = 'incidents'
    ordering = ['-created_at']


    def dispatch(self, request, *args, **kwargs):
        # Только staff/superuser
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser or request.user.profile.is_mentor_user):
            return render(request, '403.html', status=403)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        from django.utils import timezone
        import datetime
        
        queryset = super().get_queryset()

        # pyright: reportUnreachable=false
        # Оптимизация: предзагрузка ManyToMany полей
        queryset = queryset.prefetch_related('assigned_to', 'violators').select_related('user')

        # Фильтр по дате создания
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        date_from_datetime = None
        date_to_datetime = None
        
        # Если нет параметров в GET запросе (первичная загрузка), устанавливаем дефолтные значения
        if not self.request.GET:
            # Устанавливаем период с начала 2025 года до сегодняшней даты
            date_from = '2025-01-01'
            date_to = timezone.now().date().strftime('%Y-%m-%d')
        
        if date_from:
            date_from_parsed = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
            date_from_datetime = timezone.make_aware(datetime.datetime.combine(date_from_parsed, datetime.time.min))
            queryset = queryset.filter(created_at__gte=date_from_datetime)
        if date_to:
            date_to_parsed = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            date_to_datetime = timezone.make_aware(datetime.datetime.combine(date_to_parsed, datetime.time.max))
            queryset = queryset.filter(created_at__lte=date_to_datetime)
        
        # Фильтр по статусу
        statuses = self.request.GET.getlist('status')
        
        # Если нет параметров в GET запросе (первичная загрузка), устанавливаем дефолтные статусы
        if not self.request.GET:
            statuses = ['new', 'accepted', 'assigned', 'studies_completed']
        
        if statuses:
            queryset = queryset.filter(status__in=statuses)
        
        # Фильтр по типу
        incident_type = self.request.GET.get('incident_type')
        if incident_type:
            queryset = queryset.filter(incident_type=incident_type)
        
        # Добавляем аннотации для подсчета назначенных и завершивших курс пользователей
        # Считаем только пользователей из assigned_to (сигнал синхронизирует доступ к курсу)
        # Исключаем админов, суперпользователей и авторов курса
        queryset = queryset.annotate(
            assigned_users_count=Count(
                'assigned_to',
                distinct=True,
                filter=Q(
                    course__isnull=False,
                    assigned_to__is_staff=False,
                    assigned_to__is_superuser=False
                ) & ~Q(assigned_to=F('course__author'))
            ),
            completed_users_count=Count(
                'assigned_to',
                distinct=True,
                filter=Q(
                    course__isnull=False,
                    course__usercourse__status='completed',
                    course__usercourse__user=F('assigned_to'),
                    assigned_to__is_staff=False,
                    assigned_to__is_superuser=False
                ) & ~Q(assigned_to=F('course__author'))
            )
        )
        
        # Обновляем статусы инцидентов с неоцененными открытыми ответами
        # Делаем это только для инцидентов с курсами, чтобы не делать лишних запросов
        # Проверяем до применения фильтров по статусу, чтобы не пропустить инциденты
        from builder.signals import check_and_update_incident_studies_completed_status
        from builder.models import Incident
        
        # Получаем все инциденты с курсами, которые могут иметь неоцененные ответы
        # Проверяем только те, которые могут быть в текущем queryset (по датам)
        incidents_to_check = Incident.objects.filter(
            course__isnull=False,
            status__in=['new', 'accepted', 'assigned', 'studies_completed']
        )
        
        # Применяем фильтры по дате, если они есть
        if date_from_datetime:
            incidents_to_check = incidents_to_check.filter(created_at__gte=date_from_datetime)
        if date_to_datetime:
            incidents_to_check = incidents_to_check.filter(created_at__lte=date_to_datetime)
        
        # Обновляем статусы
        for incident in incidents_to_check:
            try:
                check_and_update_incident_studies_completed_status(incident)
            except Exception as e:
                import logging
                logger = logging.getLogger(__name__)
                logger.error(f"Ошибка при проверке статуса инцидента {incident.id}: {e}")
        
        return queryset


    def get_context_data(self, **kwargs):
        from django.utils import timezone
        
        context = super().get_context_data(**kwargs)
        readonly = False
        is_mentor = self.request.user.profile.is_mentor_user and not self.request.user.is_staff and not self.request.user.is_superuser
        if is_mentor:
            readonly = True
        context['readonly'] = readonly
        context['status_choices'] = Incident.STATUS_CHOICES
        context['incident_type_choices'] = Incident.INCIDENT_TYPE_CHOICES
        context['now'] = timezone.now()  # Текущая дата и время для проверки просроченных дедлайнов
        
        # Если нет параметров в GET запросе (первичная загрузка), устанавливаем дефолтные значения
        if not self.request.GET:
            context['selected_statuses'] = ['new', 'accepted', 'assigned', 'studies_completed']
            context['selected_incident_type'] = ''
            context['date_from'] = '2025-01-01'
            context['date_to'] = timezone.now().date().strftime('%Y-%m-%d')
        else:
            # Передаем текущие значения фильтров в контекст
            context['selected_statuses'] = self.request.GET.getlist('status', [])
            context['selected_incident_type'] = self.request.GET.get('incident_type', '')
            context['date_from'] = self.request.GET.get('date_from', '')
            context['date_to'] = self.request.GET.get('date_to', '')
        
        return context


def _apply_header_style(ws, row_num, col_count):
    """Применяет стили к заголовкам таблицы."""
    for col_num in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal='center')


def _set_column_widths(ws, widths):
    """Устанавливает ширину колонок."""
    for col_num, width in enumerate(widths, 1):
        col_letter = get_column_letter(col_num)
        ws.column_dimensions[col_letter].width = width


def _insert_rows(ws, rows):
    """Устанавливает пустые строки"""
    
    return ws.append([])
            
        


@login_required
def incidents_export_excel_report(request):
    if not (request.user.is_staff or request.user.is_superuser):
        return HttpResponse("Доступ запрещен", status=403)

    incidents = Incident.objects.all()
    wb = Workbook()
    
    # ================== ЛИСТ 1: Общая сводка ==================
    ws_summary = wb.active
    ws_summary.title = "Общая сводка"
    
    # Статистика за последнюю неделю
    from django.utils import timezone
    now = timezone.now()
    week_ago = now - timedelta(days=7)
    
    # Инциденты, созданные за последнюю неделю
    incidents_last_week = incidents.filter(created_at__gte=week_ago)
    total_last_week = incidents_last_week.count()
    
    # Сколько из них в статусе "Назначен" (назначен курс-инцидент)
    assigned_last_week = incidents_last_week.filter(status='assigned').count()
    
    # Сколько из них в статусе "Принят" и "Новый"
    accepted_and_new_last_week = incidents_last_week.filter(status__in=['accepted', 'new']).count()

    # Сколько из них в статусе "Обуч. завершено" и "Завершено"
    completed_last_week = incidents_last_week.filter(status__in=['studies_completed', 'resolved']).count()
    
    # Количество назначений курсов-инцидентов для инцидентов за последнюю неделю
    incidents_with_course_last_week = incidents_last_week.filter(course__isnull=False)
    course_ids_last_week = incidents_with_course_last_week.values_list('course_id', flat=True).distinct()
    course_assignments_last_week = UserCourse.objects.filter(
        course_id__in=course_ids_last_week
    ).count()
    

    # Добавляем статистику за последнюю неделю в начало листа
    ws_summary.append(["Статистика за последнюю неделю"])
    # Применяем стиль только к первому столбцу первой строки
    header_cell = ws_summary.cell(row=1, column=1)
    header_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_cell.font = Font(bold=True, size=12, color="FFFFFF")
    
    week_stats_headers = [
        "Создано за неделю",
        "Назначено курс (статус 'Назначен')",
        "В статусе 'Принят' и 'Новый'",
        "Обуч. завершено и завершено",
        "Назначений курсов-инцидентов"
    ]
    ws_summary.append(week_stats_headers)
    week_stats_header_row = ws_summary.max_row
    _apply_header_style(ws_summary, 2, len(week_stats_headers))
    ws_summary.row_dimensions[week_stats_header_row].height = 30
    for col_num in range(1, len(week_stats_headers) + 1):
        cell = ws_summary.cell(row=week_stats_header_row, column=col_num)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _set_column_widths(ws_summary, [40, 30, 30, 30, 30, 25, 20])

    # _set_column_widths(ws_summary, [45, 40, 40, 40, 40, 25, 20])
    
    ws_summary.append([
        total_last_week,
        assigned_last_week,
        accepted_and_new_last_week,
        completed_last_week,
        course_assignments_last_week
    ])
    
    # Отступ перед общей сводкой
    ws_summary.append([])
    
    # Подсчёт данных
    total_incidents = incidents.count()
    info_incidents_count = incidents.filter(incident_type='informational').count()
    edu_incidents_count = incidents.filter(incident_type='educational').count()
    
    # Всего назначений (assigned_to + violators)
    total_assignments = get_total_incidents_students(incidents)
    
    # Уникальные назначения (каждый пользователь считается 1 раз)
    unique_assigned_users = set()
    for incident in incidents:
        unique_assigned_users.update(incident.assigned_to.values_list('id', flat=True))
        unique_assigned_users.update(incident.violators.values_list('id', flat=True))
    total_unique_assignments = len(unique_assigned_users)
    
    # Завершённые обучения по курсам-инцидентам
    incident_courses = Course.objects.filter(is_incident=True)
    total_completed = UserCourse.objects.filter(
        course__in=incident_courses,
        status='completed'
    ).count()
    
    # Статусы инцидентов
    status_counts = {
        'new': incidents.filter(status='new').count(),
        'accepted': incidents.filter(status='accepted').count(),
        'assigned': incidents.filter(status='assigned').count(),
        'studies_completed': incidents.filter(status='studies_completed').count(),
        'resolved': incidents.filter(status='resolved').count(),
        'declined': incidents.filter(status='declined').count(),
    }
    
    # Заголовки и данные для общей сводки
    summary_headers = [
        "Всего инцидентов", "Информационных", "Обучающих",
        "Всего назначений", "Уникальных назначений", "Завершено обучений"
    ]
    ws_summary.append(summary_headers)
    # Применяем стиль к строке 5 (после статистики за неделю: строка 1, строка 2, строка 3, строка 4 пустая, строка 5 - заголовки)
    _apply_header_style(ws_summary, ws_summary.max_row, len(summary_headers))
    # _set_column_widths(ws_summary, [50, 30, 30, 20, 22, 22])
    
    ws_summary.append([
        total_incidents, info_incidents_count, edu_incidents_count,
        total_assignments, total_unique_assignments, total_completed
    ])
    
    # Отступ и статусы инцидентов
    ws_summary.append([])
    ws_summary.append(["Статусы инцидентов"])
    # Применяем стиль только к первому столбцу строки со статусами
    status_title_row = ws_summary.max_row
    status_title_cell = ws_summary.cell(row=status_title_row, column=1)
    status_title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    status_title_cell.font = Font(bold=True, color="FFFFFF")
    
    status_headers = ["Новый", "Принят", "Назначен", "Обучение завершено", "Завершён", "Отклонён"]
    ws_summary.append(status_headers)
    # Применяем стиль к заголовкам статусов (строка после "Статусы инцидентов")
    _apply_header_style(ws_summary, ws_summary.max_row, len(status_headers))
    
    ws_summary.append([
        status_counts['new'], status_counts['accepted'], status_counts['assigned'],
        status_counts['studies_completed'], status_counts['resolved'], status_counts['declined']
    ])

    # Подразделения и инциденты для блоков «Просрочены дедлайны» и «По подразделениям»
    departments_involved = list(
        Department.objects.filter(profile__user__id__in=unique_assigned_users)
        .values_list('name', flat=True)
        .distinct()
    )
    incidents_prefetched = incidents.prefetch_related('assigned_to', 'violators')

    # Просрочены дедлайны по подразделениям: считаем по UserCourse.deadline (срок курса у пользователя), не по Incident.deadline
    # now уже определен выше для статистики за неделю
    incidents_with_course = [inc for inc in incidents_prefetched if inc.course_id is not None]
    # По подразделению: (department_name, set(user_ids) просрочивших, set(incident_ids) просроченных для подразделения)
    overdue_by_department = {}
    for department_name in departments_involved:
        department_user_ids = set(
            User.objects.filter(profile__department__name=department_name)
            .filter(id__in=unique_assigned_users)
            .values_list('id', flat=True)
        )
        overdue_user_ids = set()
        overdue_incident_ids = set()
        for inc in incidents_with_course:
            assigned_ids = set(inc.assigned_to.values_list('id', flat=True))
            violator_ids = set(inc.violators.values_list('id', flat=True))
            in_department = (assigned_ids | violator_ids) & department_user_ids
            if not in_department:
                continue
            # Просрочили: у пользователя есть UserCourse по курсу инцидента с дедлайном < now и статус не 'completed'
            overdue_user_ids_for_inc = set(
                UserCourse.objects.filter(
                    user_id__in=in_department,
                    course_id=inc.course_id,
                    deadline__isnull=False,
                    deadline__lt=now,
                ).exclude(status='completed').values_list('user_id', flat=True)
            )
            if overdue_user_ids_for_inc:
                overdue_incident_ids.add(inc.id)
                overdue_user_ids |= overdue_user_ids_for_inc
        overdue_by_department[department_name] = (len(overdue_user_ids), len(overdue_incident_ids))
    # Сортируем подразделения по количеству просроченных инцидентов (убывание)
    overdue_rows = [
        [department_name, num_employees, num_incidents]
        for department_name, (num_employees, num_incidents) in sorted(
            overdue_by_department.items(), key=lambda x: -x[1][1]
        )
    ]

    ws_summary.append([])
    ws_summary.append(['Просрочены дедлайны в подразделении', 'Количество сотрудников', 'Количество инцидентов'])
    _apply_header_style(ws_summary, ws_summary.max_row, 3)
    for row in overdue_rows:
        ws_summary.append(row)
    ws_summary.append([])
    ws_summary.append([])
    ws_summary.append(['По подразделениям'])
    _apply_header_style(ws_summary, ws_summary.max_row, 1)

    departments_headers = ["Подразделение", "Всего", "Обучающие", "Информационные", "Завершены", "Не завершены", "Повторяющиеся"]
    ws_summary.append(departments_headers)
    _apply_header_style(ws_summary, ws_summary.max_row, 8)

    start_row = ws_summary.max_row + 1
    for idx, department_name in enumerate(departments_involved):
        department_user_ids = set(
            User.objects.filter(profile__department__name=department_name)
            .filter(id__in=unique_assigned_users)
            .values_list('id', flat=True)
        )
        if not department_user_ids:
            ws_summary.append([department_name, 0, 0, 0, 0, 0, "Не выявлено"])
            row_num = start_row + idx
            for col in range(1,8):
                ws_summary.cell(row=row_num, column=col).alignment = Alignment(wrap_text=True)
            continue

        department_total = 0
        department_edu = 0
        department_info = 0
        titles_list = []

        for inc in incidents_prefetched:
            assigned_ids = set(inc.assigned_to.values_list('id', flat=True))
            violator_ids = set(inc.violators.values_list('id', flat=True))
            in_department = (assigned_ids | violator_ids) & department_user_ids
            n = len(in_department)
            if n == 0:
                continue
            department_total += n
            if inc.incident_type == 'educational':
                department_edu += n
            else:
                department_info += n
            titles_list.append(inc.title)

        completed = UserCourse.objects.filter(
            user_id__in=department_user_ids,
            course__in=incident_courses,
            status='completed'
        ).count()
        not_completed = max(0, department_total - completed)
        
        has_duplicate_titles = len(titles_list) != len(set(titles_list))
        repeat_value = "Выявлено" if has_duplicate_titles else "Не выявлено"

        ws_summary.append([
            department_name,
            department_total,
            department_edu,
            department_info,
            completed,
            not_completed,
            repeat_value,
        ])


    # ================== ЛИСТ 2: Диаграммы ==================
    ws_charts = wb.create_sheet("Диаграммы")
    
    # --- Данные для круговой диаграммы: Типы инцидентов ---
    ws_charts.append(["Тип инцидента", "Количество"])
    ws_charts.append(["Информационные", info_incidents_count])
    ws_charts.append(["Обучающие", edu_incidents_count])
    
    # Круговая диаграмма: Типы инцидентов
    pie_chart = PieChart()
    pie_chart.title = "Типы инцидентов"
    labels = Reference(ws_charts, min_col=1, min_row=2, max_row=3)
    data = Reference(ws_charts, min_col=2, min_row=1, max_row=3)
    pie_chart.add_data(data, titles_from_data=True)
    pie_chart.set_categories(labels)
    pie_chart.dataLabels = DataLabelList()
    pie_chart.dataLabels.showPercent = True
    pie_chart.dataLabels.showVal = True
    pie_chart.dataLabels.showCatName = False
    pie_chart.width = 12
    pie_chart.height = 8
    ws_charts.add_chart(pie_chart, "D2")
    
    # --- Данные для столбчатой диаграммы: Статусы инцидентов ---
    ws_charts.append([])  # Пустая строка
    ws_charts.append(["Статус", "Количество"])
    status_data_start_row = 6
    ws_charts.append(["Новый", status_counts['new']])
    ws_charts.append(["Принят", status_counts['accepted']])
    ws_charts.append(["Назначен", status_counts['assigned']])
    ws_charts.append(["Обучение завершено", status_counts['studies_completed']])
    ws_charts.append(["Завершён", status_counts['resolved']])
    ws_charts.append(["Отклонён", status_counts['declined']])
    
    # Столбчатая диаграмма: Статусы инцидентов
    bar_chart = BarChart()
    bar_chart.title = "Статусы инцидентов"
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.y_axis.title = "Количество"
    bar_chart.x_axis.title = "Статус"
    
    bar_labels = Reference(ws_charts, min_col=1, min_row=status_data_start_row, max_row=status_data_start_row + 5)
    bar_data = Reference(ws_charts, min_col=2, min_row=status_data_start_row - 1, max_row=status_data_start_row + 5)
    bar_chart.add_data(bar_data, titles_from_data=True)
    bar_chart.set_categories(bar_labels)
    bar_chart.shape = 4
    bar_chart.width = 16
    bar_chart.height = 10
    ws_charts.add_chart(bar_chart, "D14")
    
    # --- Данные для диаграммы: Назначения и завершения ---
    ws_charts.append([])
    ws_charts.append(["Метрика", "Значение"])
    metrics_start_row = 14
    ws_charts.append(["Всего назначений", total_assignments])
    ws_charts.append(["Уникальных назначений", total_unique_assignments])
    ws_charts.append(["Завершено обучений", total_completed])
    
    # Столбчатая диаграмма: Назначения
    bar_chart2 = BarChart()
    bar_chart2.title = "Назначения и завершения"
    bar_chart2.type = "col"
    bar_chart2.style = 12
    bar_chart2.y_axis.title = "Количество"
    
    bar2_labels = Reference(ws_charts, min_col=1, min_row=metrics_start_row, max_row=metrics_start_row + 2)
    bar2_data = Reference(ws_charts, min_col=2, min_row=metrics_start_row - 1, max_row=metrics_start_row + 2)
    bar_chart2.add_data(bar2_data, titles_from_data=True)
    bar_chart2.set_categories(bar2_labels)
    bar_chart2.width = 12
    bar_chart2.height = 8
    ws_charts.add_chart(bar_chart2, "D28")
    
    # Устанавливаем ширину колонок для данных
    _set_column_widths(ws_charts, [25, 15])
    
    # ================== ЛИСТ 3: Группировка по названию ==================
    ws_by_title = wb.create_sheet("По названиям")
    
    # Группируем инциденты по названию
    incidents_by_title = incidents.values('title').annotate(
        count=Count('id'),
        info_count=Count('id', filter=Q(incident_type='informational')),
        edu_count=Count('id', filter=Q(incident_type='educational'))
    ).order_by('-count')
    
    title_headers = ["Название инцидента", "Всего", "Информационных", "Обучающих"]
    ws_by_title.append(title_headers)
    _apply_header_style(ws_by_title, 1, len(title_headers))
    _set_column_widths(ws_by_title, [50, 12, 18, 15])
    
    for item in incidents_by_title:
        ws_by_title.append([
            item['title'], item['count'], item['info_count'], item['edu_count']
        ])
    
    # ================== ЛИСТ 4: Кто зафиксировал ==================
    ws_by_user = wb.create_sheet("Кто зафиксировал")
    
    # Группируем по пользователю, который создал инцидент
    incidents_by_user = incidents.values(
        'user__id', 'user__username', 'user__first_name', 'user__last_name'
    ).annotate(count=Count('id')).order_by('-count')
    
    user_headers = ["Пользователь", "Количество инцидентов"]
    ws_by_user.append(user_headers)
    _apply_header_style(ws_by_user, 1, len(user_headers))
    _set_column_widths(ws_by_user, [40, 25])
    
    for item in incidents_by_user:
        # Формируем имя пользователя
        full_name = f"{item['user__last_name'] or ''} {item['user__first_name'] or ''}".strip()
        display_name = full_name if full_name else item['user__username']
        ws_by_user.append([display_name, item['count']])
    
    # ================== ЛИСТ 5: Детализация по инцидентам ==================
    ws_details = wb.create_sheet("Детализация")
    
    details_headers = [
        "ID", "Название", "Тип", "Статус", "Зафиксировал",
        "Назначено (чел.)", "Нарушителей", "Курс-инцидент", "Дата создания"
    ]
    ws_details.append(details_headers)
    _apply_header_style(ws_details, 1, len(details_headers))
    _set_column_widths(ws_details, [8, 40, 18, 22, 30, 18, 15, 35, 20])
    
    for incident in incidents.select_related('user', 'course').prefetch_related('assigned_to', 'violators'):
        # Имя зафиксировавшего
        user = incident.user
        user_full_name = f"{user.last_name or ''} {user.first_name or ''}".strip()
        user_display = user_full_name if user_full_name else user.username
        
        # Тип и статус
        incident_type_display = "Информационный" if incident.incident_type == 'informational' else "Обучающий"
        status_display = dict(Incident.STATUS_CHOICES).get(incident.status, incident.status)
        
        ws_details.append([
            incident.id,
            incident.title,
            incident_type_display,
            status_display,
            user_display,
            incident.assigned_to.count(),
            incident.violators.count(),
            incident.course.title if incident.course else "—",
            incident.created_at.strftime('%Y-%m-%d %H:%M') if incident.created_at else "—"
        ])

    # ================== ЛИСТ 6: Моё подразделение ==================
    ws_my_dept = wb.create_sheet("Моё подразделение")

    # Пользователи из подразделения текущего пользователя
    my_department = None
    if hasattr(request.user, 'profile') and request.user.profile and request.user.profile.department:
        my_department = request.user.profile.department
        my_department_user_ids = set(
            User.objects.filter(profile__department=my_department).values_list('id', flat=True)
        )
    else:
        my_department_user_ids = set()

    # Первая таблица: назначения курсов-инцидентов по подразделению
    table1_headers = [
        "Дата назначения (курса-инцидента)", "ФИО (кому назначено)", "Подразделение",
        "Дедлайн", "Статус", "Название инцидента"
    ]
    ws_my_dept.append(table1_headers)
    _apply_header_style(ws_my_dept, 1, len(table1_headers))
    _set_column_widths(ws_my_dept, [28, 35, 30, 22, 22, 45])

    if my_department_user_ids:
        incident_courses = Course.objects.filter(is_incident=True)
        user_courses_list = (
            UserCourse.objects.filter(
                user_id__in=my_department_user_ids,
                course__in=incident_courses,
            )
            .select_related('user', 'course')
            .order_by('-start_date')
        )
        incidents_by_course = {
            inc.course_id: inc
            for inc in Incident.objects.filter(course_id__isnull=False).select_related('course')
        }
        for uc in user_courses_list:
            incident = incidents_by_course.get(uc.course_id)
            if not incident:
                continue
            user = uc.user
            fio = f"{user.last_name or ''} {user.first_name or ''}".strip() or user.username
            subdivision = user.profile.department.name if (hasattr(user, 'profile') and user.profile and user.profile.department) else "—"
            status_display = dict(Incident.STATUS_CHOICES).get(incident.status, incident.status)
            ws_my_dept.append([
                uc.start_date.strftime('%Y-%m-%d %H:%M') if uc.start_date else "—",
                fio,
                subdivision,
                uc.deadline.strftime('%Y-%m-%d %H:%M') if uc.deadline else "—",
                status_display,
                incident.title,
            ])

    # Пустая строка и вторая таблица: сводка по ФИО
    ws_my_dept.append([])
    table2_headers = [
        "ФИО",
        "Количество просроченных курсов-инцидентов",
        "Количество завершенных",
        "Количество назначенных",
        "Количество в статусе «Принят»/«Новый»",
        "Всего инцидентов",
    ]
    ws_my_dept.append(table2_headers)
    table2_header_row = ws_my_dept.max_row
    _apply_header_style(ws_my_dept, table2_header_row, len(table2_headers))
    # Высота строки в 2 раза и перенос по словам для заголовков второй таблицы
    default_height = 15
    ws_my_dept.row_dimensions[table2_header_row].height = default_height * 2
    for col_num in range(1, len(table2_headers) + 1):
        cell = ws_my_dept.cell(row=table2_header_row, column=col_num)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    _set_column_widths(ws_my_dept, [35, 30, 42, 28, 28, 42])

    if my_department_user_ids:
        incident_courses = Course.objects.filter(is_incident=True)
        now = timezone.now()
        # Всего инцидентов по пользователям (assigned_to + violators + expert), уникально
        total_incident_ids_by_user = {}
        for inc in Incident.objects.prefetch_related('assigned_to', 'violators').select_related('expert'):
            for u in list(inc.assigned_to.all()) + list(inc.violators.all()):
                total_incident_ids_by_user.setdefault(u.id, set()).add(inc.id)
            if inc.expert_id:
                total_incident_ids_by_user.setdefault(inc.expert_id, set()).add(inc.id)
        total_incidents_by_user = {uid: len(s) for uid, s in total_incident_ids_by_user.items()}

        # Количество инцидентов в статусе Принят/Новый по пользователям (assigned_to + violators)
        incidents_prefetch = Incident.objects.prefetch_related('assigned_to', 'violators').filter(
            status__in=['accepted', 'new']
        )
        accepted_new_by_user = {}
        for inc in incidents_prefetch:
            for u in list(inc.assigned_to.all()) + list(inc.violators.all()):
                accepted_new_by_user[u.id] = accepted_new_by_user.get(u.id, 0) + 1

        for user in User.objects.filter(id__in=my_department_user_ids).prefetch_related('profile').order_by('last_name', 'first_name'):
            fio = f"{user.last_name or ''} {user.first_name or ''}".strip() or user.username
            total_incidents = total_incidents_by_user.get(user.id, 0)
            ucs = list(
                UserCourse.objects.filter(
                    user=user,
                    course__in=incident_courses,
                ).select_related('course')
            )
            assigned_count = len(ucs)
            completed_count = sum(1 for uc in ucs if uc.status == 'completed')
            overdue_count = sum(
                1 for uc in ucs
                if uc.status != 'completed' and uc.deadline and uc.deadline < now
            )
            accepted_new_count = accepted_new_by_user.get(user.id, 0)
            if total_incidents > 0 or assigned_count > 0 or accepted_new_count > 0:
                ws_my_dept.append([
                    fio,
                    overdue_count,
                    completed_count,
                    assigned_count,
                    accepted_new_count,
                    total_incidents,
                ])

    # Сохраняем файл
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"incidents_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    
    audit_logger.info(
        f'Экспортировал отчет по инцидентам в Excel', 
        extra={
            'user': request.user.email if request.user.is_authenticated else 'Anonymous',
            'target_user': request.user.email
        }
    )
    return response



class IncidentCreateView(CreateView, AuditLoggerMixin):
    """
    Создание инцидента (ручное или автоматическое).
    """
    model = Incident
    form_class = IncidentForm
    template_name = 'builder/incidents/incident_form.html'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            return render(request, '403.html', status=403)
        return super().dispatch(request, *args, **kwargs)


    def get_success_url(self):
        """
        Возвращает URL для редиректа после успешного создания инцидента.
        Перенаправляет на страницу редактирования созданного инцидента.
        """
        return reverse('builder:incident_edit', kwargs={'pk': self.object.pk})


    def form_valid(self, form):
        # Устанавливаем статус "Принят" для нового инцидента
        form.instance.status = 'accepted'
        response = super().form_valid(form)
        # Логируем создание инцидента
        self.log_create_action(self.object, "Создан новый инцидент")
        return response




class IncidentUpdateView(UpdateView, AuditLoggerMixin):
    """
    Редактирование инцидента.
    """
    model = Incident
    form_class = IncidentForm
    template_name = 'builder/incidents/incident_form.html'
    success_url = reverse_lazy('builder:incidents')

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            return render(request, '403.html', status=403)
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        # Сохраняем старый список назначенных пользователей до сохранения формы
        old_assigned_users = set(self.object.assigned_to.all())
        
        # Получаем новый список назначенных пользователей из формы (до сохранения)
        new_assigned_users = set(form.cleaned_data.get('assigned_to', []))
        
        # Определяем, какие пользователи были удалены и добавлены
        removed_users = old_assigned_users - new_assigned_users
        added_users = new_assigned_users - old_assigned_users
        
        # Сохраняем форму
        response = super().form_valid(form)
        
        # Если у инцидента есть связанный курс, обновляем назначения курса
        if self.object.course:
            course = self.object.course
            
            # Удаляем назначение курса для пользователей, которые были удалены из списка назначенных
            for user in removed_users:
                UserCourse.objects.filter(user=user, course=course).delete()
            
            # Назначаем курс новым пользователям
            # Определяем дедлайн:
            # 1) приоритет у incident.assigned_to_time_to_complete (если задано и > 0),
            # 2) затем используем course.default_deadline_days (если задано и > 0),
            # 3) иначе берём значение по умолчанию (3 дня).
            time_to_complete = self.object.assigned_to_time_to_complete
            if not time_to_complete or time_to_complete <= 0:
                if course.default_deadline_days and course.default_deadline_days > 0:
                    time_to_complete = course.default_deadline_days
                else:
                    time_to_complete = 3
            deadline = timezone.now() + timedelta(days=time_to_complete)
            
            for user in added_users:
                UserCourse.objects.get_or_create(
                    user=user,
                    course=course,
                    defaults={'status': 'available', 'deadline': deadline}
                )
        
        # Логируем обновление инцидента
        self.log_update_action(self.object, "Инцидент обновлён")
        return response




@method_decorator(login_required, name='dispatch')
class IncidentDeclineView(View, AuditLoggerMixin):
    """
    Отклонение или возобновление инцидента.
    """
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            return render(request, '403.html', status=403)
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        incident_id = kwargs.get('pk')
        incident = get_object_or_404(Incident, pk=incident_id)
        
        # Сохраняем старые значения для аудита
        old_values = serialize_model_data(incident)
        
        if incident.status == 'declined':
            # Возобновляем инцидент - возвращаем предыдущий статус
            if incident.previous_status:
                incident.status = incident.previous_status
                incident.previous_status = None
                comment = f"Инцидент возобновлён. Статус изменён на '{incident.get_status_display()}'"
            else:
                # Если предыдущий статус не сохранён, устанавливаем 'new'
                incident.status = 'new'
                incident.previous_status = None
                comment = "Инцидент возобновлён. Статус изменён на 'Новый'"
            
            # Если у инцидента есть связанный курс-инцидент, назначаем его снова всем пользователям из assigned_to
            if incident.course:
                course = incident.course
                # Получаем всех пользователей, назначенных на инцидент
                assigned_users = incident.assigned_to.all()
                
                # Определяем дедлайн:
                # 1) приоритет у incident.assigned_to_time_to_complete (если задано и > 0),
                # 2) затем используем course.default_deadline_days (если задано и > 0),
                # 3) иначе берём значение по умолчанию (3 дня).
                time_to_complete = incident.assigned_to_time_to_complete
                if not time_to_complete or time_to_complete <= 0:
                    if course.default_deadline_days and course.default_deadline_days > 0:
                        time_to_complete = course.default_deadline_days
                    else:
                        time_to_complete = 3
                deadline = timezone.now() + timedelta(days=time_to_complete)
                
                for user in assigned_users:
                    # Удаляем запись из ManualCourseUnassignment, если она есть
                    ManualCourseUnassignment.objects.filter(
                        user=user,
                        course=course
                    ).delete()
                    
                    # Назначаем курс пользователю
                    UserCourse.objects.get_or_create(
                        user=user,
                        course=course,
                        defaults={'status': 'available', 'deadline': deadline}
                    )
        else:
            # Отклоняем инцидент - сохраняем текущий статус и устанавливаем 'declined'
            previous_status_display = dict(Incident.STATUS_CHOICES).get(incident.status, incident.status)
            incident.previous_status = incident.status
            incident.status = 'declined'
            comment = f"Инцидент отклонён. Предыдущий статус: '{previous_status_display}'"

            if incident.course:
                user_courses = list(UserCourse.objects.filter(course=incident.course))
                for user_course in user_courses:
                    ManualCourseUnassignment.objects.get_or_create(
                        user=user_course.user,
                        course=incident.course,
                        defaults={
                            'unassigned_by': request.user,
                            'reason': f'Инцидент "{incident.title}" переведен в статус "Отклонен".'
                        }
                    )
                    user_course.delete()

        incident.save(update_fields=['status', 'previous_status', 'updated_at'])
        
        # Логируем действие
        self.log_update_action(incident, old_values, comment)
        
        return redirect('builder:incidents')


@method_decorator(login_required, name='dispatch')
class IncidentDeleteView(View, AuditLoggerMixin):
    """
    Полное удаление инцидента: запись и связанные назначения на курс-инцидент.
    """
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            return render(request, '403.html', status=403)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        incident = get_object_or_404(Incident, pk=kwargs.get('pk'))
        title = incident.title
        course = incident.course

        self.log_delete_action(incident, f'Инцидент удалён: «{title}»')

        if course:
            incident.assigned_to.clear()
            incident.violators.clear()
            ManualCourseUnassignment.objects.filter(course=course).delete()
            UserCourse.objects.filter(course=course).delete()

        incident.delete()
        messages.success(request, f'Инцидент «{title}» удалён')
        return redirect('builder:incidents')


@method_decorator(login_required, name='dispatch')
class CreateCourseFromIncidentView(View):
    """
    Создание курса-инцидента из инцидента.
    """
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
            return render(request, '403.html', status=403)
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        incident_id = kwargs.get('pk')
        try:
            incident = get_object_or_404(Incident, pk=incident_id)
            
            # Проверяем, не создан ли уже курс для этого инцидента
            if incident.course:
                return redirect('courses:course_detail', slug=incident.course.slug)
            
            # Создаем курс с названием инцидента
            course = Course.objects.create(
                title=incident.title,
                description='', # Не выводить описание инцидента в описание курса-инцидента
                author=request.user,
                is_incident=True,
                responsible_mentor=incident.responsible_mentor,
                mentors_time_to_check=incident.mentors_time_to_check or 2
            )
            
            # Связываем инцидент с курсом
            # Статус остается 'accepted', так как курс еще не назначен сотрудникам
            incident.course = course
            # Если статус был 'new', меняем на 'accepted'
            if incident.status == 'new':
                incident.status = 'accepted'
            incident.save(update_fields=['course', 'status', 'updated_at'])
            
            # Автоназначение курса отключено - назначение происходит вручную через кнопки в деталке курса
            
            # Перенаправляем на страницу курса
            return redirect('courses:course_detail', slug=course.slug)
        except Exception as e:
            # В случае ошибки перенаправляем обратно на форму инцидента с сообщением об ошибке
            from django.contrib import messages
            import traceback
            messages.error(request, f'Ошибка при создании курса: {str(e)}')
            if incident_id:
                return redirect('builder:incident_edit', pk=incident_id)
            return redirect('builder:incidents')




class IncidentDetailListView(ListView):
    """
    Список по прогрессу пользователей по всем инцидентам.
    Ответ кэшируется по пользователю и полному URL (включая GET-параметры).
    """
    model = Incident
    template_name = 'builder/incidents/incident_detail.html'
    context_object_name = 'incidents'
    ordering = ['-created_at']


    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser or request.user.profile.is_mentor_user):
            return render(request, '403.html', status=403)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        from django.utils import timezone
        import datetime
        
        queryset = super().get_queryset()
        
        # Оптимизация: предзагрузка ManyToMany полей и связанных объектов
        queryset = queryset.prefetch_related('assigned_to', 'violators').select_related('user', 'responsible_mentor', 'expert', 'course')
        
        # Фильтр по названию инцидента (поиск)
        search = self.request.GET.get('search', '').strip()
        if search:
            queryset = queryset.filter(title__icontains=search)
        
        # Фильтр по дате создания
        date_from = self.request.GET.get('date_from')
        date_to = self.request.GET.get('date_to')
        
        # Если нет параметров в GET запросе (первичная загрузка), устанавливаем дефолтные значения
        if not self.request.GET:
            # Устанавливаем период с начала 2025 года до сегодняшней даты
            date_from = '2025-01-01'
            date_to = timezone.now().date().strftime('%Y-%m-%d')
        
        if date_from:
            date_from_parsed = datetime.datetime.strptime(date_from, '%Y-%m-%d').date()
            date_from_datetime = timezone.make_aware(datetime.datetime.combine(date_from_parsed, datetime.time.min))
            queryset = queryset.filter(created_at__gte=date_from_datetime)
        if date_to:
            date_to_parsed = datetime.datetime.strptime(date_to, '%Y-%m-%d').date()
            date_to_datetime = timezone.make_aware(datetime.datetime.combine(date_to_parsed, datetime.time.max))
            queryset = queryset.filter(created_at__lte=date_to_datetime)
        
        return queryset

    def get_context_data(self, **kwargs):
        from django.utils import timezone
        from django.contrib.auth import get_user_model
        from myapp.models import QuizResult
        
        context = super().get_context_data(**kwargs)
        context['now'] = timezone.now()  # Текущая дата и время для проверки просроченных дедлайнов
        is_mentor_only = (
            self.request.user.profile.is_mentor_user
            and not self.request.user.is_staff
            and not self.request.user.is_superuser
        )
        mentor_department = self.request.user.profile.department if is_mentor_only else None
        mentor_department_id = mentor_department.id if mentor_department else None
        mentor_department_name = mentor_department.name if mentor_department else None
        context['show_department_filter'] = not is_mentor_only
        
        # Получаем список всех активных пользователей для фильтра
        User = get_user_model()
        users_queryset = User.objects.filter(is_active=True)
        if mentor_department_id is not None:
            users_queryset = users_queryset.filter(profile__department_id=mentor_department_id)
        context['users'] = users_queryset.order_by('last_name', 'first_name')
        
        # Параметры фильтров
        search = self.request.GET.get('search', '').strip()
        selected_user_id = self.request.GET.get('assigned_user', '')
        violator_filter = self.request.GET.get('violator_filter', 'all')  # 'all', 'yes', 'no'
        
        # Фильтр по статусу курса (UserCourse.status) + виртуальный статус "Обучение завершено"
        status_choices = list(UserCourse.STATUS_CHOICES) + [('studies_completed', 'Обучение завершено')]
        context['status_choices'] = status_choices

        departments = Department.objects.all().order_by('name')
        if mentor_department_id is not None:
            departments = departments.filter(id=mentor_department_id)
        context['departments'] = departments
        
        # Если нет параметров в GET запросе (первичная загрузка), устанавливаем дефолтные значения
        if not self.request.GET:
            context['date_from'] = '2025-01-01'
            context['date_to'] = timezone.now().date().strftime('%Y-%m-%d')
            context['search'] = ''
            context['selected_statuses'] = []
            context['selected_user_id'] = None
            context['violator_filter'] = 'all'
            context['violator_filter_locked'] = False
            context['department_filter'] = mentor_department_name or ''
            context['selected_department_filters'] = [mentor_department_name] if mentor_department_name else []
            context['only_overdue'] = False
        else:
            date_from = self.request.GET.get('date_from', '')
            date_to = self.request.GET.get('date_to', '')
            if mentor_department_name:
                selected_department_filters = [mentor_department_name]
                department_filter = mentor_department_name
            else:
                selected_department_filters = self.request.GET.getlist('department_filter')
                department_filter = self.request.GET.get('department_filter', '')  # для обратной совместимости
            only_overdue = self.request.GET.get('only_overdue', '') == 'on'

            
            # Если violator_filter=yes и даты не указаны, устанавливаем последние 30 дней
            if violator_filter == 'yes' and not date_from and not date_to:
                import datetime
                today = timezone.now().date()
                date_from = (today - datetime.timedelta(days=30)).strftime('%Y-%m-%d')
                date_to = today.strftime('%Y-%m-%d')
            
            context['date_from'] = date_from
            context['date_to'] = date_to
            context['search'] = search
            # Статусы, выбранные в фильтре (чекбоксы)
            context['selected_statuses'] = self.request.GET.getlist('status', [])
            context['department_filter'] = department_filter
            context['selected_department_filters'] = selected_department_filters
            context['only_overdue'] = only_overdue
            
            try:
                context['selected_user_id'] = int(selected_user_id) if selected_user_id else None
            except (ValueError, TypeError):
                context['selected_user_id'] = None
            context['violator_filter'] = violator_filter
            # Блокируем фильтр по нарушителям, если он установлен в 'yes' (переход с кнопки "Нарушители")
            context['violator_filter_locked'] = (violator_filter == 'yes')
        
        incident_user_list, has_more_items = self._get_incident_user_slice(limit=50, offset=0)
        context['incident_user_list'] = incident_user_list
        context['has_more_incident_users'] = has_more_items
        context['incident_users_next_offset'] = len(incident_user_list)

        context['is_admin'] = is_admin(self.request.user)
        return context

    def _get_incident_user_slice(self, limit=50, offset=0):
        from myapp.models import QuizResult

        try:
            limit = int(limit)
        except (TypeError, ValueError):
            limit = 50
        try:
            offset = int(offset)
        except (TypeError, ValueError):
            offset = 0
        limit = max(1, limit)
        offset = max(0, offset)

        incidents = self.get_queryset()
        now = timezone.now()
        is_mentor_only = (
            self.request.user.profile.is_mentor_user
            and not self.request.user.is_staff
            and not self.request.user.is_superuser
        )
        mentor_department = self.request.user.profile.department if is_mentor_only else None
        mentor_department_id = mentor_department.id if mentor_department else None

        selected_user_id = self.request.GET.get('assigned_user', '')
        try:
            selected_user_id = int(selected_user_id) if selected_user_id else None
        except (TypeError, ValueError):
            selected_user_id = None

        violator_filter = self.request.GET.get('violator_filter', 'all')
        selected_department_filters = self.request.GET.getlist('department_filter')
        if mentor_department_id is not None and mentor_department is not None:
            selected_department_filters = [mentor_department.name]
        only_overdue = self.request.GET.get('only_overdue', '') == 'on'
        selected_course_statuses = self.request.GET.getlist('status')

        max_items_to_collect = offset + limit + 1
        collected = []
        stop_collecting = False

        for incident in incidents:
            if stop_collecting:
                break
            assigned_users = incident.assigned_to.all()
            violators = incident.violators.all()

            for user in assigned_users:
                if mentor_department_id is not None:
                    user_department_id = getattr(getattr(user, 'profile', None), 'department_id', None)
                    if user_department_id != mentor_department_id:
                        continue

                if selected_department_filters:
                    user_department_name = None
                    if hasattr(user, 'profile') and user.profile and user.profile.department:
                        user_department_name = user.profile.department.name
                    if user_department_name not in selected_department_filters:
                        continue

                if selected_user_id and user.id != selected_user_id:
                    continue

                is_violator = user in violators
                if violator_filter == 'yes' and not is_violator:
                    continue
                if violator_filter == 'no' and is_violator:
                    continue

                if only_overdue and not incident.course:
                    continue

                if incident.course:
                    user_course_qs = UserCourse.objects.filter(user=user, course=incident.course)
                    if not user_course_qs.exists():
                        continue
                elif selected_course_statuses:
                    # Если выбран фильтр статусов курса, инциденты без курса не показываем.
                    continue

                progress_percent = None
                course_deadline = None
                course_status = None
                course_status_display = None

                if incident.course:
                    course = incident.course
                    user_course = UserCourse.objects.filter(user=user, course=course).first()
                    if user_course:
                        course_deadline = user_course.deadline
                        course_status = user_course.status
                        course_status_display = user_course.get_status_display()

                        # Виртуальный статус "Обучение завершено" для ожидания проверки наставником.
                        has_pending_quiz_review = QuizResult.objects.filter(
                            user=user,
                            course=course,
                            status='pending'
                        ).exists()
                        has_pending_homework_review = HomeworkSubmission.objects.filter(
                            user=user,
                            course=course,
                            status='pending'
                        ).exists()
                        if (has_pending_quiz_review or has_pending_homework_review) and course_status != 'completed':
                            course_status = 'studies_completed'
                            course_status_display = 'Обучение завершено'

                    if selected_course_statuses and course_status not in selected_course_statuses:
                        continue

                    if only_overdue:
                        if not course_deadline or course_deadline >= now or course_status == 'completed' or incident.status == 'declined':
                            continue

                    trajectory = UserLessonTrajectory.objects.filter(user=user, course=course).first()
                    if trajectory:
                        lessons = trajectory.lessons.all().order_by('order')
                        total_lessons = lessons.count()
                        lesson_ids = lessons.values_list('id', flat=True)
                        completed_lessons = UserProgress.objects.filter(
                            user=user, course=course, completed=True, lesson_id__in=lesson_ids
                        ).count()
                    else:
                        lessons = course.lessons.all().order_by('order')
                        total_lessons = lessons.count()
                        completed_lessons = UserProgress.objects.filter(
                            user=user, course=course, completed=True
                        ).count()

                    completed_quizzes = QuizResult.objects.filter(
                        user=user,
                        course=course,
                        quiz_title__in=[quiz.name for quiz in course.quizzes.all()],
                        passed=True
                    ).values('quiz_title').distinct().count()
                    total_quizzes = course.quizzes.count()
                    total_materials = total_lessons + total_quizzes
                    completed_materials = completed_lessons + completed_quizzes
                    progress_percent = int((completed_materials / total_materials) * 100) if total_materials > 0 else 0

                collected.append({
                    'incident': incident,
                    'user': user,
                    'is_violator': is_violator,
                    'is_expert': False,
                    'progress_percent': progress_percent,
                    'course_deadline': course_deadline,
                    'incident_status': incident.status,
                    'course_status': course_status,
                    'course_status_display': course_status_display,
                    'incident_status_display': incident.get_status_display(),
                })
                if len(collected) >= max_items_to_collect:
                    stop_collecting = True
                    break

            if stop_collecting:
                continue

            if incident.expert:
                expert = incident.expert
                if expert not in assigned_users:
                    should_add_expert = True
                    if mentor_department_id is not None:
                        expert_department_id = getattr(getattr(expert, 'profile', None), 'department_id', None)
                        if expert_department_id != mentor_department_id:
                            should_add_expert = False
                    if selected_department_filters:
                        expert_department_name = None
                        if hasattr(expert, 'profile') and expert.profile and expert.profile.department:
                            expert_department_name = expert.profile.department.name
                        if expert_department_name not in selected_department_filters:
                            should_add_expert = False
                    if selected_user_id and expert.id != selected_user_id:
                        should_add_expert = False
                    if violator_filter == 'yes':
                        should_add_expert = False
                    if only_overdue and not incident.course:
                        should_add_expert = False
                    if incident.course:
                        expert_course_qs = UserCourse.objects.filter(user=expert, course=incident.course)
                        if not expert_course_qs.exists():
                            should_add_expert = False
                    elif selected_course_statuses:
                        should_add_expert = False

                    if should_add_expert:
                        progress_percent = None
                        course_deadline = None
                        course_status = None
                        course_status_display = None
                        if incident.course:
                            course = incident.course
                            user_course = UserCourse.objects.filter(user=expert, course=course).first()
                            if user_course:
                                course_deadline = user_course.deadline
                                course_status = user_course.status
                                course_status_display = user_course.get_status_display()

                                has_pending_quiz_review = QuizResult.objects.filter(
                                    user=expert,
                                    course=course,
                                    status='pending'
                                ).exists()
                                has_pending_homework_review = HomeworkSubmission.objects.filter(
                                    user=expert,
                                    course=course,
                                    status='pending'
                                ).exists()
                                if (has_pending_quiz_review or has_pending_homework_review) and course_status != 'completed':
                                    course_status = 'studies_completed'
                                    course_status_display = 'Обучение завершено'

                            if selected_course_statuses and course_status not in selected_course_statuses:
                                should_add_expert = False
                            if only_overdue:
                                if not course_deadline or course_deadline >= now or course_status == 'completed' or incident.status == 'declined':
                                    should_add_expert = False
                            trajectory = UserLessonTrajectory.objects.filter(user=expert, course=course).first()
                            if trajectory:
                                lessons = trajectory.lessons.all().order_by('order')
                                total_lessons = lessons.count()
                                lesson_ids = lessons.values_list('id', flat=True)
                                completed_lessons = UserProgress.objects.filter(
                                    user=expert, course=course, completed=True, lesson_id__in=lesson_ids
                                ).count()
                            else:
                                lessons = course.lessons.all().order_by('order')
                                total_lessons = lessons.count()
                                completed_lessons = UserProgress.objects.filter(
                                    user=expert, course=course, completed=True
                                ).count()
                            completed_quizzes = QuizResult.objects.filter(
                                user=expert,
                                course=course,
                                quiz_title__in=[quiz.name for quiz in course.quizzes.all()],
                                passed=True
                            ).values('quiz_title').distinct().count()
                            total_quizzes = course.quizzes.count()
                            total_materials = total_lessons + total_quizzes
                            completed_materials = completed_lessons + completed_quizzes
                            progress_percent = int((completed_materials / total_materials) * 100) if total_materials > 0 else 0

                        if should_add_expert:
                            collected.append({
                                'incident': incident,
                                'user': expert,
                                'is_violator': False,
                                'is_expert': True,
                                'progress_percent': progress_percent,
                                'course_deadline': course_deadline,
                                'incident_status': incident.status,
                                'course_status': course_status,
                                'course_status_display': course_status_display,
                                'incident_status_display': incident.get_status_display(),
                            })
                            if len(collected) >= max_items_to_collect:
                                stop_collecting = True
                                break

        has_more = len(collected) > (offset + limit)
        return collected[offset:offset + limit], has_more




@method_decorator(login_required, name='dispatch')
class IncidentDetailLoadMoreView(IncidentDetailListView):
    def get(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser or request.user.profile.is_mentor_user):
            return JsonResponse({'error': 'forbidden'}, status=403)

        try:
            offset = int(request.GET.get('offset', 0))
        except (TypeError, ValueError):
            offset = 0
        try:
            limit = int(request.GET.get('limit', 50))
        except (TypeError, ValueError):
            limit = 50
        offset = max(0, offset)
        limit = max(1, limit)

        incident_user_list, has_more = self._get_incident_user_slice(limit=limit, offset=offset)
        rows_html = render_to_string(
            'builder/incidents/_incidents_detail_rows.html',
            {
                'incident_user_list': incident_user_list,
                'row_start_index': offset,
                'is_admin': is_admin(request.user),
            },
            request=request
        )
        return JsonResponse({
            'rows_html': rows_html,
            'loaded_count': len(incident_user_list),
            'next_offset': offset + len(incident_user_list),
            'has_more': has_more,
        })


@method_decorator(login_required, name='dispatch')
class UnassignIncidentUserView(View, AuditLoggerMixin):
    """
    Отмена назначения пользователя на инцидент.
    Удаляет пользователя из поля assigned_to инцидента.
    """
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser or request.user.profile.is_mentor_user):
            return render(request, '403.html', status=403)
        return super().dispatch(request, *args, **kwargs)
    
    def post(self, request, *args, **kwargs):
        from django.contrib.auth import get_user_model
        from django.contrib import messages
        
        User = get_user_model()
        incident_id = kwargs.get('incident_id')
        user_id = kwargs.get('user_id')
        
        incident = get_object_or_404(Incident, pk=incident_id)
        user = get_object_or_404(User, pk=user_id)
        
        # Проверяем, что пользователь действительно назначен на инцидент
        if user not in incident.assigned_to.all():
            messages.error(request, f'Пользователь {user.get_full_name() or user.username} не назначен на инцидент "{incident.title}"')
            return redirect('builder:incident_detail')
        
        # Сохраняем старые значения для аудита
        old_values = serialize_model_data(incident)
        
        # Удаляем пользователя из assigned_to
        # Сигнал автоматически удалит доступ к курсу, если он есть
        incident.assigned_to.remove(user)
        
        # Логируем действие
        comment = f"Отменено назначение пользователя {user.get_full_name() or user.username} на инцидент"
        self.log_update_action(incident, old_values, comment)
        
        messages.success(request, f'Назначение пользователя {user.get_full_name() or user.username} на инцидент "{incident.title}" отменено')
        
        # Перенаправляем обратно на страницу деталей с сохранением фильтров
        redirect_url = reverse('builder:incident_detail')
        # Получаем параметры фильтров из POST (они передаются как скрытые поля формы)
        # или из GET (если они есть)
        from urllib.parse import urlencode
        query_params = []
        for key in ['search', 'date_from', 'date_to', 'assigned_user', 'violator_filter']:
            value = request.POST.get(key) or request.GET.get(key)
            if value:
                query_params.append((key, value))
        
        if query_params:
            redirect_url += '?' + urlencode(query_params)
        
        return redirect(redirect_url)


@method_decorator(login_required, name='dispatch')
class BulkUnassignIncidentUsersView(View, AuditLoggerMixin):
    """
    Массовая отмена назначений: удаляет пользователей из assigned_to по списку пар (инцидент, пользователь).
    """
    MAX_ASSIGNMENTS = 300

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (
            request.user.is_staff or request.user.is_superuser or request.user.profile.is_mentor_user
        ):
            return render(request, '403.html', status=403)
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        from urllib.parse import urlencode

        redirect_url = reverse('builder:incident_detail')
        query_params = []
        for key in ['search', 'date_from', 'date_to', 'assigned_user', 'violator_filter']:
            value = request.POST.get(key) or request.GET.get(key)
            if value:
                query_params.append((key, value))
        if query_params:
            redirect_url += '?' + urlencode(query_params)

        if not is_admin(request.user):
            messages.error(request, 'Недостаточно прав для этой операции')
            return redirect(redirect_url)

        raw = request.POST.get('assignments', '')
        try:
            pairs = json.loads(raw) if raw else []
        except json.JSONDecodeError:
            messages.error(request, 'Некорректный формат списка назначений')
            return redirect(redirect_url)

        if not isinstance(pairs, list) or not pairs:
            messages.warning(request, 'Не выбрано ни одной строки для отмены назначения')
            return redirect(redirect_url)

        normalized = []
        seen = set()
        for item in pairs[: self.MAX_ASSIGNMENTS]:
            if not isinstance(item, dict):
                continue
            try:
                incident_id = int(item.get('incident_id'))
                user_id = int(item.get('user_id'))
            except (TypeError, ValueError):
                continue
            if incident_id <= 0 or user_id <= 0:
                continue
            key = (incident_id, user_id)
            if key in seen:
                continue
            seen.add(key)
            normalized.append((incident_id, user_id))

        if not normalized:
            messages.warning(request, 'Не удалось обработать выбранные строки')
            return redirect(redirect_url)

        success_count = 0
        skipped_count = 0

        for incident_id, user_id in normalized:
            incident = Incident.objects.filter(pk=incident_id).first()
            user = User.objects.filter(pk=user_id).first()
            if not incident or not user:
                skipped_count += 1
                continue
            if user not in incident.assigned_to.all():
                skipped_count += 1
                continue

            old_values = serialize_model_data(incident)
            incident.assigned_to.remove(user)
            comment = (
                f'Массовая отмена: пользователь {user.get_full_name() or user.username} снят с инцидента'
            )
            self.log_update_action(incident, old_values, comment)
            success_count += 1

        if success_count:
            messages.success(
                request,
                f'Отменено назначений: {success_count}.',
            )
        if skipped_count:
            messages.info(
                request,
                f'Пропущено строк (не назначены или не найдены): {skipped_count}.',
            )
        if not success_count and not skipped_count:
            messages.error(request, 'Не удалось отменить назначения')

        return redirect(redirect_url)


@method_decorator(login_required, name='dispatch')
class IncidentStatusesReportView(ListView):
    """
    Отчет за последнюю неделю по инцидентам.
    Показывает статистику по каждому пользователю:
    - ФИО
    - Подразделение
    - Назначено (количество инцидентов со статусом 'assigned')
    - Просрочено (количество инцидентов с просроченным дедлайном курса)
    - Завершено (количество инцидентов со статусом 'resolved')
    - Обучение завершено (количество инцидентов со статусом 'studies_completed')
    """
    template_name = 'builder/incidents/incident_statuses_report.html'
    context_object_name = 'report_data'

    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser or request.user.profile.is_mentor_user):
            return render(request, '403.html', status=403)
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        from django.contrib.auth import get_user_model
        from django.db.models import Q
        import datetime as dt
        
        User = get_user_model()
        now = timezone.now()
        
        # Получаем параметры дат из GET запроса
        date_from_str = self.request.GET.get('date_from')
        date_to_str = self.request.GET.get('date_to')
        department_filter = self.request.GET.get('department_filter')


        # Если даты не указаны, устанавливаем диапазон с начала месяца
        if not date_from_str or not date_to_str:
            today = now.date()
            # Начало текущего месяца
            month_start = dt.date(today.year, today.month, 1)
            date_from_str = month_start.strftime('%Y-%m-%d')
            date_to_str = today.strftime('%Y-%m-%d')
        
        # Преобразуем строки в datetime объекты
        date_from = timezone.make_aware(dt.datetime.combine(
            dt.datetime.strptime(date_from_str, '%Y-%m-%d').date(),
            dt.time.min
        ))
        date_to = timezone.make_aware(dt.datetime.combine(
            dt.datetime.strptime(date_to_str, '%Y-%m-%d').date(),
            dt.time.max
        ))
        
        # Получаем все инциденты за выбранный период
        incidents = Incident.objects.filter(
            created_at__gte=date_from,
            created_at__lte=date_to
        ).prefetch_related('assigned_to', 'violators', 'course').select_related('course')
        
        # Собираем уникальных пользователей, которые были назначены на инциденты
        users_with_incidents = set()
        for incident in incidents:
            users_with_incidents.update(incident.assigned_to.all())
            users_with_incidents.update(incident.violators.all())
            if incident.expert:
                users_with_incidents.add(incident.expert)
        
        # Формируем отчет для каждого пользователя
        report_data = []
        for user in users_with_incidents:
            # Получаем профиль пользователя
            if not hasattr(user, 'profile') or not user.profile:
                continue
            
            # Применяем фильтр по подразделению, если он указан
            if department_filter:
                user_department = user.profile.department.name if user.profile.department else '—'
                if user_department != department_filter:
                    continue
            
            # Фильтруем инциденты для этого пользователя
            user_incidents = incidents.filter(
                Q(assigned_to=user) | Q(violators=user) | Q(expert=user)
            ).distinct()
            
            # Подсчитываем статистику
            assigned_count = user_incidents.filter(status='assigned').count()
            resolved_count = user_incidents.filter(status='resolved').count()
            studies_completed_count = user_incidents.filter(status='studies_completed').count()
            
            # Подсчитываем просроченные инциденты
            overdue_count = 0
            for incident in user_incidents:
                if incident.course:
                    # Проверяем, есть ли у пользователя UserCourse для этого курса
                    user_course = UserCourse.objects.filter(
                        user=user,
                        course=incident.course
                    ).first()
                    
                    if user_course and user_course.deadline:
                        # Проверяем, просрочен ли дедлайн и не завершен ли курс
                        if user_course.deadline < now and user_course.status != 'completed':
                            overdue_count += 1
            
            report_data.append({
                'user': user,
                'full_name': user.get_full_name() or user.username,
                'department': user.profile.department.name if user.profile.department else '—',
                'assigned_count': assigned_count,
                'overdue_count': overdue_count,
                'resolved_count': resolved_count,
                'studies_completed_count': studies_completed_count,
            })
        
        # Сортируем по ФИО
        report_data.sort(key=lambda x: x['full_name'])
        
        return report_data

    def get_context_data(self, **kwargs):
        import datetime as dt
        
        context = super().get_context_data(**kwargs)
        now = timezone.now()
        
        # Получаем параметры дат из GET запроса
        date_from_str = self.request.GET.get('date_from')
        date_to_str = self.request.GET.get('date_to')
        department_filter = self.request.GET.get('department_filter', '')
        
        # Если даты не указаны, устанавливаем диапазон с начала месяца
        if not date_from_str or not date_to_str:
            today = now.date()
            # Начало текущего месяца
            month_start = dt.date(today.year, today.month, 1)
            date_from_str = month_start.strftime('%Y-%m-%d')
            date_to_str = today.strftime('%Y-%m-%d')
        
        # Получаем список всех подразделений для выпадающего списка
        departments = Department.objects.all().order_by('name')
        
        context['date_from'] = date_from_str
        context['date_to'] = date_to_str
        context['week_start'] = dt.datetime.strptime(date_from_str, '%Y-%m-%d').date()
        context['week_end'] = dt.datetime.strptime(date_to_str, '%Y-%m-%d').date()
        context['departments'] = departments
        context['department_filter'] = department_filter
        return context
